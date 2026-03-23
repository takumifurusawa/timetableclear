"""
================================================================================
小学校時間割自動生成アプリ  timetable_app.py
================================================================================

【実行方法】
1. ライブラリをインストール:
   pip install streamlit openpyxl ortools pandas

2. アプリを起動:
   streamlit run timetable_app.py

3. ブラウザで http://localhost:8501 を開く

【動作環境】
  Python 3.10 以上 / streamlit >= 1.28 / openpyxl >= 3.1
  ortools >= 9.0 / pandas >= 1.5

================================================================================
"""

# ============================================================
# インポート
# ============================================================
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io
import pandas as pd
from datetime import datetime
import time
import copy

# ============================================================
# ① 固定定数（変更禁止）
# ============================================================

SUBJECTS = [
    "国語", "算数", "理科", "社会", "外国語",
    "体育", "図工", "音楽", "家庭", "道徳",
    "総合", "学活", "生活", "その他1", "その他2", "その他3"
]

DAYS   = ["月", "火", "水", "木", "金"]
GRADES = [1, 2, 3, 4, 5, 6]

COMMON_SUBJECTS = ["体育", "図工", "音楽", "家庭", "外国語"]

COLOR_HEADER = "D9D9D9"
COLOR_MANUAL = "CCE5FF"
COLOR_AUTO   = "CCFFCC"
COLOR_SG     = "FFE5CC"
COLOR_ABSENT = "C0C0C0"

STEP_LABELS = {
    1: "STEP 1: クラス数設定",
    2: "STEP 2: 教員登録",
    3: "STEP 3: コマ数設定",
    4: "STEP 4: 週コマ数設定",
    5: "STEP 5: 担当教員割り当て",
    6: "STEP 6: 不在コマ設定",
    7: "STEP 7: 特別教室設定",
    8: "STEP 8: 手動時間割入力",
    9: "STEP 9: 少人数学級設定",
    0: "🎲 時間割生成・出力",
}

# ============================================================
# ② session_state 全キー初期化
# ============================================================

def init_session_state():
    defaults = {
        "grade_classes":          {g: 1 for g in GRADES},
        "teachers":               [],
        "periods_per_day":        {d: 6 for d in DAYS},
        "periods_per_day_grade":  {g: {d: 6 for d in DAYS} for g in GRADES},
        "weekly_periods":         {},
        "assignments":            {},
        "unavailable":            {},
        "special_rooms":          [],
        "manual_timetable":       {},
        "manual_timetable_sg":    {},
        "small_group_classes":    {},
        "cross_grade_sync":        [],
        "soft_grade_grouping":    True,
        "soft_priority_subjects": [],
        "soft_first_subjects":    [],
        "generated_timetable":    {},
        "generated_timetable_sg": {},
        "timetable_history":      [],   # 過去の生成結果を保持するリスト
        "current_step":           1,
        "timetable_pattern_no":   0,
        "r7_exempt_teachers":     [],   # R7制約（1日最低1コマ空き）を免除する教員リスト
        "homeroom_teachers":      {},   # 担任設定 {クラス名: 教員名}
        "swap_undo_stack":        [],   # 手動編集の undo スタック
        "swap2_slot_a":           None, # 2コマ入れ替え: 選択スロットA
        "swap2_slot_b":           None, # 2コマ入れ替え: 選択スロットB
        "swap2_result":           None, # 2コマ入れ替え: 確認結果
        "swap2_last_teacher":     None, # 2コマ入れ替え: 直前の選択教員
        "edit_cls_slot_a":        None, # 手動編集/クラスから選ぶ: 選択スロットA(day,period,cls,subj)
        "edit_cls_slot_b":        None, # 手動編集/クラスから選ぶ: 選択スロットB(day,period,cls,subj)
        "edit_cls_last":          None, # 手動編集/クラスから選ぶ: 直前の選択クラス
        "edit_cls_cands":         None, # 手動編集/クラスから選ぶ: 候補リスト(キャッシュ)
        "edit2_slot_a":           None, # 手動編集/教員から選ぶ: 2コマ入れ替えスロットA
        "edit2_slot_b":           None, # 手動編集/教員から選ぶ: 2コマ入れ替えスロットB
        "edit2_result":           None, # 手動編集/教員から選ぶ: 2コマ入れ替え確認結果
        "edit2_last_teacher":     None, # 手動編集/教員から選ぶ: 直前の選択教員
        "edit2_cands":            None, # 手動編集/教員から選ぶ: スロットAの候補キャッシュ
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val

# ============================================================
# ③ ヘルパー関数
# ============================================================

def get_all_classes() -> list[str]:
    result = []
    for g in GRADES:
        count = st.session_state["grade_classes"].get(g, 0)
        for c in range(1, count + 1):
            result.append(f"{g}年{c}組")
    return result

def get_grade(class_name: str) -> int:
    return int(class_name[0])

def get_class_index(class_name: str) -> int:
    classes = get_all_classes()
    return classes.index(class_name) if class_name in classes else 999

def get_periods_for_class(cls: str, day: str) -> int:
    """クラスの学年に応じた1日のコマ数を返す。"""
    grade = get_grade(cls)
    return st.session_state["periods_per_day_grade"].get(grade, {}).get(
        day, st.session_state["periods_per_day"].get(day, 6)
    )

def get_total_periods_per_week() -> int:
    return sum(st.session_state["periods_per_day"].values())

def get_max_periods() -> int:
    return max(st.session_state["periods_per_day"].values())


def compute_combined_subjects(cls_weekly_periods: dict) -> list:
    """
    0.5刻みの週コマ数から合成教科リストを生成する。
    小数部(.5)がある教科を SUBJECTS の順番でペアリングして "A/B" 合成教科を返す。

    例: 音楽=1.5, 家庭=1.5 → [{"name": "音楽/家庭", "subjects": ["音楽", "家庭"]}]
    奇数個の場合は最後の1つはペアに入らない（UIで警告）。

    Returns: list of {"name": str, "subjects": [str, str]}
    """
    fractional = [
        s for s in SUBJECTS
        if s in cls_weekly_periods and round(cls_weekly_periods[s] * 2) % 2 == 1
    ]
    combined = []
    for i in range(0, len(fractional) - 1, 2):
        s1, s2 = fractional[i], fractional[i + 1]
        combined.append({"name": f"{s1}/{s2}", "subjects": [s1, s2]})
    return combined


def get_teachers_for_slot(class_name: str, subject: str) -> list[str]:
    raw = st.session_state["assignments"].get(class_name, {}).get(subject, [])
    # 合成教科（"音楽/家庭" など）は構成教科それぞれの担当教員を合算して返す
    if not raw and "/" in subject:
        raw = []
        for part in subject.split("/"):
            raw.extend(
                st.session_state["assignments"].get(class_name, {}).get(part.strip(), [])
            )
    # 万一 ["山八,金場"] のようにカンマ区切りが1要素リストに入っていた場合も正しく展開する
    result = []
    seen = set()
    for item in raw:
        for t in str(item).split(","):
            t = t.strip()
            if t and t not in seen:
                result.append(t)
                seen.add(t)
    return result

def get_sg_teachers(sg_name: str, sg_classes: dict = None) -> list:
    """少人数学級の担当教員リストを subject_settings から取得する。"""
    if sg_classes is None:
        sg_classes = st.session_state.get("small_group_classes", {})
    subj_settings = sg_classes.get(sg_name, {}).get("subject_settings", {})
    return [t for t in {s.get("teacher") for s in subj_settings.values()} if t]


def is_special_room_subject(subject: str, cls: str = None) -> tuple[bool, int]:
    grade = get_grade(cls) if cls else None
    for room in st.session_state["special_rooms"]:
        gs = room.get("grade_subjects", [])
        for g, s in gs:
            if s == subject and (grade is None or g == grade):
                return True, room["capacity"]
    return False, 999

def ensure_class_keys():
    cls_list = get_all_classes()
    wp   = st.session_state["weekly_periods"]
    asgn = st.session_state["assignments"]
    for cls in cls_list:
        wp.setdefault(cls, {s: 0 for s in SUBJECTS})
        asgn.setdefault(cls, {s: [] for s in SUBJECTS})
    for cls in list(wp.keys()):
        if cls not in cls_list:
            del wp[cls]
    for cls in list(asgn.keys()):
        if cls not in cls_list:
            del asgn[cls]

# ============================================================
# ④ Excel 共通書式ヘルパー
# ============================================================

def _xl_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _xl_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _xl_align(wrap: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _xl_write(ws, row: int, col: int, value,
              color: str | None = None, bold: bool = False,
              width: int | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border    = _xl_border()
    cell.alignment = _xl_align()
    if color:
        cell.fill = _xl_fill(color)
    if bold:
        cell.font = Font(bold=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell

# ============================================================
# Part 2: Excel I/O 補助スタイル関数
# ============================================================

def _apply_header_style(ws, row_num: int = 1):
    """指定行にヘッダースタイル（太字・薄グレー・罫線）を適用する。"""
    fill   = PatternFill("solid", fgColor=COLOR_HEADER)
    font   = Font(bold=True)
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.font      = font
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_col_widths(ws, width: int = 14):
    """全列の幅を設定する。"""
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = width


def _apply_cell_border(ws):
    """全データセルに罫線を適用する。"""
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


# ============================================================
# Part 2: Excel関数① テンプレート生成
# ============================================================

@st.cache_data
def generate_template_excel() -> bytes:
    """
    全10シートのテンプレートExcelをBytesIOで返す。
    サイドバーのダウンロードボタンに渡す。
    内容は固定なので st.cache_data でキャッシュする。
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── シート1: クラス設定 ──────────────────────────────────
    ws = wb.create_sheet("クラス設定")
    ws.append(["学年", "クラス数"])
    for g in GRADES:
        ws.append([f"{g}年", 1])
    _apply_header_style(ws)
    _set_col_widths(ws, 12)

    # ── シート2: 先生リスト (ヘッダーなし: A=先生名) ─────────────
    ws = wb.create_sheet("先生リスト")
    ws.append(["（例）古澤"])
    _set_col_widths(ws, 16)

    # ── シート2b: 教科リスト (ヘッダーなし: A=教科名, B=週コマ数, C=必修フラグ) ──
    ws = wb.create_sheet("教科リスト")
    for subj in SUBJECTS:
        ws.append([subj, 0, 1])
    _set_col_widths(ws, 14)

    # ── シート3: コマ数設定 ──────────────────────────────────
    ws = wb.create_sheet("コマ数設定")
    ws.append(["曜日", "1日のコマ数"])
    for d in DAYS:
        ws.append([d, 6])
    _apply_header_style(ws)
    _set_col_widths(ws, 14)

    # ── シート4: 週コマ数 ────────────────────────────────────
    ws = wb.create_sheet("週コマ数")
    ws.append(["学年クラス"] + SUBJECTS)
    ws.append(["（例）1年1組"] + [0] * len(SUBJECTS))
    _apply_header_style(ws)
    _set_col_widths(ws, 10)
    ws.column_dimensions["A"].width = 14

    # ── シート5: 担当割り当て ────────────────────────────────
    ws = wb.create_sheet("担当割り当て")
    ws.append(["学年クラス"] + SUBJECTS)
    ws.append(["（例）1年1組"] + [""] * len(SUBJECTS))
    ws.cell(row=3, column=1).value = "※複数教員はカンマ区切り（例: 古澤,田中）"
    _apply_header_style(ws)
    _set_col_widths(ws, 10)
    ws.column_dimensions["A"].width = 14

    # ── シート6: 不在コマ ────────────────────────────────────
    ws = wb.create_sheet("不在コマ")
    ws.append(["教員名", "曜日", "時限"])
    ws.append(["（例）古澤", "月", 4])
    ws.append(["（例）古澤", "水", 5])
    _apply_header_style(ws)
    _set_col_widths(ws, 14)

    # ── シート7: 特別教室 ────────────────────────────────────
    ws = wb.create_sheet("特別教室")
    ws.append(["教室名", "対応教科", "同時使用可能クラス数"])
    ws.append(["理科室", "理科", 1])
    ws.append(["体育館", "保体", 1])
    ws.append(["音楽室", "音楽", 1])
    _apply_header_style(ws)
    _set_col_widths(ws, 18)

    # ── シート8: 手動時間割 ──────────────────────────────────
    ws = wb.create_sheet("手動時間割")
    ws.append(["学年クラス", "曜日", "時限", "教科"])
    ws.append(["（例）1年1組", "月", 1, "国語"])
    _apply_header_style(ws)
    _set_col_widths(ws, 14)

    # ── シート9: 少人数学級設定 ──────────────────────────────
    ws = wb.create_sheet("少人数学級設定")
    ws.append([
        "少人数学級名", "生徒名", "在籍クラス",
        "🔴少人数教科（カンマ区切り）", "🔵通常教科（カンマ区切り）", "🟡引率教科（カンマ区切り）"
    ])
    ws.append(["少人数A", "生徒A", "1年1組", "国語,算数", "体育,音楽", "理科"])
    ws.append(["少人数A", "生徒B", "2年2組", "国語", "体育", ""])
    _apply_header_style(ws)
    _set_col_widths(ws, 22)

    # ── シート10: 少人数教科設定 ─────────────────────────────
    ws = wb.create_sheet("少人数教科設定")
    ws.append(["少人数学級名", "教科", "週コマ数", "担当教員"])
    ws.append(["少人数A", "国語", 2, "田中"])
    ws.append(["少人数A", "算数", 1, "鈴木"])
    _apply_header_style(ws)
    _set_col_widths(ws, 16)

    # ── シート11: 少人数手動時間割（カラー） ──────────────────
    ws = wb.create_sheet("少人数手動時間割")
    ws.append(["少人数学級名", "生徒名", "曜日", "時限", "色（red/blue/yellow）"])
    ws.append(["少人数A", "生徒A", "月", 1, "red"])
    ws.append(["少人数A", "生徒A", "月", 2, "blue"])
    _apply_header_style(ws)
    _set_col_widths(ws, 16)

    # ── シート12: 担任設定 ────────────────────────────────────
    ws = wb.create_sheet("担任設定")
    ws.append(["学年クラス", "担任教員"])
    for g in GRADES:
        ws.append([f"{g}年1組", ""])
    _apply_header_style(ws)
    _set_col_widths(ws, 16)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Part 1b: 入力正規化ヘルパー
# ============================================================
import unicodedata as _ud

def normalize_value(val) -> str:
    """全角→半角 + strip"""
    if val is None:
        return ""
    return _ud.normalize("NFKC", str(val)).strip()

def normalize_str(val) -> str:
    """normalize_value の別名（互換性）"""
    return normalize_value(val)

def safe_int(val, default: int = 0) -> int:
    """正規化して整数に変換。失敗時は default を返す"""
    try:
        return int(normalize_value(val))
    except (ValueError, TypeError):
        return default

# ============================================================
# Part 2: Excel関数② 設定ファイル読み込み
# ============================================================

def load_settings_from_excel(uploaded_file):
    """
    Excelファイルを読み込み session_state を全て上書きする。
    シートが存在しない場合はそのSTEPをスキップし警告を表示する。
    """
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)

        # ── シート1: クラス設定 ──────────────────────────────
        if "クラス設定" in wb.sheetnames:
            ws = wb["クラス設定"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    try:
                        g = int(str(row[0]).replace("年", ""))
                        val = int(row[1])
                        st.session_state["grade_classes"][g] = val
                        # ウィジェットキャッシュも更新
                        st.session_state[f"grade_{g}"] = val
                    except (ValueError, TypeError):
                        pass

        # ── シート2: 先生リスト (ヘッダーなし, row1から) / 旧: 教員リスト 両対応 ──
        _teacher_sheet = None
        if "先生リスト" in wb.sheetnames:
            _teacher_sheet = wb["先生リスト"]
        elif "教員リスト" in wb.sheetnames:
            _teacher_sheet = wb["教員リスト"]
        if _teacher_sheet is not None:
            teachers = []
            # 先生リストはヘッダーなし (row1 = データ先頭)
            # 教員リストはヘッダーあり (row1 = ヘッダー) → どちらも全行チェック
            for row in _teacher_sheet.iter_rows(min_row=1, values_only=True):
                val = row[0] if row else None
                if (val and str(val).strip()
                        and str(val).strip() not in ("教員名", "先生名")
                        and not str(val).startswith("（例）")):
                    teachers.append(normalize_str(str(val)))
            st.session_state["teachers"] = teachers

        # ── シート2b: 教科リスト (ヘッダーなし: A=教科名, B=週コマ数, C=必修フラグ) ──
        if "教科リスト" in wb.sheetnames:
            ws = wb["教科リスト"]
            subject_defaults = {}
            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or not row[0]:
                    continue
                subj = normalize_str(str(row[0]))
                if subj not in SUBJECTS:
                    continue
                weekly_cnt = safe_int(row[1]) if len(row) > 1 else 0
                # required_flag = row[2] if len(row) > 2 else 1  # 将来利用可
                subject_defaults[subj] = weekly_cnt
            # 教科リストの週コマ数を全クラスのデフォルトとして適用
            if subject_defaults:
                classes = [f"{g}年{c}組"
                           for g in GRADES
                           for c in range(1, st.session_state["grade_classes"].get(g, 0) + 1)]
                if classes:
                    wp = st.session_state.get("weekly_periods", {})
                    for cls in classes:
                        if cls not in wp:
                            wp[cls] = {}
                        for subj, cnt in subject_defaults.items():
                            if wp[cls].get(subj, 0) == 0:
                                wp[cls][subj] = cnt
                    st.session_state["weekly_periods"] = wp

        # ── シート3: コマ数設定 ──────────────────────────────
        if "コマ数設定" in wb.sheetnames:
            ws = wb["コマ数設定"]
            headers = [c.value for c in ws[1]]
            ppd_grade = st.session_state.setdefault(
                "periods_per_day_grade", {g: {d: 6 for d in DAYS} for g in GRADES}
            )
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                cell0 = str(row[0]).strip()
                # 学年別フォーマット（新）: "1年", "2年"...
                if "年" in cell0:
                    try:
                        g = int(cell0.replace("年", ""))
                    except ValueError:
                        continue
                    ppd_grade.setdefault(g, {})
                    for col_idx, d in enumerate(DAYS, start=1):
                        if col_idx < len(row) and row[col_idx] is not None:
                            try:
                                val = int(row[col_idx])
                                ppd_grade[g][d] = val
                                st.session_state[f"ppd_{g}_{d}"] = val
                            except (ValueError, TypeError):
                                pass
                # 旧フォーマット（曜日別）: "月", "火"...
                elif cell0 in DAYS and len(row) > 1 and row[1] is not None:
                    try:
                        val = int(row[1])
                        for g in GRADES:
                            ppd_grade.setdefault(g, {})[cell0] = val
                            st.session_state[f"ppd_{g}_{cell0}"] = val
                    except (ValueError, TypeError):
                        pass
            # periods_per_day（最大値）を更新
            for d in DAYS:
                st.session_state["periods_per_day"][d] = max(
                    ppd_grade.get(g, {}).get(d, 6) for g in GRADES
                )

        # ── シート4: 週コマ数 ────────────────────────────────
        if "週コマ数" in wb.sheetnames:
            ws = wb["週コマ数"]
            headers = [c.value for c in ws[1]][1:]
            weekly = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cls = row[0]
                if not cls or str(cls).startswith("（例）"):
                    continue
                cls = normalize_str(cls)
                weekly[cls] = {}
                for i, subj in enumerate(headers):
                    if subj and i + 1 < len(row):
                        raw = row[i + 1]
                        if raw is None:
                            weekly[cls][subj] = 0.0
                        else:
                            fval = float(raw)
                            # 0.5刻みに丸める（浮動小数点誤差対策）
                            weekly[cls][subj] = round(fval * 2) / 2
            st.session_state["weekly_periods"] = weekly

        # ── シート5: 担当割り当て ────────────────────────────
        if "担当割り当て" in wb.sheetnames:
            ws = wb["担当割り当て"]
            headers = [c.value for c in ws[1]][1:]
            assignments = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cls = row[0]
                if not cls or str(cls).startswith("（例）") \
                        or str(cls).startswith("※"):
                    continue
                cls = normalize_str(cls)
                assignments[cls] = {}
                for i, subj in enumerate(headers):
                    if subj and i + 1 < len(row) and row[i + 1]:
                        teachers = [t.strip()
                                    for t in str(row[i + 1]).split(",")
                                    if t.strip()]
                        assignments[cls][subj] = teachers
                    else:
                        assignments[cls][subj] = []
            st.session_state["assignments"] = assignments

        # ── シート6: 不在コマ ────────────────────────────────
        if "不在コマ" in wb.sheetnames:
            ws = wb["不在コマ"]
            unavail = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                teacher, day, period = row[0], row[1], row[2]
                if not teacher or str(teacher).startswith("（例）"):
                    continue
                teacher = str(teacher).strip()
                if day in DAYS and period:
                    unavail.setdefault(teacher, {}).setdefault(
                        str(day), []).append(int(period))
            st.session_state["unavailable"] = unavail

        # ── シート7: 特別教室 ────────────────────────────────
        if "特別教室" in wb.sheetnames:
            ws = wb["特別教室"]
            rooms = []
            room_map = {}  # name -> room dict
            for row in ws.iter_rows(min_row=2, values_only=True):
                name, grade_val, subj, cap = row[0], row[1], row[2], row[3]
                if not name or str(name).startswith("（例）"):
                    continue
                if not name or not cap:
                    continue
                name = str(name).strip()
                try:
                    cap_int = int(cap)
                except (ValueError, TypeError):
                    cap_int = 1
                if name not in room_map:
                    room_map[name] = {"name": name, "capacity": cap_int, "grade_subjects": []}
                if grade_val and subj:
                    try:
                        grade_int = int(grade_val)
                    except (ValueError, TypeError):
                        grade_int = GRADES[0]
                    room_map[name]["grade_subjects"].append([grade_int, str(subj).strip()])
            rooms = list(room_map.values())
            st.session_state["special_rooms"] = rooms

        # ── シート8: 手動時間割 ──────────────────────────────
        if "手動時間割" in wb.sheetnames:
            ws = wb["手動時間割"]
            manual = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cls, day, period, subj = row[0], row[1], row[2], row[3]
                if not cls or str(cls).startswith("（例）"):
                    continue
                if cls and day in DAYS and period and subj:
                    cls = normalize_str(cls)
                    manual.setdefault(cls, {}).setdefault(
                        str(day), {})[int(period)] = normalize_str(subj)
            st.session_state["manual_timetable"] = manual

        # ── シート9: 少人数学級設定（新形式） ───────────────
        if "少人数学級設定" in wb.sheetnames:
            ws = wb["少人数学級設定"]
            sg = {}
            # 旧形式（5列以下）の場合はスキップ（列数チェック）
            _sg_new_format = ws.max_column >= 6
            for row in ws.iter_rows(min_row=2, values_only=True):
                sg_name = row[0] if row else None
                if not sg_name or str(sg_name).startswith("（例）"):
                    continue
                if not _sg_new_format:
                    continue  # 旧形式シートはデータをスキップ
                sg_name = normalize_str(sg_name)
                if sg_name not in sg:
                    sg[sg_name] = {"students": [], "subject_settings": {}, "student_timetable": {}}
                # Read student
                student_name = normalize_str(row[1]) if len(row) > 1 and row[1] else ""
                home_class = normalize_str(row[2]) if len(row) > 2 and row[2] else ""
                if student_name and home_class:
                    red_subjs = [normalize_str(s) for s in str(row[3]).split(",") if s.strip()] if len(row) > 3 and row[3] else []
                    blue_subjs = [normalize_str(s) for s in str(row[4]).split(",") if s.strip()] if len(row) > 4 and row[4] else []
                    yellow_subjs = [normalize_str(s) for s in str(row[5]).split(",") if s.strip()] if len(row) > 5 and row[5] else []
                    sg[sg_name]["students"].append({
                        "name": student_name,
                        "home_class": home_class,
                        "attributes": {"red": red_subjs, "blue": blue_subjs, "yellow": yellow_subjs}
                    })
            st.session_state["small_group_classes"] = sg

        # ── シート10: 少人数教科設定 ─────────────────────────
        if "少人数教科設定" in wb.sheetnames:
            ws = wb["少人数教科設定"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                sg_name = normalize_str(row[0]) if row[0] else ""
                subj = normalize_str(row[1]) if row[1] else ""
                if not sg_name or not subj:
                    continue
                if sg_name not in st.session_state["small_group_classes"]:
                    continue
                weekly = safe_int(row[2], 0)
                teacher = normalize_str(row[3]) if row[3] else ""
                st.session_state["small_group_classes"][sg_name]["subject_settings"].setdefault(subj, {})
                st.session_state["small_group_classes"][sg_name]["subject_settings"][subj]["weekly"] = weekly
                st.session_state["small_group_classes"][sg_name]["subject_settings"][subj]["teacher"] = teacher

        # ── シート11: 少人数手動時間割（カラー） ────────────
        if "少人数手動時間割" in wb.sheetnames:
            ws = wb["少人数手動時間割"]
            # 旧形式（4列以下: 少人数学級名,曜日,時限,教科）はスキップ
            _sgt_new_format = ws.max_column >= 5
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not _sgt_new_format:
                    continue  # 旧形式シートはスキップ
                if len(row) < 5:
                    continue
                sg_name = normalize_str(row[0]) if row[0] else ""
                sname = normalize_str(row[1]) if row[1] else ""
                day_ = normalize_str(row[2]) if row[2] else ""
                period_ = safe_int(row[3], 0)
                color = normalize_str(row[4]) if row[4] else ""
                if not sg_name or not sname or not day_ or not period_ or not color:
                    continue
                if sg_name not in st.session_state["small_group_classes"]:
                    continue
                sg_data = st.session_state["small_group_classes"][sg_name]
                sg_data.setdefault("student_timetable", {}).setdefault(sname, {}).setdefault(day_, {})[period_] = color

        # ── シート12: 担任設定 ────────────────────────────────
        if "担任設定" in wb.sheetnames:
            ws = wb["担任設定"]
            homeroom = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cls_val = normalize_str(row[0]) if row and row[0] else ""
                teacher_val = normalize_str(row[1]) if row and len(row) > 1 and row[1] else ""
                if cls_val:
                    homeroom[cls_val] = teacher_val
            st.session_state["homeroom_teachers"] = homeroom

        # ── ウィジェットキーをクリアして各STEPの入力欄を再初期化 ──
        # Streamlit はウィジェットキーが session_state に残っていると
        # value= パラメータより優先するため、Excel ロード後に古いキーを
        # 削除しておかないと読み込んだ値が画面に反映されない。
        # ※ "grade_" は "grade_classes" を誤って削除しないよう、
        #   数字サフィックスのもの（grade_1, grade_2, grade_3）のみ対象にする。
        for _k in list(st.session_state.keys()):
            if _k.startswith("grade_") and _k[6:].isdigit():
                del st.session_state[_k]
            elif _k.startswith(("ppd_", "wp_", "assign_", "unavail_",
                                 "sgt_",            # 少人数生徒時間割カラー選択
                                 "attr_red_",       # 少人数生徒赤教科マルチセレクト
                                 "attr_blue_",      # 少人数生徒青教科マルチセレクト
                                 "attr_yellow_",    # 少人数生徒黄教科マルチセレクト
                                 )):
                del st.session_state[_k]

        st.success("✅ Excelファイルから設定を読み込みました（全STEP上書き）")

    except Exception as e:
        st.error(f"❌ 読み込みエラー: {e}")
        st.info("💡 テンプレートExcelを使用して入力してください")


# ============================================================
# Part 2: Excel関数③ 設定ファイル書き出し
# ============================================================

def save_settings_to_excel() -> bytes:
    """
    現在のsession_stateをExcelに書き出してBytesIOで返す。
    generate_template_excel()と同じシート構成・書式で出力する。
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # シート1: クラス設定
    ws = wb.create_sheet("クラス設定")
    ws.append(["学年", "クラス数"])
    for g in GRADES:
        ws.append([f"{g}年", st.session_state["grade_classes"].get(g, 0)])
    _apply_header_style(ws); _set_col_widths(ws, 12)

    # シート2: 先生リスト (ヘッダーなし)
    ws = wb.create_sheet("先生リスト")
    for t in st.session_state["teachers"]:
        ws.append([t])
    _set_col_widths(ws, 16)

    # シート2b: 教科リスト (ヘッダーなし)
    ws = wb.create_sheet("教科リスト")
    weekly_p = st.session_state.get("weekly_periods", {})
    classes = [f"{g}年{c}組"
               for g in GRADES
               for c in range(1, st.session_state["grade_classes"].get(g, 0) + 1)]
    first_cls = classes[0] if classes else None
    for subj in SUBJECTS:
        cnt = weekly_p.get(first_cls, {}).get(subj, 0) if first_cls else 0
        ws.append([subj, cnt, 1])
    _set_col_widths(ws, 14)

    # シート3: コマ数設定（学年別）
    ws = wb.create_sheet("コマ数設定")
    ws.append(["学年"] + DAYS)
    ppd_grade = st.session_state.get("periods_per_day_grade", {})
    for g in GRADES:
        row = [f"{g}年"] + [ppd_grade.get(g, {}).get(d, 6) for d in DAYS]
        ws.append(row)
    _apply_header_style(ws); _set_col_widths(ws, 10)

    # シート4: 週コマ数
    ws = wb.create_sheet("週コマ数")
    ws.append(["学年クラス"] + SUBJECTS)
    for cls in get_all_classes():
        row = [cls]
        for subj in SUBJECTS:
            row.append(
                st.session_state["weekly_periods"].get(cls, {}).get(subj, 0))
        ws.append(row)
    _apply_header_style(ws); _set_col_widths(ws, 10)
    ws.column_dimensions["A"].width = 14

    # シート5: 担当割り当て
    ws = wb.create_sheet("担当割り当て")
    ws.append(["学年クラス"] + SUBJECTS)
    for cls in get_all_classes():
        row = [cls]
        for subj in SUBJECTS:
            teachers = st.session_state["assignments"].get(
                cls, {}).get(subj, [])
            row.append(",".join(teachers))
        ws.append(row)
    _apply_header_style(ws); _set_col_widths(ws, 10)
    ws.column_dimensions["A"].width = 14

    # シート6: 不在コマ
    ws = wb.create_sheet("不在コマ")
    ws.append(["教員名", "曜日", "時限"])
    for teacher, day_map in st.session_state["unavailable"].items():
        for day, periods in day_map.items():
            for p in periods:
                ws.append([teacher, day, p])
    _apply_header_style(ws); _set_col_widths(ws, 14)

    # シート7: 特別教室
    ws = wb.create_sheet("特別教室")
    ws.append(["教室名", "対象学年", "対応教科", "同時使用可能クラス数"])
    for room in st.session_state["special_rooms"]:
        gs = room.get("grade_subjects", [])
        if gs:
            for g, s in gs:
                ws.append([room["name"], g, s, room["capacity"]])
        else:
            ws.append([room["name"], "", "", room["capacity"]])
    _apply_header_style(ws); _set_col_widths(ws, 18)

    # シート8: 手動時間割
    ws = wb.create_sheet("手動時間割")
    ws.append(["学年クラス", "曜日", "時限", "教科"])
    for cls, day_map in st.session_state["manual_timetable"].items():
        for day, period_map in day_map.items():
            for period, subj in period_map.items():
                ws.append([cls, day, period, subj])
    _apply_header_style(ws); _set_col_widths(ws, 14)

    # シート9: 少人数学級設定（新形式）
    ws = wb.create_sheet("少人数学級設定")
    ws.append(["少人数学級名", "生徒名", "在籍クラス", "🔴少人数教科（カンマ区切り）",
               "🔵通常教科（カンマ区切り）", "🟡引率教科（カンマ区切り）"])
    for sg_name, sg_data in st.session_state["small_group_classes"].items():
        students = sg_data.get("students", [])
        if not students:
            ws.append([sg_name, "", "", "", "", ""])
        for student in students:
            attrs = student.get("attributes", {})
            ws.append([
                sg_name,
                student.get("name", ""),
                student.get("home_class", ""),
                ",".join(attrs.get("red", [])),
                ",".join(attrs.get("blue", [])),
                ",".join(attrs.get("yellow", []))
            ])
    _apply_header_style(ws); _set_col_widths(ws, 18)

    # シート10: 少人数教科設定
    ws = wb.create_sheet("少人数教科設定")
    ws.append(["少人数学級名", "教科", "週コマ数", "担当教員"])
    for sg_name, sg_data in st.session_state["small_group_classes"].items():
        for subj, settings in sg_data.get("subject_settings", {}).items():
            weekly = settings.get("weekly", 0)
            teacher = settings.get("teacher", "")
            if weekly > 0 or teacher:
                ws.append([sg_name, subj, weekly, teacher])
    _apply_header_style(ws); _set_col_widths(ws, 16)

    # シート11: 少人数手動時間割（カラー）
    ws = wb.create_sheet("少人数手動時間割")
    ws.append(["少人数学級名", "生徒名", "曜日", "時限", "色（red/blue/yellow）"])
    for sg_name, sg_data in st.session_state["small_group_classes"].items():
        for sname, day_map in sg_data.get("student_timetable", {}).items():
            for day_, period_map in day_map.items():
                for period_, color in period_map.items():
                    if color:
                        ws.append([sg_name, sname, day_, period_, color])
    _apply_header_style(ws); _set_col_widths(ws, 16)

    # シート12: 担任設定
    ws = wb.create_sheet("担任設定")
    ws.append(["学年クラス", "担任教員"])
    for cls, teacher in st.session_state.get("homeroom_teachers", {}).items():
        if cls and teacher:
            ws.append([cls, teacher])
    _apply_header_style(ws); _set_col_widths(ws, 16)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Part 2: サイドバー UI
# ============================================================

def render_sidebar():
    """
    サイドバーにExcelボタン群・STEPナビゲーションを表示する。
    """
    st.sidebar.title("📅 小学校時間割生成")
    st.sidebar.markdown("---")

    # Excelテンプレートダウンロード
    st.sidebar.download_button(
        label="📥 テンプレートDL",
        data=generate_template_excel(),
        file_name="時間割テンプレート.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="入力用テンプレートをダウンロードします"
    )

    # 設定ファイルアップロード
    uploaded = st.sidebar.file_uploader(
        "📤 設定ファイルをアップロード",
        type=["xlsx"],
        help="保存済みの設定ファイルをアップロードします（全STEP上書き）"
    )
    if uploaded:
        # ファイルの同一性を名前+サイズで判定し、新規アップロード時のみ読み込む
        # （画面遷移のたびに再読み込みされて手動入力が消えるのを防ぐ）
        file_key = f"{uploaded.name}_{uploaded.size}"
        if st.session_state.get("_last_uploaded_excel_key") != file_key:
            load_settings_from_excel(uploaded)
            st.session_state["_last_uploaded_excel_key"] = file_key

    # 現在設定を保存
    st.sidebar.download_button(
        label="💾 現在の設定を保存",
        data=save_settings_to_excel(),
        file_name=f"時間割設定_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="現在の設定をExcelに保存します"
    )

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📋 設定STEP")

    # STEPナビゲーションボタン
    step_labels = {
        1: "① クラス数設定",
        2: "② 教員登録",
        3: "③ コマ数設定",
        4: "④ 週コマ数",
        5: "⑤ 担当教員割り当て",
        6: "⑥ 不在コマ設定",
        7: "⑦ 特別教室設定",
        8: "⑧ 手動時間割入力",
        9: "⑨ 少人数学級設定",
        10: "⑩ 学年間教科同期設定",
    }
    for step_num, label in step_labels.items():
        if st.sidebar.button(label, key=f"nav_{step_num}",
                             use_container_width=True):
            st.session_state["current_step"] = step_num
            st.rerun()

    st.sidebar.markdown("---")
    if st.sidebar.button("🚀 時間割を生成・確認",
                         type="primary", use_container_width=True):
        st.session_state["current_step"] = 0
        st.rerun()


# ============================================================
# Part 2: STEP 1〜5 UI
# ============================================================

def render_step1():
    st.header("STEP 1｜学年・クラス数設定")
    st.caption("各学年のクラス数を設定してください（1〜6クラス）")

    cols = st.columns(len(GRADES))
    for i, g in enumerate(GRADES):
        with cols[i]:
            st.session_state["grade_classes"][g] = st.number_input(
                f"{g}年生のクラス数",
                min_value=1, max_value=6,
                value=st.session_state["grade_classes"].get(g, 1),
                step=1,
                key=f"grade_{g}"
            )

    classes = get_all_classes()
    st.markdown("---")
    st.info(f"📋 設定クラス一覧（全 **{len(classes)}** クラス）\n\n"
            + "　".join(classes))


def render_step2():
    st.header("STEP 2｜教員登録")
    st.caption("時間割に関わる教員を全員登録してください")

    col1, col2 = st.columns([3, 1])
    with col1:
        new_name = st.text_input(
            "教員名を入力",
            placeholder="例: 古澤",
            key="new_teacher_input",
            label_visibility="collapsed"
        )
    with col2:
        if st.button("➕ 追加", use_container_width=True):
            name = new_name.strip()
            if name and name not in st.session_state["teachers"]:
                st.session_state["teachers"].append(name)
                st.rerun()
            elif name in st.session_state["teachers"]:
                st.warning("同じ名前の教員がすでに登録されています")

    st.markdown("---")
    st.markdown(f"#### 登録済み教員（{len(st.session_state['teachers'])}名）")

    if not st.session_state["teachers"]:
        st.info("教員が登録されていません")
    else:
        exempt = st.session_state.get("r7_exempt_teachers", [])
        _header = st.columns([4, 2, 1])
        _header[1].caption("R7免除（終日授業OK）")
        for i, teacher in enumerate(st.session_state["teachers"]):
            col1, col2, col3 = st.columns([4, 2, 1])
            col1.write(f"👤 {teacher}")
            is_exempt = col2.checkbox(
                "R7免除",
                value=teacher in exempt,
                key=f"r7_exempt_{i}",
                label_visibility="collapsed"
            )
            if is_exempt and teacher not in exempt:
                exempt.append(teacher)
                st.session_state["r7_exempt_teachers"] = exempt
            elif not is_exempt and teacher in exempt:
                exempt.remove(teacher)
                st.session_state["r7_exempt_teachers"] = exempt
            if col3.button("🗑️", key=f"del_teacher_{i}",
                           use_container_width=True):
                st.session_state["teachers"].pop(i)
                if teacher in st.session_state["r7_exempt_teachers"]:
                    st.session_state["r7_exempt_teachers"].remove(teacher)
                st.rerun()
        if exempt:
            st.caption(f"R7免除中: {', '.join(exempt)}")


def render_step3():
    st.header("STEP 3｜曜日別1日コマ数設定（学年別）")
    st.caption("各学年・各曜日の1日のコマ数を設定してください（1〜8コマ）")

    ppd_grade = st.session_state["periods_per_day_grade"]
    # 学年がGRADESに合わせて初期化されているか確認
    for g in GRADES:
        ppd_grade.setdefault(g, {d: 6 for d in DAYS})

    header_cols = st.columns([2] + [1] * len(DAYS))
    with header_cols[0]:
        st.markdown("**学年**")
    for i, d in enumerate(DAYS):
        with header_cols[i + 1]:
            st.markdown(f"**{d}曜**")

    for g in GRADES:
        row_cols = st.columns([2] + [1] * len(DAYS))
        with row_cols[0]:
            st.markdown(f"**{g}年生**")
        for i, d in enumerate(DAYS):
            with row_cols[i + 1]:
                val = st.number_input(
                    f"{g}年{d}曜",
                    min_value=1, max_value=8,
                    value=ppd_grade[g].get(d, 6),
                    step=1,
                    key=f"ppd_{g}_{d}",
                    label_visibility="collapsed"
                )
                ppd_grade[g][d] = val

    # periods_per_day（最大値）を自動計算して更新
    for d in DAYS:
        st.session_state["periods_per_day"][d] = max(
            ppd_grade[g].get(d, 6) for g in GRADES
        )

    st.markdown("---")
    rows = []
    for g in GRADES:
        total_g = sum(ppd_grade[g].get(d, 6) for d in DAYS)
        rows.append(f"{g}年: 週{total_g}コマ")
    st.info("📊 学年別 週合計コマ数\n\n" + "　".join(rows))


def render_step4():
    st.header("STEP 4｜週コマ数設定")
    st.caption("各クラス・各教科の週あたりのコマ数を入力してください")

    classes = get_all_classes()
    total_per_week = get_total_periods_per_week()

    if not classes:
        st.warning("先にSTEP 1でクラス数を設定してください")
        return

    for cls in classes:
        with st.expander(f"📘 {cls}", expanded=False):
            if cls not in st.session_state["weekly_periods"]:
                st.session_state["weekly_periods"][cls] = {s: 0.0 for s in SUBJECTS}

            total = 0.0
            cols_per_row = 5
            for row_start in range(0, len(SUBJECTS), cols_per_row):
                row_subjects = SUBJECTS[row_start:row_start + cols_per_row]
                cols = st.columns(cols_per_row)
                for j, subj in enumerate(row_subjects):
                    with cols[j]:
                        raw_val = st.session_state["weekly_periods"][cls].get(subj, 0)
                        val = st.number_input(
                            subj,
                            min_value=0.0,
                            max_value=float(total_per_week),
                            value=float(raw_val),
                            step=0.5,
                            format="%.1f",
                            key=f"wp_{cls}_{subj}"
                        )
                        st.session_state["weekly_periods"][cls][subj] = val
                        total += val

            # 合成教科プレビュー
            combined_list = compute_combined_subjects(
                st.session_state["weekly_periods"][cls])
            if combined_list:
                combined_strs = [f"**{c['name']}** × 1コマ" for c in combined_list]
                st.info("合成教科: " + "、".join(combined_strs))
                # 小数部ありで奇数個の場合は警告
                fractional_count = sum(
                    1 for s in SUBJECTS
                    if round(st.session_state["weekly_periods"][cls].get(s, 0) * 2) % 2 == 1
                )
                if fractional_count % 2 != 0:
                    st.warning("⚠️ 0.5コマが奇数個あります。最後の1つはペアに入れず0.5が切り捨てられます。")

            # 実効コマ数 = 各教科の整数部合計 + 合成教科数
            effective_total = (
                sum(int(v) for v in st.session_state["weekly_periods"][cls].values())
                + len(combined_list)
            )
            if effective_total > total_per_week:
                st.error(
                    f"⚠️ 実効合計 {effective_total} コマ → 週上限 {total_per_week} コマを超過！")
            elif effective_total == total_per_week:
                st.success(f"✅ 実効合計 {effective_total} コマ（週上限ちょうど）")
            else:
                st.warning(
                    f"📝 実効合計 {effective_total} コマ（週上限まで残り {total_per_week - effective_total} コマ）")


def render_step5():
    st.header("STEP 5｜教科担当教員割り当て")
    st.caption("各クラス・各教科の担当教員を選択してください（複数選択可）")

    classes  = get_all_classes()
    teachers = st.session_state["teachers"]

    if not teachers:
        st.warning("先にSTEP 2で教員を登録してください")
        return
    if not classes:
        st.warning("先にSTEP 1でクラス数を設定してください")
        return

    for cls in classes:
        with st.expander(f"📘 {cls}", expanded=False):
            if cls not in st.session_state["assignments"]:
                st.session_state["assignments"][cls] = {s: [] for s in SUBJECTS}

            # 担任設定
            current_homeroom = st.session_state["homeroom_teachers"].get(cls, "")
            homeroom_options = ["（未設定）"] + teachers
            homeroom_default = current_homeroom if current_homeroom in teachers else "（未設定）"
            selected_homeroom = st.selectbox(
                "担任",
                options=homeroom_options,
                index=homeroom_options.index(homeroom_default),
                key=f"homeroom_{cls}"
            )
            st.session_state["homeroom_teachers"][cls] = (
                selected_homeroom if selected_homeroom != "（未設定）" else ""
            )
            st.markdown("---")

            cols_per_row = 5
            for row_start in range(0, len(SUBJECTS), cols_per_row):
                row_subjects = SUBJECTS[row_start:row_start + cols_per_row]
                cols = st.columns(cols_per_row)
                for j, subj in enumerate(row_subjects):
                    with cols[j]:
                        current = st.session_state["assignments"][cls].get(subj, [])
                        valid_current = [t for t in current if t in teachers]
                        selected = st.multiselect(
                            subj,
                            options=teachers,
                            default=valid_current,
                            key=f"assign_{cls}_{subj}"
                        )
                        st.session_state["assignments"][cls][subj] = selected


# ============================================================
# Part 3: STEP 6〜9 UI
# ============================================================

def render_step6():
    st.header("STEP 6｜教員不在コマ設定")
    st.caption("各教員が授業に入れない時限をチェックしてください")

    teachers = st.session_state["teachers"]
    if not teachers:
        st.warning("先にSTEP 2で教員を登録してください")
        return

    for teacher in teachers:
        with st.expander(f"👤 {teacher}", expanded=False):
            if teacher not in st.session_state["unavailable"]:
                st.session_state["unavailable"][teacher] = {}

            for day in DAYS:
                periods = st.session_state["periods_per_day"].get(day, 6)
                current_unavail = st.session_state["unavailable"][teacher].get(day, [])

                st.markdown(f"**{day}曜日**")
                cols = st.columns(periods)
                selected = []
                for p in range(1, periods + 1):
                    with cols[p - 1]:
                        checked = st.checkbox(
                            f"{p}限",
                            value=(p in current_unavail),
                            key=f"unavail_{teacher}_{day}_{p}"
                        )
                        if checked:
                            selected.append(p)
                st.session_state["unavailable"][teacher][day] = selected

            total_unavail = sum(
                len(v) for v in st.session_state["unavailable"][teacher].values()
            )
            if total_unavail > 0:
                st.caption(f"合計不在コマ数: {total_unavail} コマ/週")


def render_step7():
    st.header("STEP 7｜特別教室設定")
    st.caption(
        "特別教室を登録し、使用する学年×教科の組み合わせと同時使用可能クラス数を設定してください。\n\n"
        "例：「第一体育館」に 1年体育・2年体育・3年体育 を登録 → "
        "同じ時限にこれらのクラスが合計で設定した上限数まで使用できます。"
    )

    if st.button("➕ 特別教室を追加", key="add_room"):
        st.session_state["special_rooms"].append(
            {"name": "", "capacity": 1, "grade_subjects": []}
        )
        st.rerun()

    if not st.session_state["special_rooms"]:
        st.info("特別教室が登録されていません。上のボタンで追加してください。")
        return

    for i, room in enumerate(st.session_state["special_rooms"]):
        # 旧フォーマット（grade/subject）を新フォーマットに自動変換
        if "grade" in room and "subject" in room and "grade_subjects" not in room:
            room["grade_subjects"] = [(room.pop("grade"), room.pop("subject"))]

        with st.expander(f"🏫 {room.get('name') or f'教室{i+1}'}", expanded=True):
            col1, col2, col3 = st.columns([4, 2, 1])
            with col1:
                room["name"] = st.text_input(
                    "教室名",
                    value=room.get("name", ""),
                    placeholder="例: 第一体育館",
                    key=f"room_name_{i}"
                )
            with col2:
                room["capacity"] = st.number_input(
                    "同時使用可能クラス数",
                    min_value=1, max_value=10,
                    value=room.get("capacity", 1),
                    step=1,
                    key=f"room_cap_{i}"
                )
            with col3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("🗑️ 削除", key=f"del_room_{i}", use_container_width=True):
                    st.session_state["special_rooms"].pop(i)
                    st.rerun()

            # 学年×教科ペアの追加
            st.markdown("**使用する学年×教科**")
            col_g, col_s, col_add = st.columns([2, 3, 2])
            with col_g:
                add_grade = st.selectbox(
                    "学年", GRADES, key=f"room_add_g_{i}",
                    format_func=lambda g: f"{g}年"
                )
            with col_s:
                add_subj = st.selectbox(
                    "教科", SUBJECTS, key=f"room_add_s_{i}"
                )
            with col_add:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("➕ 追加", key=f"room_add_btn_{i}", use_container_width=True):
                    gs = room.setdefault("grade_subjects", [])
                    pair = [add_grade, add_subj]
                    if pair not in gs:
                        gs.append(pair)
                        st.rerun()

            # 登録済みペア一覧
            gs_list = room.get("grade_subjects", [])
            if gs_list:
                for j, (g, s) in enumerate(gs_list):
                    c1, c2 = st.columns([6, 1])
                    with c1:
                        st.markdown(f"・{g}年　{s}")
                    with c2:
                        if st.button("✕", key=f"room_del_gs_{i}_{j}"):
                            gs_list.pop(j)
                            st.rerun()
            else:
                st.info("学年×教科の組み合わせを追加してください")


def render_step8():
    st.header("STEP 8｜手動時間割入力")
    st.caption(
        "自動生成前に手動で配置したいコマを入力してください。"
        "手動入力したコマは自動生成で上書きされません（水色で表示）。"
    )

    tab_normal, tab_sg = st.tabs(["📘 通常学級", "📗 少人数学級"])

    # ── 通常学級タブ ─────────────────────────────────────────
    with tab_normal:
        classes = get_all_classes()
        if not classes:
            st.warning("先にSTEP 1でクラス数を設定してください")
        else:
            cls = st.selectbox("クラスを選択", classes, key="manual_cls_select")

            if cls not in st.session_state["manual_timetable"]:
                st.session_state["manual_timetable"][cls] = {}

            # 少人数カラー制約を取得
            color_constraints = get_color_constraints(
                st.session_state.get("small_group_classes", {})
            )
            cls_constraints = color_constraints.get(cls, {})

            COLOR_BG = {
                "red":    "#FFD0D0",
                "blue":   "#D0E8FF",
                "yellow": "#FFF5CC",
                "mixed":  "#D0F0FF",
            }
            COLOR_LABEL = {
                "red":    "🔴",
                "blue":   "🔵",
                "yellow": "🟡",
                "mixed":  "🔵🟡",
            }

            for day in DAYS:
                periods = st.session_state["periods_per_day"].get(day, 6)
                st.markdown(f"**{day}曜日**")
                cols = st.columns(periods)
                day_data = st.session_state["manual_timetable"][cls].setdefault(day, {})

                for p in range(1, periods + 1):
                    with cols[p - 1]:
                        slot_info = cls_constraints.get(day, {}).get(p)

                        if slot_info and slot_info["color"] == "red":
                            color = slot_info["color"]
                            allowed = sorted(slot_info["allowed"])
                            label = COLOR_LABEL.get(color, "")
                            bg    = COLOR_BG.get(color, "#FFFFFF")
                            # 色付きラベル表示（赤スロットのみ: 少人数教科のみ選択可）
                            st.markdown(
                                f'<div style="background:{bg};border-radius:4px;'
                                f'padding:2px 6px;margin-bottom:4px;font-size:0.85em;">'
                                f'{p}限 {label}</div>',
                                unsafe_allow_html=True
                            )
                            options = ["（未入力）"] + allowed
                        else:
                            st.markdown(f"**{p}限**")
                            options = ["（未入力）"] + SUBJECTS

                        current_val = day_data.get(p, "（未入力）")
                        # 制約外の値が入っていたらリセット
                        if current_val not in options:
                            current_val = "（未入力）"
                            day_data.pop(p, None)
                        idx = options.index(current_val) if current_val in options else 0
                        chosen = st.selectbox(
                            f"{p}限",
                            options=options,
                            index=idx,
                            key=f"manual_{cls}_{day}_{p}",
                            label_visibility="collapsed"
                        )
                        if chosen == "（未入力）":
                            day_data.pop(p, None)
                        else:
                            day_data[p] = chosen

            total_manual = sum(
                len(v) for v in
                st.session_state["manual_timetable"].get(cls, {}).values()
            )
            st.markdown("---")
            col1, col2 = st.columns([3, 1])
            col1.info(f"📌 {cls} の手動入力コマ数: {total_manual} コマ")
            if col2.button(f"🗑️ {cls} をすべてクリア",
                           key=f"clear_manual_{cls}"):
                st.session_state["manual_timetable"][cls] = {}
                st.rerun()

    # ── 少人数学級タブ ───────────────────────────────────────
    with tab_sg:
        sg_list = list(st.session_state["small_group_classes"].keys())
        if not sg_list:
            st.info("先にSTEP 9で少人数学級を登録してください")
        else:
            sg_name = st.selectbox(
                "少人数学級を選択", sg_list, key="manual_sg_select"
            )

            if sg_name not in st.session_state["manual_timetable_sg"]:
                st.session_state["manual_timetable_sg"][sg_name] = {}

            options = ["（未入力）"] + SUBJECTS

            for day in DAYS:
                periods = st.session_state["periods_per_day"].get(day, 6)
                st.markdown(f"**{day}曜日**")
                cols = st.columns(periods)
                day_data = st.session_state["manual_timetable_sg"][sg_name].setdefault(day, {})

                for p in range(1, periods + 1):
                    with cols[p - 1]:
                        current_val = day_data.get(p, "（未入力）")
                        idx = options.index(current_val) \
                            if current_val in options else 0
                        chosen = st.selectbox(
                            f"{p}限",
                            options=options,
                            index=idx,
                            key=f"manual_sg_{sg_name}_{day}_{p}"
                        )
                        if chosen == "（未入力）":
                            day_data.pop(p, None)
                        else:
                            day_data[p] = chosen

            total_manual_sg = sum(
                len(v) for v in
                st.session_state["manual_timetable_sg"].get(sg_name, {}).values()
            )
            st.markdown("---")
            col1, col2 = st.columns([3, 1])
            col1.info(f"📌 {sg_name} の手動入力コマ数: {total_manual_sg} コマ")
            if col2.button(f"🗑️ {sg_name} をすべてクリア",
                           key=f"clear_manual_sg_{sg_name}"):
                st.session_state["manual_timetable_sg"][sg_name] = {}
                st.rerun()


def get_color_constraints(small_group_classes: dict) -> dict:
    """
    少人数学級の手動カラー設定から、各在籍クラスの各スロットに課される教科制約を計算する。

    Returns:
        {
            class_name: {
                day: {
                    period: {
                        "color": "red" | "blue" | "yellow" | "mixed",
                        "allowed": set of allowed subjects (union of all students' color subjects),
                        "students": {student_name: [allowed subjects for this student]}
                    }
                }
            }
        }

    Notes:
        - red → only red subjects allowed
        - blue or yellow → blue + yellow subjects allowed (both treated as "normal class")
        - Multiple students in same home class: union of allowed subjects
        - If a slot has no color set for any student, no constraint
    """
    constraints = {}

    for sg_name, sg_data in small_group_classes.items():
        students = sg_data.get("students", [])
        student_timetable = sg_data.get("student_timetable", {})

        for student in students:
            sname = student["name"]
            home_class = student.get("home_class", "")
            if not home_class:
                continue
            attrs = student.get("attributes", {})
            red_subjects = set(attrs.get("red", []))
            blue_subjects = set(attrs.get("blue", []))
            yellow_subjects = set(attrs.get("yellow", []))
            normal_subjects = blue_subjects | yellow_subjects  # blue and yellow treated same in regular class

            s_tt = student_timetable.get(sname, {})

            for day, period_map in s_tt.items():
                for period, color in period_map.items():
                    if not color:
                        continue

                    # Determine allowed subjects for this student at this slot
                    if color == "red":
                        allowed = red_subjects
                        effective_color = "red"
                    else:  # blue or yellow → normal class, blue+yellow subjects
                        allowed = normal_subjects
                        effective_color = "mixed"

                    # Update constraints dict
                    if home_class not in constraints:
                        constraints[home_class] = {}
                    if day not in constraints[home_class]:
                        constraints[home_class][day] = {}

                    if period not in constraints[home_class][day]:
                        constraints[home_class][day][period] = {
                            "color": effective_color,
                            "allowed": set(allowed),
                            "students": {sname: list(allowed)}
                        }
                    else:
                        existing = constraints[home_class][day][period]
                        # Union of allowed subjects
                        existing["allowed"] |= allowed
                        existing["students"][sname] = list(allowed)
                        # Color: if mixed colors, mark as mixed
                        if existing["color"] != effective_color:
                            existing["color"] = "mixed"

    return constraints


def _format_red_allowed_per_student(slot_info: dict) -> str:
    """
    赤スロットの許容教科を生徒ごとに整形して返す（括弧付き）。
    例: "(AさんB:国語・算数\nBさんC:国語・算数・理科)"
    生徒情報がない場合は全体の union を返す。
    """
    if not slot_info or slot_info.get("color") != "red":
        return ""
    students = slot_info.get("students", {})
    if not students:
        allowed = slot_info.get("allowed", set())
        return ("(" + "・".join(sorted(allowed)) + ")") if allowed else ""
    parts = []
    for sname, subjs in sorted(students.items()):
        if subjs:
            parts.append(f"{sname}:{'・'.join(sorted(subjs))}")
    if not parts:
        return ""
    return "(" + "\n".join(parts) + ")"


def render_sg_students(sg_name: str, sg_data: dict):
    """生徒登録タブ"""
    classes = get_all_classes()
    students = sg_data.get("students", [])

    # 新規生徒追加
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        new_student_name = st.text_input("生徒の識別名", placeholder="例: 生徒A", key=f"new_student_name_{sg_name}")
    with col2:
        new_student_class = st.selectbox("在籍クラス", options=[""] + classes, key=f"new_student_class_{sg_name}")
    with col3:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ 追加", key=f"add_student_{sg_name}", use_container_width=True):
            name = new_student_name.strip()
            if name and new_student_class:
                # 重複チェック
                if any(s["name"] == name for s in students):
                    st.warning(f"「{name}」はすでに登録されています")
                else:
                    new_student = {
                        "name": name,
                        "home_class": new_student_class,
                        "attributes": {"red": [], "blue": [], "yellow": []}
                    }
                    st.session_state["small_group_classes"][sg_name]["students"].append(new_student)
                    st.rerun()
            else:
                st.warning("名前と在籍クラスを入力してください")

    if not students:
        st.info("生徒が登録されていません")
        return

    st.markdown("---")
    st.markdown("**登録済み生徒と教科属性設定**")
    st.caption("各生徒の教科を3つの属性に分類してください\n- 🔴 少人数：少人数学級で受ける\n- 🔵 通常：通常学級で受ける\n- 🟡 引率：通常学級で受けるが少人数の先生が引率")

    COLOR_OPTIONS = {
        "":       "　（なし）",
        "red":    "🔴 少人数",
        "blue":   "🔵 通常",
        "yellow": "🟡 引率",
    }

    for i, student in enumerate(students):
        sname = student["name"]
        with st.expander(f"👤 {sname}（{student['home_class']}）", expanded=False):
            attrs = student.get("attributes", {"red": [], "blue": [], "yellow": []})

            # ── 教科属性設定 ──────────────────────────────────
            col_r, col_b, col_y = st.columns(3)
            with col_r:
                key_r = f"attr_red_{sg_name}_{i}"
                # keyがない場合のみ初期値をセット（Excelロード後はキーが削除済みのため正しい値で初期化）
                if key_r not in st.session_state:
                    st.session_state[key_r] = attrs.get("red", [])
                st.multiselect("🔴 少人数で受ける教科", options=SUBJECTS, key=key_r)
                st.session_state["small_group_classes"][sg_name]["students"][i]["attributes"]["red"] = st.session_state[key_r]

            with col_b:
                key_b = f"attr_blue_{sg_name}_{i}"
                if key_b not in st.session_state:
                    st.session_state[key_b] = attrs.get("blue", [])
                st.multiselect("🔵 通常で受ける教科", options=SUBJECTS, key=key_b)
                st.session_state["small_group_classes"][sg_name]["students"][i]["attributes"]["blue"] = st.session_state[key_b]

            with col_y:
                key_y = f"attr_yellow_{sg_name}_{i}"
                if key_y not in st.session_state:
                    st.session_state[key_y] = attrs.get("yellow", [])
                st.multiselect("🟡 引率が必要な教科", options=SUBJECTS, key=key_y)
                st.session_state["small_group_classes"][sg_name]["students"][i]["attributes"]["yellow"] = st.session_state[key_y]

            st.markdown("---")

            # ── 時間割カラーグリッド ──────────────────────────
            st.markdown("**🗓️ 時間割（色設定）**")
            st.caption("各時限に🔴🔵🟡を設定してください")

            periods_per_day = st.session_state.get("periods_per_day", {d: 6 for d in DAYS})

            # sname の student_timetable を初期化
            sg_tt = st.session_state["small_group_classes"][sg_name].setdefault("student_timetable", {})
            if sname not in sg_tt:
                sg_tt[sname] = {}

            day_tabs = st.tabs(DAYS)
            for day in DAYS:
                with day_tabs[DAYS.index(day)]:
                    max_p = periods_per_day.get(day, 6)
                    if day not in sg_tt[sname]:
                        sg_tt[sname][day] = {}

                    header_cols = st.columns(max_p)
                    for p in range(1, max_p + 1):
                        with header_cols[p - 1]:
                            st.markdown(f"**{p}限**")

                    val_cols = st.columns(max_p)
                    for p in range(1, max_p + 1):
                        with val_cols[p - 1]:
                            key_cell = f"sgt_{sg_name}_{i}_{day}_{p}"
                            current = sg_tt[sname][day].get(p, "")
                            options = list(COLOR_OPTIONS.keys())
                            idx = options.index(current) if current in options else 0
                            st.selectbox(
                                f"{p}限",
                                options=options,
                                format_func=lambda x: COLOR_OPTIONS[x],
                                index=idx,
                                key=key_cell,
                                label_visibility="collapsed"
                            )
                            sg_tt[sname][day][p] = st.session_state[key_cell]

            st.markdown("---")
            if st.button(f"🗑️ {sname} を削除", key=f"del_student_{sg_name}_{i}"):
                st.session_state["small_group_classes"][sg_name]["students"].pop(i)
                st.rerun()


def render_sg_subjects(sg_name: str, sg_data: dict):
    """教科設定タブ"""
    teachers = st.session_state["teachers"]
    subject_settings = sg_data.get("subject_settings", {})

    st.markdown("**少人数学級で教える教科の週コマ数・担当教員を設定**")

    # 赤属性を持つ教科を自動収集
    auto_subjects = set()
    for student in sg_data.get("students", []):
        for subj in student.get("attributes", {}).get("red", []):
            auto_subjects.add(subj)

    if auto_subjects:
        st.caption(f"💡 生徒の🔴属性に設定された教科: {', '.join(sorted(auto_subjects))}")

    for subj in SUBJECTS:
        settings = subject_settings.get(subj, {"weekly": 0, "teacher": ""})
        col1, col2, col3 = st.columns([3, 2, 3])
        with col1:
            st.markdown(f"**{subj}**")
        with col2:
            key_w = f"sg_subj_w_{sg_name}_{subj}"
            if key_w not in st.session_state:
                st.session_state[key_w] = settings.get("weekly", 0)
            st.number_input("週コマ数", min_value=0, max_value=10, step=1, key=key_w, label_visibility="collapsed")
            st.session_state["small_group_classes"][sg_name]["subject_settings"].setdefault(subj, {})["weekly"] = st.session_state[key_w]
        with col3:
            key_t = f"sg_subj_t_{sg_name}_{subj}"
            if key_t not in st.session_state:
                st.session_state[key_t] = settings.get("teacher", "")
            options = [""] + teachers
            idx = options.index(st.session_state[key_t]) if st.session_state[key_t] in options else 0
            st.selectbox("担当教員", options=options, index=idx, key=key_t, label_visibility="collapsed")
            st.session_state["small_group_classes"][sg_name]["subject_settings"].setdefault(subj, {})["teacher"] = st.session_state[key_t]


def render_sg_timetable(sg_name: str, sg_data: dict):
    """手動時間割タブ：在籍クラスへの制約プレビューのみ（色入力は生徒登録タブ）"""
    students = sg_data.get("students", [])

    if not students:
        st.info("先に「生徒登録」タブで生徒を登録してください")
        return

    st.caption("「生徒登録」タブの各生徒に設定した色が、在籍クラスの時間割制約として反映されます")

    constraints = get_color_constraints(st.session_state["small_group_classes"])
    home_classes = set(s["home_class"] for s in students)

    COLOR_BG    = {"red": "#FFD0D0", "blue": "#D0E8FF", "yellow": "#FFF5CC", "mixed": "#D0F0FF"}
    COLOR_LABEL = {"red": "🔴少人数", "blue": "🔵通常", "yellow": "🟡引率", "mixed": "🔵🟡通常"}

    if not any(hcls in constraints for hcls in home_classes):
        st.info("まだ色が設定されていません。「生徒登録」タブの各生徒の時間割で🔴🔵🟡を設定してください。")
        return

    for hcls in sorted(home_classes):
        if hcls not in constraints:
            continue
        st.markdown(f"#### 🏫 {hcls} の制約")
        for day in DAYS:
            day_constraints = constraints.get(hcls, {}).get(day, {})
            if not day_constraints:
                continue
            max_p = st.session_state.get("periods_per_day", {}).get(day, 6)
            st.markdown(f"**{day}曜日**")
            cols = st.columns(max_p)
            for p in range(1, max_p + 1):
                with cols[p - 1]:
                    info = day_constraints.get(p)
                    if info:
                        color = info["color"]
                        label = COLOR_LABEL.get(color, color)
                        bg    = COLOR_BG.get(color, "#FFFFFF")
                        allowed_str = "・".join(sorted(info["allowed"])) if info["allowed"] else "制限なし"
                        st.markdown(
                            f'<div style="background:{bg};border-radius:4px;padding:4px;'
                            f'font-size:0.8em;text-align:center;">'
                            f'<b>{p}限</b><br>{label}<br><span style="font-size:0.85em">{allowed_str}</span></div>',
                            unsafe_allow_html=True
                        )
                    else:
                        st.markdown(
                            f'<div style="border:1px solid #ddd;border-radius:4px;padding:4px;'
                            f'font-size:0.8em;text-align:center;color:#aaa;">{p}限<br>制限なし</div>',
                            unsafe_allow_html=True
                        )

def render_step9():
    st.header("STEP 9｜少人数学級設定")
    st.caption("少人数学級を作成し、生徒・教科・手動時間割を設定します")

    # 少人数学級の追加
    col1, col2 = st.columns([3, 1])
    with col1:
        new_sg = st.text_input("少人数学級名を入力", placeholder="例: 少人数A", key="new_sg_name_input", label_visibility="collapsed")
    with col2:
        if st.button("➕ 追加", key="add_sg", use_container_width=True):
            name = new_sg.strip()
            if name and name not in st.session_state["small_group_classes"]:
                st.session_state["small_group_classes"][name] = {
                    "students": [],
                    "subject_settings": {},
                    "student_timetable": {}
                }
                st.rerun()
            elif name in st.session_state["small_group_classes"]:
                st.warning("同じ名前がすでに存在します")

    if not st.session_state["small_group_classes"]:
        st.info("少人数学級が登録されていません")
        return

    # 各少人数学級をexpanderで表示
    for sg_name in list(st.session_state["small_group_classes"].keys()):
        sg_data = st.session_state["small_group_classes"][sg_name]
        with st.expander(f"📗 {sg_name}", expanded=True):
            inner_tabs = st.tabs(["👤 生徒登録", "📚 教科設定", "🗓️ 手動時間割"])

            # ── タブ1: 生徒登録 ──────────────────────────────
            with inner_tabs[0]:
                render_sg_students(sg_name, sg_data)

            # ── タブ2: 教科設定 ──────────────────────────────
            with inner_tabs[1]:
                render_sg_subjects(sg_name, sg_data)

            # ── タブ3: 手動時間割 ─────────────────────────────
            with inner_tabs[2]:
                render_sg_timetable(sg_name, sg_data)

            # 削除ボタン
            st.markdown("---")
            if st.button(f"🗑️ {sg_name} を削除", key=f"del_sg_{sg_name}"):
                del st.session_state["small_group_classes"][sg_name]
                st.rerun()


# ============================================================
# STEP 10: 学年間教科同期設定
# ============================================================

def render_step10():
    st.header("STEP 10｜学年間教科同期設定")
    st.caption(
        "異なる学年の教科を同じ曜日・同じ時限に揃えるハード制約を設定します。\n\n"
        "例：「5年 家庭」と「6年 音楽」を同期 → 5年の各クラスが家庭の時限に、"
        "6年のいずれかのクラスが音楽を担当する時限が一致します。"
    )

    pairs = st.session_state["cross_grade_sync"]

    # ── 新規ペア追加 ───────────────────────────────────────
    st.markdown("#### 同期ペアを追加")
    col1, col2, col3, col4, col5 = st.columns([2, 3, 2, 3, 2])
    with col1:
        new_g1 = st.selectbox("学年①", GRADES, key="cgs_g1")
    with col2:
        new_s1 = st.selectbox("教科①", SUBJECTS, key="cgs_s1")
    with col3:
        new_g2 = st.selectbox("学年②", GRADES, key="cgs_g2")
    with col4:
        new_s2 = st.selectbox("教科②", SUBJECTS, key="cgs_s2")
    with col5:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕ 追加", key="cgs_add", use_container_width=True):
            if new_g1 == new_g2 and new_s1 == new_s2:
                st.warning("同じ学年・同じ教科のペアは追加できません")
            else:
                new_pair = {"grade1": new_g1, "subject1": new_s1,
                            "grade2": new_g2, "subject2": new_s2}
                # 重複チェック（順序問わず）
                reverse = {"grade1": new_g2, "subject1": new_s2,
                           "grade2": new_g1, "subject2": new_s1}
                if new_pair in pairs or reverse in pairs:
                    st.warning("同じ組み合わせがすでに登録されています")
                else:
                    pairs.append(new_pair)
                    st.rerun()

    st.markdown("---")

    # ── 登録済みペア一覧 ───────────────────────────────────
    if not pairs:
        st.info("同期ペアが登録されていません")
        return

    st.markdown("#### 登録済み同期ペア")
    for i, pair in enumerate(pairs):
        col_a, col_b = st.columns([6, 1])
        with col_a:
            st.markdown(
                f"**{i+1}.** {pair['grade1']}年「{pair['subject1']}」　↔　"
                f"{pair['grade2']}年「{pair['subject2']}」"
            )
        with col_b:
            if st.button("🗑️", key=f"cgs_del_{i}", help="削除"):
                pairs.pop(i)
                st.rerun()


# ============================================================
# Part 3: ソフト制約設定UI
# ============================================================

def render_soft_constraints():
    """
    ソフト制約（S1学年集約・S2優先教科）の設定UIを表示する。
    render_generate()内の生成ボタン上部に配置する。
    """
    st.markdown("### 🎛️ ソフト制約設定")

    st.session_state["soft_grade_grouping"] = st.checkbox(
        "S1: 学年集約を有効にする",
        value=st.session_state.get("soft_grade_grouping", True),
        help=(
            "同一教員が同じ曜日に複数クラスを担当する場合、"
            "できるだけ同じ学年の授業が連続するように調整します"
        )
    )

    if st.session_state["soft_grade_grouping"]:
        st.session_state["soft_priority_subjects"] = st.multiselect(
            "S2: 優先教科（学年集約を特に強く適用する教科）",
            options=SUBJECTS,
            default=st.session_state.get("soft_priority_subjects", []),
            help=(
                "選択した教科は学年集約のペナルティが10倍になります。"
                "例: 理科・音楽・保体など準備が必要な教科を選択してください"
            )
        )

    st.markdown("---")
    st.markdown("#### 📚 教科別段階生成")
    st.session_state["soft_first_subjects"] = st.multiselect(
        "先に配置する教科（Phase 1）",
        options=SUBJECTS,
        default=st.session_state.get("soft_first_subjects", []),
        help=(
            "選択した教科を Phase 1 として先に全クラスに配置し、"
            "その後に残りの教科（Phase 2）を配置します。\n"
            "「📚 教科別段階生成」ボタンで実行してください。"
        )
    )


# ============================================================
# Part 3: バリデーション関数
# ============================================================

def validate_all_settings() -> list[str]:
    """
    生成前に全設定の整合性チェックを行い、警告メッセージのリストを返す。
    リストが空なら問題なし。
    """
    warnings_list = []
    classes        = get_all_classes()
    total_per_week = get_total_periods_per_week()

    # ① 教員未登録チェック
    if not st.session_state["teachers"]:
        warnings_list.append("教員が1人も登録されていません（STEP 2）")

    # ② 週コマ数の合計チェック
    for cls in classes:
        total = sum(st.session_state["weekly_periods"].get(cls, {}).values())
        if total != total_per_week:
            warnings_list.append(
                f"{cls}: 週コマ数の合計が {total} コマ "
                f"（週上限 {total_per_week} コマと不一致）"
            )

    # ③ 担当教員未設定チェック（週コマ数>0なのに担当なし）
    # 合成教科は両方の基教科に担当がいれば自動設定されるのでスキップ
    for cls in classes:
        for subj in SUBJECTS:
            weekly = st.session_state["weekly_periods"].get(cls, {}).get(subj, 0)
            if weekly > 0:
                teachers = st.session_state["assignments"].get(
                    cls, {}).get(subj, [])
                if not teachers:
                    warnings_list.append(
                        f"{cls} / {subj}: 週{weekly}コマ設定だが担当教員が未設定"
                    )

    # ④ 特別教室の教科名整合性チェック
    for room in st.session_state["special_rooms"]:
        gs = room.get("grade_subjects", [])
        if not gs:
            warnings_list.append(
                f"特別教室「{room.get('name', '?')}」の学年×教科が未設定です"
            )
        else:
            for g, s in gs:
                if s not in SUBJECTS:
                    warnings_list.append(
                        f"特別教室「{room.get('name', '?')}」の対応教科「{s}」が不正です"
                    )

    # ⑤ 少人数学級の整合性チェック
    for sg_name, sg_data in st.session_state["small_group_classes"].items():
        students = sg_data.get("students", [])
        subject_settings = sg_data.get("subject_settings", {})
        all_teachers = st.session_state["teachers"]
        classes = get_all_classes()

        for i, student in enumerate(students):
            # 在籍クラスチェック
            if student.get("home_class") not in classes:
                warnings_list.append(
                    f"少人数学級「{sg_name}」生徒「{student.get('name')}」: "
                    f"在籍クラス「{student.get('home_class')}」が存在しません"
                )

        # 教科設定の担当教員チェック
        for subj, settings in subject_settings.items():
            weekly = settings.get("weekly", 0)
            teacher = settings.get("teacher", "")
            if weekly > 0 and not teacher:
                warnings_list.append(
                    f"少人数学級「{sg_name}」教科「{subj}」: 週{weekly}コマ設定だが担当教員が未設定"
                )
            if teacher and teacher not in all_teachers:
                warnings_list.append(
                    f"少人数学級「{sg_name}」教科「{subj}」: 担当教員「{teacher}」が教員リストにありません"
                )

    # ⑥ R7事前チェック（教員の週総授業コマ数が上限内か）
    for teacher in st.session_state["teachers"]:
        weekly_load = sum(
            st.session_state["weekly_periods"].get(cls, {}).get(subj, 0)
            for cls in classes
            for subj, tlist in st.session_state["assignments"].get(cls, {}).items()
            if teacher in tlist
        )
        if weekly_load == 0:
            continue
        max_weekly = sum(
            max(0, st.session_state["periods_per_day"].get(day, 6)
                - len(st.session_state["unavailable"].get(teacher, {}).get(day, []))
                - 1)
            for day in DAYS
        )
        if weekly_load > max_weekly:
            warnings_list.append(
                f"教員「{teacher}」: 週{weekly_load}コマ必要ですが"
                f"R7制約後の上限は{max_weekly}コマです（担当クラスを減らすか不在設定を見直してください）"
            )

    # ⑦ 手動コマのR1・R5チェック
    manual_tt = st.session_state.get("manual_timetable", {})
    if manual_tt:
        for day in DAYS:
            total_p = st.session_state["periods_per_day"].get(day, 6)
            for p in range(1, total_p + 1):
                seen_t: dict = {}
                for cls in classes:
                    subj = manual_tt.get(cls, {}).get(day, {}).get(p)
                    if not subj:
                        continue
                    # R5: 同日同教科
                    day_map = manual_tt.get(cls, {}).get(day, {})
                    same = [s for pk, s in day_map.items() if pk != p and s == subj]
                    if same:
                        warnings_list.append(
                            f"手動コマR5違反: {cls} {day}曜に「{subj}」が複数コマ設定されています")
                    # R1: 教員重複
                    for t in st.session_state["assignments"].get(cls, {}).get(subj, []):
                        if t in seen_t:
                            warnings_list.append(
                                f"手動コマR1違反: {t}先生が{day}曜{p}限に"
                                f"{seen_t[t]}と{cls}の両方に設定されています")
                        seen_t[t] = cls
                        # ⑧ 手動コマと不在コマの競合チェック（R4）
                        if p in st.session_state["unavailable"].get(t, {}).get(day, []):
                            warnings_list.append(
                                f"手動コマと不在コマが競合: {t}先生は{day}曜{p}限が不在設定ですが、"
                                f"{cls}の「{subj}」が手動配置されています（時間割を生成できません）"
                            )

    return warnings_list


# ============================================================
# Part 4: 制約チェック関数
# ============================================================

def is_valid_placement(
    timetable: dict,
    timetable_sg: dict,
    cls: str,
    day: str,
    period: int,
    subject: str,
    assignments: dict,
    special_rooms: list,
    unavailable: dict,
    periods_per_day: dict,
    small_group_classes: dict,
    weekly_remaining: dict
) -> bool:
    """R1〜R8を全チェックする。すべてパスしたらTrue。"""

    # ── 学年別コマ数チェック ─────────────────────────────────
    if period > get_periods_for_class(cls, day):
        return False

    teachers = assignments.get(cls, {}).get(subject, [])

    # ── R1: 教員重複禁止 ─────────────────────────────────────
    for t in teachers:
        for other_cls, schedule in timetable.items():
            if other_cls == cls:
                continue
            other_subj = schedule.get(day, {}).get(period)
            if other_subj:
                if t in assignments.get(other_cls, {}).get(other_subj, []):
                    return False
        # 少人数学級の教科設定に登録された担当教員との重複チェック
        for sg_name, sg_data in small_group_classes.items():
            subj_settings = sg_data.get("subject_settings", {})
            student_timetable = sg_data.get("student_timetable", {})
            # 当該スロットに少人数授業（red）がある場合、その教科の担当教員はbusy
            for student in sg_data.get("students", []):
                sname = student["name"]
                color = student_timetable.get(sname, {}).get(day, {}).get(period, "")
                if color == "red":
                    for red_subj in student.get("attributes", {}).get("red", []):
                        sg_teacher = subj_settings.get(red_subj, {}).get("teacher", "")
                        if sg_teacher and t == sg_teacher:
                            return False

    # ── R3: 特別教室同時使用制限 ─────────────────────────────
    cls_grade = get_grade(cls)
    for room in special_rooms:
        gs = room.get("grade_subjects", [])
        if [cls_grade, subject] not in gs and (cls_grade, subject) not in gs:
            continue
        count = sum(
            1 for other_cls, schedule in timetable.items()
            if other_cls != cls
            and ([get_grade(other_cls), schedule.get(day, {}).get(period)] in gs
                 or (get_grade(other_cls), schedule.get(day, {}).get(period)) in gs)
        )
        if count >= room["capacity"]:
            return False

    # ── R4: 教員不在コマ尊重 ─────────────────────────────────
    for t in teachers:
        if period in unavailable.get(t, {}).get(day, []):
            return False

    # ── R5: 同一クラス・同一曜日に同一教科は1コマまで ───────
    for p, s in timetable.get(cls, {}).get(day, {}).items():
        if p != period and s == subject:
            return False

    # ── R6: 少人数カラー制約（ハード制約・緩和なし） ────────────────────────────────
    # NOTE: この関数(is_valid_placement)は現在どこからも呼ばれていない。
    #       実際のチェックは solve_ortools 内の制約として実装されている。
    # 赤スロットには許容教科以外は絶対に入れない。許容教科が0コマでも緩和しない。
    color_constraints = get_color_constraints(small_group_classes)
    if cls in color_constraints:
        slot_info = color_constraints.get(cls, {}).get(day, {}).get(period)
        _is_red_slot = bool(slot_info and slot_info.get("color") == "red")
        if slot_info:
            allowed = slot_info["allowed"]
            if allowed and subject not in allowed:
                return False  # 非許容教科はカラー制約スロットに入れない（無条件ハード）
        # 非赤スロット（blue/yellow/mixed/制約なし）：赤スロット向け教科の「総供給」予約チェック
        # fast_valid と同じロジック：slot_infoの有無によらず非赤スロットに対して適用
        if not _is_red_slot:
            _cls_cc = color_constraints[cls]
            _all_red_allowed_v = set()
            for _dm in _cls_cc.values():
                for _si in _dm.values():
                    if _si.get("color") == "red":
                        _all_red_allowed_v |= _si.get("allowed", set())
            if subject in _all_red_allowed_v:
                _unfilled_red_v = sum(
                    1
                    for _d2, _pm2 in _cls_cc.items()
                    for _p2, _si2 in _pm2.items()
                    if _si2.get("color") == "red" and timetable.get(cls, {}).get(_d2, {}).get(_p2) is None
                )
                if _unfilled_red_v > 0:
                    _total_allowed_rem_v = sum(
                        weekly_remaining.get(cls, {}).get(_s, 0)
                        for _s in _all_red_allowed_v
                    )
                    if _total_allowed_rem_v - 1 < _unfilled_red_v:
                        return False

    # ── R8: 学年間教科同期制約（一方向・少ない方が基準） ────────
    # コマ数の少ない側のすべてのコマは、多い側のクラスと必ず同期する。
    # 多い側の余りコマは、少ない側が存在しない時限に自由配置。
    cross_grade_sync = st.session_state.get("cross_grade_sync", [])
    cls_grade = get_grade(cls)
    all_classes = get_all_classes()
    for pair in cross_grade_sync:
        g1, s1, g2, s2 = pair["grade1"], pair["subject1"], pair["grade2"], pair["subject2"]
        grade1_classes = [c for c in all_classes if get_grade(c) == g1]
        grade2_classes = [c for c in all_classes if get_grade(c) == g2]

        # 残コマ数で少ない方（minority）を動的に判定
        total_g1 = sum(weekly_remaining.get(c, {}).get(s1, 0) for c in grade1_classes)
        total_g2 = sum(weekly_remaining.get(c, {}).get(s2, 0) for c in grade2_classes)

        if total_g1 <= total_g2:
            # g1(s1)がminority: s1を配置するなら同時限にg2のs2が必要
            minority_grade, minority_subj = g1, s1
            majority_grade, majority_subj = g2, s2
            minority_classes, majority_classes = grade1_classes, grade2_classes
        else:
            # g2(s2)がminority: s2を配置するなら同時限にg1のs1が必要
            minority_grade, minority_subj = g2, s2
            majority_grade, majority_subj = g1, s1
            minority_classes, majority_classes = grade2_classes, grade1_classes

        # minority側を配置するとき: majority側が全て別教科で埋まっていたらNG
        if cls_grade == minority_grade and subject == minority_subj:
            filled_maj = [
                timetable.get(c, {}).get(day, {}).get(period)
                for c in majority_classes
            ]
            filled_maj_vals = [s for s in filled_maj if s is not None]
            if len(filled_maj_vals) == len(majority_classes) and \
                    not any(s == majority_subj for s in filled_maj_vals):
                return False

        # majority側がminority_subjを持っているとき:
        # minority側を別教科で埋めてしまい、minority側が全て埋まったらNG
        if cls_grade == minority_grade and subject != minority_subj:
            has_maj_subj = any(
                timetable.get(c, {}).get(day, {}).get(period) == majority_subj
                for c in majority_classes
            )
            if has_maj_subj:
                other_min = [c for c in minority_classes if c != cls]
                filled_other = [
                    timetable.get(c, {}).get(day, {}).get(period)
                    for c in other_min
                ]
                if all(s is not None for s in filled_other) and \
                        not any(s == minority_subj for s in filled_other):
                    return False

    # ── R7: 各教員の当日最低1コマ空き ───────────────────────
    r7_exempt = st.session_state.get("r7_exempt_teachers", [])
    total_periods_today = periods_per_day.get(day, 6)
    for t in teachers:
        if t in r7_exempt:
            continue  # R7免除教員はスキップ
        assigned_today = sum(
            1 for oc, sc in timetable.items()
            for p_ in range(1, total_periods_today + 1)
            if sc.get(day, {}).get(p_) and
               t in assignments.get(oc, {}).get(sc[day].get(p_, ""), [])
        )
        assigned_today += sum(
            1 for sg_name, sg_sc in timetable_sg.items()
            for p_ in range(1, total_periods_today + 1)
            if sg_sc.get(day, {}).get(p_) and
               t in get_sg_teachers(sg_name, small_group_classes)
        )
        unavail_today = len(unavailable.get(t, {}).get(day, []))
        max_assignable = total_periods_today - unavail_today - 1
        if assigned_today >= max_assignable:
            return False

    return True




# ============================================================
# Part 4: 前処理関数
# ============================================================

def preprocess(
    timetable: dict,
    timetable_sg: dict,
    manual_timetable: dict,
    manual_timetable_sg: dict,
    weekly_periods: dict,
    weekly_periods_sg: dict
) -> tuple[dict, dict]:
    """
    手動入力コマをロックし、残り週コマ数を計算して返す。
    戻り値: (weekly_remaining, weekly_remaining_sg)
    """
    # 通常学級の残り週コマ数を初期化
    weekly_remaining = {
        cls: dict(weekly_periods.get(cls, {}))
        for cls in timetable
    }

    # 0.5刻み合成教科の処理
    # 小数部(.5)のある教科をペアリングして合成教科を追加し、基教科は切り捨て
    for cls in list(weekly_remaining.keys()):
        cls_wp = weekly_remaining[cls]
        combined_list = compute_combined_subjects(cls_wp)
        # 端数のある基教科を切り捨て（1.5 → 1、0.5 → 0）
        for s in SUBJECTS:
            if s in cls_wp and round(cls_wp[s] * 2) % 2 == 1:
                cls_wp[s] = int(cls_wp[s])
        # 合成教科を追加（1ペアにつき1コマ）
        for c in combined_list:
            cls_wp[c["name"]] = 1

    # 手動入力コマをロック
    for cls, day_map in manual_timetable.items():
        if cls not in timetable:
            continue
        for day, period_map in day_map.items():
            for period, subj in period_map.items():
                timetable[cls].setdefault(day, {})[period] = subj
                if subj in weekly_remaining.get(cls, {}):
                    weekly_remaining[cls][subj] = max(
                        0, weekly_remaining[cls][subj] - 1
                    )

    # 少人数学級の残り週コマ数を初期化
    weekly_remaining_sg = {
        sg: dict(weekly_periods_sg.get(sg, {}))
        for sg in timetable_sg
    }

    # 少人数手動入力コマをロック
    for sg_name, day_map in manual_timetable_sg.items():
        if sg_name not in timetable_sg:
            continue
        for day, period_map in day_map.items():
            for period, subj in period_map.items():
                timetable_sg[sg_name].setdefault(day, {})[period] = subj
                if subj in weekly_remaining_sg.get(sg_name, {}):
                    weekly_remaining_sg[sg_name][subj] = max(
                        0, weekly_remaining_sg[sg_name][subj] - 1
                    )

    return weekly_remaining, weekly_remaining_sg


# ============================================================
# Part 4: OR-Tools CP-SAT ソルバー
# ============================================================

def solve_ortools(
    timetable: dict,
    timetable_sg: dict,
    weekly_remaining: dict,
    assignments: dict,
    special_rooms: list,
    unavailable: dict,
    periods_per_day: dict,
    small_group_classes: dict,
    priority_subjects: list,
    random_seed: int = 0,
    time_limit: float = 60.0
) -> str:
    """
    Google OR-Tools CP-SAT で時間割を生成する。
    time_limit: ソルバーの最大実行秒数（デフォルト60秒）
    戻り値: "solved" / "failed"
    """
    try:
        from ortools.sat.python import cp_model

        model   = cp_model.CpModel()
        classes = list(timetable.keys())
        x       = {}  # x[cls][day][p][subj] ∈ {0,1}

        # クラスごとの教科リスト（合成教科を含む）
        # weekly_remaining にあって SUBJECTS にない教科 = 合成教科
        _cls_subjects = {
            cls: SUBJECTS + [
                s for s in weekly_remaining.get(cls, {})
                if s not in SUBJECTS and weekly_remaining[cls].get(s, 0) > 0
            ]
            for cls in classes
        }

        # ── 変数定義 ─────────────────────────────────────────
        for cls in classes:
            x[cls] = {}
            for day in DAYS:
                x[cls][day] = {}
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    x[cls][day][p] = {
                        subj: model.NewBoolVar(f"x_{cls}_{day}_{p}_{subj}")
                        for subj in _cls_subjects[cls]
                    }

        # ── 手動ロック済みコマを固定 ─────────────────────────
        for cls in classes:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    locked = timetable[cls].get(day, {}).get(p)
                    if locked:
                        model.Add(x[cls][day][p][locked] == 1)
                        for subj in _cls_subjects[cls]:
                            if subj != locked:
                                model.Add(x[cls][day][p][subj] == 0)

        # ── 基本制約: 各コマに高々1教科 ─────────────────────
        for cls in classes:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    model.AddAtMostOne(x[cls][day][p][s] for s in SUBJECTS)
                    # 学年別コマ数を超えるスロットは全て0固定
                    if p > get_periods_for_class(cls, day):
                        for s in SUBJECTS:
                            model.Add(x[cls][day][p][s] == 0)

        # ── R1: 教員重複禁止 ─────────────────────────────────
        for t in st.session_state["teachers"]:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    tvars = [
                        x[cls][day][p][subj]
                        for cls in classes
                        for subj in _cls_subjects[cls]
                        if t in assignments.get(cls, {}).get(subj, [])
                    ]
                    if tvars:
                        model.AddAtMostOne(tvars)

        # ── 担任授業制約 ──────────────────────────────────────
        # 空きコマは担任が自クラスで授業するため、担任が他クラスに入るコマは
        # 必ず担任クラスに「担任以外の教員」の教科が配置されていなければならない。
        # 制約: 担任T が他クラスC2 の(day,p)に配置
        #      → 担任クラスC の(day,p)に T 以外の教員が担当する教科あり
        # ※ T 自身の教科は R1（教員重複禁止）で自動排除されるため hr_has_subj に含めない
        # ※ T が不在コマの場合、空きコマは「担任在室」ではないため制約不要
        _ort_homeroom = st.session_state.get("homeroom_teachers", {})
        _ort_homeroom_rev = {t: c for c, t in _ort_homeroom.items() if t}
        for t, hr_cls in _ort_homeroom_rev.items():
            if hr_cls not in classes:
                continue
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    # T が不在コマ → 担任在室の前提が崩れるため制約不要
                    if p in unavailable.get(t, {}).get(day, []):
                        continue
                    # 担任クラスのスロットがロック済みなら制約不要
                    hr_locked = timetable[hr_cls].get(day, {}).get(p)
                    # 担任が他クラスに入る変数の合計
                    t_in_other = [
                        x[cls2][day][p][subj]
                        for cls2 in classes
                        if cls2 != hr_cls
                        and not timetable[cls2].get(day, {}).get(p)
                        for subj in _cls_subjects[cls2]
                        if t in assignments.get(cls2, {}).get(subj, [])
                    ]
                    if not t_in_other:
                        continue
                    if hr_locked:
                        # ロック済みなら常に充足（制約追加不要）
                        continue
                    # 担任クラスに「T 以外の教員」の教科が入る変数の合計
                    # T 自身の教科を含めると R1 と矛盾（T は両クラス同時不可）するため除外
                    hr_has_subj = [
                        x[hr_cls][day][p][subj]
                        for subj in _cls_subjects[hr_cls]
                        if t not in assignments.get(hr_cls, {}).get(subj, [])
                    ]
                    # t_in_other の合計 <= hr_has_subj の合計
                    # （担任が他クラスにいるなら、担任クラスには別の教員の授業があること）
                    model.Add(sum(t_in_other) <= sum(hr_has_subj))

        # ── 担任不在コマの担任クラス強制充填 ────────────────────
        # 担任T が不在の (day,p) は「空きコマ=担任在室」の前提が成立しない。
        # 担任クラスに空きコマが残ると「担任が授業している」と誤解されるため、
        # T 以外の教員の教科を必ず配置してコマを埋める。
        # ただし、T 以外の教科が hr_cls に存在しない（または全消化済み）なら
        # 強制せず、自由スロットとして扱う（生成後の表示で「担任不在」と示す）。
        for t, hr_cls in _ort_homeroom_rev.items():
            if hr_cls not in classes:
                continue
            for day in DAYS:
                for p in unavailable.get(t, {}).get(day, []):
                    if p > periods_per_day.get(day, 6):
                        continue
                    if timetable[hr_cls].get(day, {}).get(p):
                        continue  # 手動ロック済みはスキップ
                    # T 以外の教員が担当する教科の x 変数
                    non_t_vars = [
                        x[hr_cls][day][p][subj]
                        for subj in _cls_subjects[hr_cls]
                        if t not in assignments.get(hr_cls, {}).get(subj, [])
                        and weekly_remaining.get(hr_cls, {}).get(subj, 0) > 0
                    ]
                    if non_t_vars:
                        # T 以外の教科が存在する場合のみ強制（実現可能な場合のみ）
                        model.Add(sum(non_t_vars) >= 1)

        # ── R2: 週コマ数完全消化 ─────────────────────────────
        for cls in classes:
            for subj in _cls_subjects[cls]:
                target = int(weekly_remaining.get(cls, {}).get(subj, 0))
                pvars = [
                    x[cls][day][p][subj]
                    for day in DAYS
                    for p in range(1, periods_per_day.get(day, 6) + 1)
                    if not timetable[cls].get(day, {}).get(p)
                ]
                if pvars:
                    model.Add(sum(pvars) == target)

        # ── R3: 特別教室同時使用制限 ─────────────────────────
        for room in special_rooms:
            cap = room["capacity"]
            gs  = room.get("grade_subjects", [])
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    room_vars = [
                        x[c][day][p][s]
                        for c in classes
                        for g, s in gs
                        if get_grade(c) == g and c in x
                        and not timetable[c].get(day, {}).get(p)
                    ]
                    # 手動配置済みも合算
                    manual_count = sum(
                        1 for c in classes
                        for g, s in gs
                        if get_grade(c) == g and timetable[c].get(day, {}).get(p) == s
                    )
                    if room_vars:
                        model.Add(sum(room_vars) + manual_count <= cap)

        # ── R4: 教員不在コマ ─────────────────────────────────
        for t in st.session_state["teachers"]:
            for day in DAYS:
                for p in unavailable.get(t, {}).get(day, []):
                    if p > periods_per_day.get(day, 6):
                        continue
                    for cls in classes:
                        for subj in _cls_subjects[cls]:
                            if t in assignments.get(cls, {}).get(subj, []):
                                model.Add(x[cls][day][p][subj] == 0)

        # ── R5: 同曜日・同教科1コマまで ─────────────────────
        for cls in classes:
            for day in DAYS:
                for subj in _cls_subjects[cls]:
                    model.Add(sum(
                        x[cls][day][p][subj]
                        for p in range(1, periods_per_day.get(day, 6) + 1)
                    ) <= 1)

        # ── R8: 学年間教科同期制約（一方向・minority→majority） ──
        # コマ数の少ない側(minority)がある時限には必ず多い側(majority)も存在する。
        # 多い側の余りコマは少ない側が存在しない時限に自由配置。
        cross_grade_sync = st.session_state.get("cross_grade_sync", [])
        all_classes_list = get_all_classes()
        for pair in cross_grade_sync:
            g1, s1, g2, s2 = pair["grade1"], pair["subject1"], pair["grade2"], pair["subject2"]
            grade1_cls = [c for c in all_classes_list if get_grade(c) == g1 and c in classes]
            grade2_cls = [c for c in all_classes_list if get_grade(c) == g2 and c in classes]
            if not grade1_cls or not grade2_cls:
                continue

            # 残コマ合計でminority/majorityを決定
            total_g1 = sum(weekly_remaining.get(c, {}).get(s1, 0) for c in grade1_cls)
            total_g2 = sum(weekly_remaining.get(c, {}).get(s2, 0) for c in grade2_cls)
            if total_g1 <= total_g2:
                min_cls, min_subj = grade1_cls, s1
                maj_cls, maj_subj = grade2_cls, s2
            else:
                min_cls, min_subj = grade2_cls, s2
                maj_cls, maj_subj = grade1_cls, s1

            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    # has_min: minority側のいずれかがmin_subjを持つ
                    has_min = model.NewBoolVar(
                        f"r8_min_{g1}_{s1}_{g2}_{s2}_{day}_{p}")
                    sum_min = sum(x[c][day][p][min_subj] for c in min_cls
                                  if not timetable[c].get(day, {}).get(p))
                    manual_min = sum(
                        1 for c in min_cls
                        if timetable[c].get(day, {}).get(p) == min_subj
                    )
                    model.Add(sum_min + manual_min >= 1).OnlyEnforceIf(has_min)
                    model.Add(sum_min + manual_min == 0).OnlyEnforceIf(has_min.Not())

                    # has_maj: majority側のいずれかがmaj_subjを持つ
                    has_maj = model.NewBoolVar(
                        f"r8_maj_{g1}_{s1}_{g2}_{s2}_{day}_{p}")
                    sum_maj = sum(x[c][day][p][maj_subj] for c in maj_cls
                                  if not timetable[c].get(day, {}).get(p))
                    manual_maj = sum(
                        1 for c in maj_cls
                        if timetable[c].get(day, {}).get(p) == maj_subj
                    )
                    model.Add(sum_maj + manual_maj >= 1).OnlyEnforceIf(has_maj)
                    model.Add(sum_maj + manual_maj == 0).OnlyEnforceIf(has_maj.Not())

                    # has_min → has_maj（一方向: minorityがあればmajorityも必要）
                    model.Add(has_maj >= has_min)

        # ── 目的関数ペナルティ項（R6ソフト＋S1学年集約） ────────
        penalty_terms = []

        # ── R6: 少人数カラー制約（ハード） ──────────────────────
        # 生徒ごとの色設定から在籍クラスの各スロットに入れられる教科を制限する。
        # 🔴スロット → 赤属性教科のみ許可
        # 🔵🟡スロット → 青+黄属性教科のみ許可
        _ort_color_constraints = get_color_constraints(small_group_classes)
        for cls in classes:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    if timetable[cls].get(day, {}).get(p):
                        continue  # 手動ロック済みは除外
                    slot_info = _ort_color_constraints.get(cls, {}).get(day, {}).get(p)
                    if slot_info and slot_info["color"] == "red":
                        allowed = slot_info["allowed"]
                        if allowed:
                            # 非許容教科は赤スロット（少人数教室）に絶対に配置しない（ハード制約・緩和なし）
                            # 青・黄（通常在籍）スロットは制約なし: 生徒は通常クラスにいるため教科制限不要
                            for subj in _cls_subjects[cls]:
                                if subj not in allowed:
                                    model.Add(x[cls][day][p][subj] == 0)

        # ── R6-supply: 赤スロット強制配置（教科ごと供給予約） ──────
        # 教科 subj の残りコマ数 ≤ その教科が入れる赤スロット数 の場合、
        # その教科を非赤スロットに配置するのを禁止する。
        _ort_all_red_cls: dict[str, set] = {}
        for _c, _dm in _ort_color_constraints.items():
            _u: set = set()
            for _dv in _dm.values():
                for _si in _dv.values():
                    if _si.get("color") == "red":
                        _u |= _si.get("allowed", set())
            if _u:
                _ort_all_red_cls[_c] = _u

        for cls in classes:
            if cls not in _ort_all_red_cls:
                continue
            for subj in _ort_all_red_cls[cls]:
                subj_rem = weekly_remaining.get(cls, {}).get(subj, 0)
                if subj_rem <= 0:
                    continue
                # この教科が入れる未ロック赤スロット数を数える
                avail_red = sum(
                    1
                    for day in DAYS
                    for p in range(1, periods_per_day.get(day, 6) + 1)
                    if not timetable[cls].get(day, {}).get(p)
                    and _ort_color_constraints.get(cls, {}).get(day, {}).get(p, {}).get("color") == "red"
                    and subj in _ort_color_constraints[cls][day][p].get("allowed", set())
                )
                if avail_red >= subj_rem:
                    # 非赤・未ロックスロットへの配置を禁止
                    for day in DAYS:
                        for p in range(1, periods_per_day.get(day, 6) + 1):
                            if timetable[cls].get(day, {}).get(p):
                                continue  # ロック済みは除外
                            _si = _ort_color_constraints.get(cls, {}).get(day, {}).get(p)
                            if _si and _si.get("color") == "red":
                                continue  # 赤スロット自体は除外
                            model.Add(x[cls][day][p][subj] == 0)

        # ── R6b: 少人数担当教員の赤スロットでの重複禁止 ──────
        # 少人数学級の教科設定に登録された担当教員は、
        # 担当生徒が🔴（少人数教室）の時間帯は通常クラスに配置不可。
        for sg_name, sg_data in small_group_classes.items():
            subj_settings = sg_data.get("subject_settings", {})
            student_timetable = sg_data.get("student_timetable", {})
            for student in sg_data.get("students", []):
                sname = student["name"]
                red_subjects = student.get("attributes", {}).get("red", [])
                for day in DAYS:
                    for p in range(1, periods_per_day.get(day, 6) + 1):
                        color = student_timetable.get(sname, {}).get(day, {}).get(p, "")
                        if color != "red":
                            continue
                        # この時間帯、赤教科の担当教員は通常クラスに入れない
                        for red_subj in red_subjects:
                            sg_teacher = subj_settings.get(red_subj, {}).get("teacher", "")
                            if not sg_teacher:
                                continue
                            for cls in classes:
                                for subj in _cls_subjects[cls]:
                                    if sg_teacher in assignments.get(cls, {}).get(subj, []):
                                        if not timetable[cls].get(day, {}).get(p):
                                            model.Add(x[cls][day][p][subj] == 0)

        # ── R7: 教員1日最低1コマ空き ─────────────────────────
        _ort_r7_exempt = set(st.session_state.get("r7_exempt_teachers", []))
        for t in st.session_state["teachers"]:
            if t in _ort_r7_exempt:
                continue  # R7免除教員はスキップ
            for day in DAYS:
                total_today = periods_per_day.get(day, 6)
                unavail_cnt = len(unavailable.get(t, {}).get(day, []))
                available   = total_today - unavail_cnt   # 出勤可能コマ数
                # 少人数学級の赤スロットで担当している時間数を減算
                sg_cnt = 0
                for sg_name, sg_data in small_group_classes.items():
                    subj_settings = sg_data.get("subject_settings", {})
                    sg_teachers_set = {
                        s["teacher"] for s in subj_settings.values()
                        if s.get("teacher")
                    }
                    if t not in sg_teachers_set:
                        continue
                    # 赤スロット数をカウント
                    red_periods = set()
                    for student in sg_data.get("students", []):
                        sname = student["name"]
                        for p in range(1, total_today + 1):
                            color = sg_data.get("student_timetable", {}).get(
                                sname, {}).get(day, {}).get(p, "")
                            if color == "red":
                                red_periods.add(p)
                    sg_cnt += sum(
                        1 for p in red_periods
                        if p not in unavailable.get(t, {}).get(day, [])
                    )
                max_ok = available - 1 - sg_cnt  # 最低1コマ空き（少人数分を減算）
                # 手動配置済みコマ数を数え、R7制約から除外する
                manual_cnt = sum(
                    1 for cls in classes
                    for p in range(1, total_today + 1)
                    if p not in unavailable.get(t, {}).get(day, [])
                    and t in assignments.get(cls, {}).get(
                        timetable[cls].get(day, {}).get(p, ""), []
                    )
                )
                effective_max_ok = max_ok - manual_cnt
                if effective_max_ok < 0:
                    # 手動配置がすでにR7上限を超えているが変更不可のためスキップ
                    continue
                # 自動配置スロットのみ対象（手動ロック済みは除外）
                tvars = [
                    x[cls][day][p][subj]
                    for cls in classes
                    for p in range(1, total_today + 1)
                    for subj in _cls_subjects[cls]
                    if t in assignments.get(cls, {}).get(subj, [])
                    and p not in unavailable.get(t, {}).get(day, [])
                    and not timetable[cls].get(day, {}).get(p)  # 手動ロック済みを除外
                ]
                if tvars:
                    model.Add(sum(tvars) <= effective_max_ok)

        # ── S1: 学年集約ペナルティ（時限集約 + 曜日集約） ──────────
        if st.session_state.get("soft_grade_grouping", True):
            PW, NW = 10, 1

            # S1a: 同教員・異学年が同日の異なる時限に配置されたらペナルティ
            for t in st.session_state["teachers"]:
                for day in DAYS:
                    dp = periods_per_day.get(day, 6)
                    for p1 in range(1, dp + 1):
                        for p2 in range(p1 + 1, dp + 1):
                            for c1 in classes:
                                for c2 in classes:
                                    if c1 == c2 or get_grade(c1) == get_grade(c2):
                                        continue
                                    for s1 in SUBJECTS:
                                        if t not in assignments.get(c1, {}).get(s1, []):
                                            continue
                                        for s2 in SUBJECTS:
                                            if t not in assignments.get(c2, {}).get(s2, []):
                                                continue
                                            w = PW if (s1 in priority_subjects
                                                       or s2 in priority_subjects) else NW
                                            pv = model.NewBoolVar(
                                                f"p_{t}_{day}_{p1}_{p2}_{c1}_{c2}")
                                            model.AddBoolAnd([
                                                x[c1][day][p1][s1],
                                                x[c2][day][p2][s2]
                                            ]).OnlyEnforceIf(pv)
                                            model.AddBoolOr([
                                                x[c1][day][p1][s1].Not(),
                                                x[c2][day][p2][s2].Not()
                                            ]).OnlyEnforceIf(pv.Not())
                                            penalty_terms.append(w * pv)

            # S1b: 同学年・同教科は同じ曜日に集約するペナルティ
            # has_on_day[c][s][d] = クラスcが教科sを曜日dに持つか
            has_on_day = {}
            for c in classes:
                has_on_day[c] = {}
                for s in SUBJECTS:
                    has_on_day[c][s] = {}
                    for d in DAYS:
                        dp = periods_per_day.get(d, 6)
                        xvars = [x[c][d][p][s] for p in range(1, dp + 1)]
                        hv = model.NewBoolVar(f"hod_{c}_{s}_{d}")
                        model.AddBoolOr(xvars).OnlyEnforceIf(hv)
                        model.AddBoolAnd([xv.Not() for xv in xvars]).OnlyEnforceIf(hv.Not())
                        has_on_day[c][s][d] = hv

            # 同学年クラスのペアで、同教科が異なる曜日に配置されたらペナルティ
            grade_cls_map: dict[int, list] = {}
            for c in classes:
                grade_cls_map.setdefault(get_grade(c), []).append(c)

            for gcls in grade_cls_map.values():
                for i, c1 in enumerate(gcls):
                    for c2 in gcls[i + 1:]:
                        for s in SUBJECTS:
                            w = PW if s in priority_subjects else NW
                            for d in DAYS:
                                # c1とc2でsの曜日配置が食い違うときペナルティ
                                mismatch = model.NewBoolVar(
                                    f"daymm_{c1}_{c2}_{s}_{d}")
                                model.AddBoolXOr([
                                    has_on_day[c1][s][d],
                                    has_on_day[c2][s][d],
                                    mismatch.Not()
                                ])
                                penalty_terms.append(w * mismatch)

            # S3: 学年交互防止（1〜4限で同一教員が連続して異学年を担当しない）
            # 優先教科が絡む遷移は S3_PW、それ以外は S3_NW でペナルティを課す
            S3_PW, S3_NW = 10, 1
            S3_TARGET = [p for p in range(1, 5)
                         if p <= min(periods_per_day.get(d, 6) for d in DAYS)]
            grades_in_school = sorted({get_grade(c) for c in classes})

            if len(grades_in_school) >= 2 and len(S3_TARGET) >= 2:
                # teaches_grade_var[t][d][p][g]: 全教科版
                teaches_grade_var = {}
                # teaches_prio_grade[t][d][p][g]: 優先教科のみ版
                teaches_prio_grade = {}
                for t in st.session_state["teachers"]:
                    teaches_grade_var[t] = {}
                    teaches_prio_grade[t] = {}
                    for d in DAYS:
                        teaches_grade_var[t][d] = {}
                        teaches_prio_grade[t][d] = {}
                        for p in S3_TARGET:
                            teaches_grade_var[t][d][p] = {}
                            teaches_prio_grade[t][d][p] = {}
                            for g in grades_in_school:
                                # 全教科
                                tvars_g = [
                                    x[c][d][p][s]
                                    for c in classes if get_grade(c) == g
                                    for s in SUBJECTS
                                    if t in assignments.get(c, {}).get(s, [])
                                    and p <= periods_per_day.get(d, 6)
                                ]
                                tg = model.NewBoolVar(f"tg_{t}_{d}_{p}_{g}")
                                if tvars_g:
                                    model.AddBoolOr(tvars_g).OnlyEnforceIf(tg)
                                    model.AddBoolAnd(
                                        [v.Not() for v in tvars_g]
                                    ).OnlyEnforceIf(tg.Not())
                                else:
                                    model.Add(tg == 0)
                                teaches_grade_var[t][d][p][g] = tg

                                # 優先教科のみ
                                pvars_g = [
                                    x[c][d][p][s]
                                    for c in classes if get_grade(c) == g
                                    for s in priority_subjects
                                    if t in assignments.get(c, {}).get(s, [])
                                    and p <= periods_per_day.get(d, 6)
                                ]
                                tpg = model.NewBoolVar(f"tpg_{t}_{d}_{p}_{g}")
                                if pvars_g:
                                    model.AddBoolOr(pvars_g).OnlyEnforceIf(tpg)
                                    model.AddBoolAnd(
                                        [v.Not() for v in pvars_g]
                                    ).OnlyEnforceIf(tpg.Not())
                                else:
                                    model.Add(tpg == 0)
                                teaches_prio_grade[t][d][p][g] = tpg

                # 連続する時限 (p, p+1) で学年が切り替わるとペナルティ
                # 優先教科が絡む遷移は S3_PW、それ以外は S3_NW
                for t in st.session_state["teachers"]:
                    for d in DAYS:
                        for p in S3_TARGET[:-1]:   # (1,2),(2,3),(3,4)
                            p2 = p + 1
                            if p2 not in S3_TARGET:
                                continue
                            for g1 in grades_in_school:
                                for g2 in grades_in_school:
                                    if g1 == g2:
                                        continue
                                    tg1 = teaches_grade_var[t][d][p][g1]
                                    tg2 = teaches_grade_var[t][d][p2][g2]
                                    # 学年遷移変数
                                    pv = model.NewBoolVar(
                                        f"s3_{t}_{d}_{p}_{g1}_{g2}")
                                    model.AddBoolAnd([tg1, tg2]).OnlyEnforceIf(pv)
                                    model.AddBoolOr(
                                        [tg1.Not(), tg2.Not()]
                                    ).OnlyEnforceIf(pv.Not())
                                    # 基本ペナルティ（全遷移に S3_NW）
                                    penalty_terms.append(S3_NW * pv)
                                    # 優先教科が絡む場合は追加ペナルティ
                                    if priority_subjects:
                                        tpg1 = teaches_prio_grade[t][d][p][g1]
                                        pv_p = model.NewBoolVar(
                                            f"s3p_{t}_{d}_{p}_{g1}_{g2}")
                                        model.AddBoolAnd(
                                            [pv, tpg1]
                                        ).OnlyEnforceIf(pv_p)
                                        model.AddBoolOr(
                                            [pv.Not(), tpg1.Not()]
                                        ).OnlyEnforceIf(pv_p.Not())
                                        # 追加分: PW - NW = 9
                                        penalty_terms.append(
                                            (S3_PW - S3_NW) * pv_p)

        if penalty_terms:
            model.Minimize(sum(penalty_terms))

        # ── ソルバー実行 ─────────────────────────────────────
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds  = float(time_limit)
        solver.parameters.num_search_workers   = 4
        if random_seed != 0:
            solver.parameters.random_seed = random_seed % (2**31 - 1)
        status = solver.Solve(model)

        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            for cls in classes:
                for day in DAYS:
                    for p in range(1, periods_per_day.get(day, 6) + 1):
                        if timetable[cls].get(day, {}).get(p):
                            continue
                        for subj in _cls_subjects[cls]:
                            if solver.Value(x[cls][day][p][subj]) == 1:
                                timetable[cls].setdefault(day, {})[p] = subj
                                # weekly_remaining を減算して整合性を保つ
                                if subj in weekly_remaining.get(cls, {}):
                                    weekly_remaining[cls][subj] = max(
                                        0, weekly_remaining[cls][subj] - 1)
                                break
            return "solved"
        return "failed"

    except Exception as e:
        st.error(f"OR-Toolsエラー: {e}")
        return "failed"



# ============================================================
# Part 4: 学年別段階生成
# ============================================================

def solve_grade_by_grade(
    timetable: dict,
    timetable_sg: dict,
    weekly_remaining: dict,
    assignments: dict,
    special_rooms: list,
    unavailable: dict,
    periods_per_day: dict,
    small_group_classes: dict,
    priority_subjects: list,
    ort_time_limit: float = 60.0
) -> str:
    """
    学年ごとに順番に解く段階的生成。

    クロス学年の制約処理:
    ・R1（教員重複）: 前学年の使用済みコマを extra_unavail として次学年に引き継ぐ
    ・R3（教室容量）: 前学年の per-slot 使用数を特別教室の容量から引いて渡す
    ・少人数同期のある学年を先に解く（制約が厳しいため）
    戻り値: "solved" / "failed_grade{学年番号}"
    """
    from collections import defaultdict

    # ── 学年別クラスグループ ─────────────────────────────────
    grade_groups: dict[int, list] = defaultdict(list)
    for cls in timetable:
        grade_groups[get_grade(cls)].append(cls)

    # ── 少人数グループを学年別にフィルタリング ───────────────────
    # 新形式（students/student_timetable）では home_class で判定する。
    # 旧形式（classes/sync_groups）でも引き続き動作するよう両方を考慮。
    def sg_for_classes(cls_list):
        cls_set = set(cls_list)
        result = {}
        for sg_name, sg_data in small_group_classes.items():
            # ── 新形式: students の home_class がこの学年に含まれるか ──
            new_format_match = any(
                s.get("home_class") in cls_set
                for s in sg_data.get("students", [])
            )
            if new_format_match:
                # 新形式はデータをそのまま渡す（色制約は get_color_constraints が処理）
                result[sg_name] = sg_data
                continue
            # ── 旧形式: classes/sync_groups キーで判定 ─────────────
            in_list = [c for c in sg_data.get("classes", []) if c in cls_set]
            if not in_list:
                continue
            filtered_groups = [
                [c for c in g if c in cls_set]
                for g in (sg_data.get("sync_groups") or [sg_data.get("classes", [])])
                if sum(1 for c in g if c in cls_set) >= 2
            ]
            if filtered_groups:
                result[sg_name] = {**sg_data,
                                   "classes": in_list,
                                   "sync_groups": filtered_groups}
        return result

    # ── 処理順: 少人数同期のある学年を優先 ──────────────────
    sorted_grades = sorted(grade_groups.keys(),
                           key=lambda g: -len(sg_for_classes(grade_groups[g])))

    # ── 追加不在（前学年で使用した教員スロット）───────────────
    # teacher -> day -> [period, ...]
    extra_unavail: dict[str, dict[str, list]] = defaultdict(
        lambda: defaultdict(list))
    # 教室使用カウント: subject -> day -> period -> count
    room_slot_used: dict[str, dict[str, dict[int, int]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(int)))

    def make_merged_unavail():
        result = {}
        all_teachers = set(unavailable) | set(extra_unavail)
        for t in all_teachers:
            result[t] = {
                d: sorted(set(
                    list(unavailable.get(t, {}).get(d, [])) +
                    list(extra_unavail[t].get(d, []))
                ))
                for d in DAYS
            }
        return result

    def make_adjusted_rooms(grade):
        """
        前学年の per-slot 使用数を引いた特別教室リストを返す。
        時限ごとに残容量が変わるが、OR-Tools/MRV の仕様上
        最も制約の厳しい時限（最小残容量）を各教室の有効容量とする。
        さらに詳細な時限制約は MRV キャッシュ経由で自動適用される。
        """
        adjusted = []
        for room in special_rooms:
            cap = room["capacity"]
            gs  = room.get("grade_subjects", [])
            # この教室が担当する全教科の使用数の最大値を求める
            max_used = 0
            for _g, _s in gs:
                for day in DAYS:
                    for p in range(1, periods_per_day.get(day, 6) + 1):
                        used = room_slot_used[_s][day].get(p, 0)
                        max_used = max(max_used, used)
            new_cap = max(0, cap - max_used)
            adjusted.append({**room, "capacity": new_cap})
        return adjusted

    # ── 学年ごとに生成 ────────────────────────────────────────
    for grade in sorted_grades:
        grade_classes = grade_groups[grade]
        sg_g = sg_for_classes(grade_classes)
        mv   = make_merged_unavail()

        # この学年の timetable/wr（手動コマはすでに preprocess 済み）
        tt_g = {cls: {d: dict(timetable[cls][d]) for d in DAYS}
                for cls in grade_classes}
        wr_g = {cls: dict(weekly_remaining[cls]) for cls in grade_classes}

        result = solve_ortools(
            tt_g, {}, wr_g, assignments,
            special_rooms, mv, periods_per_day, sg_g,
            priority_subjects, time_limit=ort_time_limit
        )

        if result != "solved":
            return f"failed_grade{grade}"

        # ── 解けた学年をメインの timetable/wr に書き戻す ──────
        for cls in grade_classes:
            timetable[cls] = tt_g[cls]
            weekly_remaining[cls] = wr_g[cls]

        # ── 教員使用スロットと教室使用スロットを次学年へ引き継ぐ ──
        for cls in grade_classes:
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    subj = timetable[cls][day].get(p)
                    if not subj:
                        continue
                    # R1: 教員スロット蓄積
                    for t in assignments.get(cls, {}).get(subj, []):
                        if p not in extra_unavail[t][day]:
                            extra_unavail[t][day].append(p)
                    # R3: 教室使用カウント
                    for room in special_rooms:
                        gs_room = room.get("grade_subjects", [])
                        if any(_s == subj for _g, _s in gs_room):
                            room_slot_used[subj][day][p] += 1

    return "solved"


# ============================================================
# Part 4: 教科別段階生成
# ============================================================

def solve_subject_by_subject(
    timetable: dict,
    timetable_sg: dict,
    weekly_remaining: dict,
    assignments: dict,
    special_rooms: list,
    unavailable: dict,
    periods_per_day: dict,
    small_group_classes: dict,
    priority_subjects: list,
    first_subjects: list,
    ort_time_limit: float = 60.0
) -> str:
    """
    選択した教科を先に配置し、残りの教科を後から配置する段階的生成。

    Phase 1: first_subjects を全クラスに配置
    Phase 2: 残りの教科を配置（Phase 1 の教員スロットを引き継ぐ）

    戻り値: "solved" / "failed_phase1" / "failed_phase2"
    """
    remaining_subjects = [s for s in SUBJECTS if s not in first_subjects]

    for phase_idx, phase_subjects in enumerate([first_subjects, remaining_subjects], 1):
        if not phase_subjects:
            continue

        # この段階で配置する教科のみに絞った weekly_remaining
        wr_phase = {
            cls: {s: cnt for s, cnt in weekly_remaining[cls].items()
                  if s in phase_subjects}
            for cls in weekly_remaining
        }

        # 全クラスで残コマが 0 なら skip
        if all(v == 0 for rem in wr_phase.values() for v in rem.values()):
            continue

        result = solve_ortools(
            timetable, timetable_sg,
            wr_phase, assignments,
            special_rooms, unavailable,
            periods_per_day, small_group_classes,
            priority_subjects, time_limit=ort_time_limit
        )

        if result != "solved":
            return f"failed_phase{phase_idx}"

        # timetable は in-place 更新済みなので次フェーズでそのまま使う

    return "solved"


# ============================================================
# Part 4: エラーレポート
# ============================================================

def report_r6_violations(timetable: dict):
    """
    生成済み時間割のR6ソフト制約の達成状況を集計して表示する。
    違反 = 同期グループ内で、ある時限に共通教科と非共通教科が混在しているコマ。
    """
    small_group_classes = st.session_state.get("small_group_classes", {})
    if not small_group_classes:
        return

    total_slots   = 0  # チェック対象スロット（ペア×時限）数
    violations    = []  # (sg_name, day, period, {cls: subj}) のリスト

    for sg_name, sg_data in small_group_classes.items():
        groups = sg_data.get("sync_groups") or [sg_data.get("classes", [])]
        for group in groups:
            valid_group = [c for c in group if c in timetable]
            if len(valid_group) < 2:
                continue
            periods_per_day = st.session_state["periods_per_day"]
            for day in DAYS:
                for p in range(1, periods_per_day.get(day, 6) + 1):
                    subjects = {
                        c: timetable[c][day].get(p)
                        for c in valid_group
                    }
                    filled = {c: s for c, s in subjects.items() if s}
                    if len(filled) < 2:
                        continue
                    total_slots += 1
                    joint_subjs = sg_data.get("joint_subjects", [])
                    flags = [s in joint_subjs for s in filled.values()]
                    if len(set(flags)) > 1:  # 混在
                        violations.append((sg_name, day, p, subjects))

    if total_slots == 0:
        return

    ok_count = total_slots - len(violations)
    rate = ok_count / total_slots * 100

    if not violations:
        st.success(
            f"✅ R6（少人数同期）: 全 {total_slots} スロット一致 "
            f"（達成率 100%）"
        )
    else:
        st.warning(
            f"⚠️ R6（少人数同期）: {ok_count}/{total_slots} スロット一致 "
            f"（達成率 {rate:.0f}%）　違反 {len(violations)} 件"
        )
        with st.expander("R6違反の詳細を見る", expanded=False):
            for sg_name, day, p, subjects in violations:
                joint_subjs = small_group_classes.get(sg_name, {}).get("joint_subjects", [])
                detail = "　".join(
                    f"{c}={'★' if s in joint_subjs else ''}{s}"
                    for c, s in subjects.items() if s
                )
                st.markdown(f"- **{sg_name}** {day}{p}限：{detail}")


def report_conflict(timetable: dict, weekly_remaining: dict):
    """解が見つからなかった場合に未配置コマと改善策を表示する。"""
    st.error("❌ 時間割を完成させることができませんでした")

    unplaced = [
        f"・{cls} ／ {subj}: あと **{rem}** コマ未配置"
        for cls, subj_map in weekly_remaining.items()
        for subj, rem in subj_map.items()
        if rem > 0
    ]
    if unplaced:
        st.markdown("**未配置コマ一覧:**")
        for item in unplaced:
            st.markdown(item)

    st.info(
        "💡 **改善策の例**\n"
        "- 週コマ数が多い教科を減らす（STEP 4）\n"
        "- 教員の不在コマを見直す（STEP 6）\n"
        "- 特別教室の同時使用可能クラス数を増やす（STEP 7）\n"
        "- 少人数学級の同期グループを再設定する（STEP 9）\n"
        "- 手動入力コマを減らす（STEP 8）"
    )


# ============================================================
# Part 4: 時間割表示
# ============================================================

def display_class_timetable(
    cls: str,
    timetable: dict,
    manual_timetable: dict,
    is_small_group: bool = False
):
    """クラスの時間割をグリッド表示する。手動=水色・自動=薄緑。少人数色制約はセルに🔴🔵🟡で表示。"""
    max_p = get_max_periods()
    data  = {}

    # 少人数カラー制約を取得（通常学級のみ）
    color_constraints = {} if is_small_group else get_color_constraints(
        st.session_state.get("small_group_classes", {})
    )
    COLOR_EMOJI = {"red": "🔴", "blue": "🔵", "yellow": "🟡", "mixed": "🔵🟡"}

    for day in DAYS:
        day_periods = st.session_state["periods_per_day"].get(day, 6)
        col_data = []
        for p in range(1, max_p + 1):
            if p > day_periods:
                col_data.append("─")
                continue
            subj = timetable.get(cls, {}).get(day, {}).get(p, "")
            # 色プレフィックス（赤スロットのみ表示: 青・黄は生徒が通常在籍のため装飾不要）
            slot_info = color_constraints.get(cls, {}).get(day, {}).get(p)
            color_prefix = (COLOR_EMOJI.get(slot_info["color"], "")
                            if slot_info and slot_info["color"] == "red" else "")
            # 赤セル：生徒ごとの許容教科を括弧付きで末尾に表示
            allowed_suffix = ""
            if slot_info and slot_info["color"] == "red":
                _per_student = _format_red_allowed_per_student(slot_info)
                if _per_student:
                    allowed_suffix = "\n" + _per_student
            if not subj:
                label = (color_prefix + allowed_suffix.lstrip("\n")) if (color_prefix or allowed_suffix) else ""
                col_data.append(label)
                continue
            if is_small_group:
                teachers = get_sg_teachers(cls)
            else:
                teachers = get_teachers_for_slot(cls, subj)
            cell = f"{color_prefix}{subj}\n{'・'.join(teachers)}{allowed_suffix}" if teachers else f"{color_prefix}{subj}{allowed_suffix}"
            col_data.append(cell)
        data[day] = col_data

    df = pd.DataFrame(data, index=[f"{p}限" for p in range(1, max_p + 1)])
    st.dataframe(df, use_container_width=True)
    st.caption("🔴 = 少人数教室コマ（許容教科のみ配置可）")


def display_teacher_timetable(
    teacher: str,
    timetable: dict,
    timetable_sg: dict
):
    """教員の時間割をグリッド表示する。空き・不在も色分け。"""
    max_p = get_max_periods()
    data  = {}
    # 担任クラスを取得（担任設定がある場合）
    _hr_cls = next(
        (cls for cls, t in st.session_state.get("homeroom_teachers", {}).items()
         if t == teacher), "")

    for day in DAYS:
        day_periods = st.session_state["periods_per_day"].get(day, 6)
        col_data = []
        for p in range(1, max_p + 1):
            if p > day_periods:
                col_data.append("─")
                continue
            if p in st.session_state["unavailable"].get(teacher, {}).get(day, []):
                col_data.append("🚫不在")
                continue
            cell = None
            for cls, schedule in timetable.items():
                subj = schedule.get(day, {}).get(p)
                if subj and teacher in get_teachers_for_slot(cls, subj):
                    cell = f"{cls}\n{subj}"
                    break
            if cell is None:
                for sg_name, sg_schedule in timetable_sg.items():
                    subj = sg_schedule.get(day, {}).get(p)
                    if subj and teacher in get_sg_teachers(sg_name):
                        cell = f"{sg_name}\n{subj}"
                        break
            if cell is None:
                if _hr_cls and timetable.get(_hr_cls, {}).get(day, {}).get(p) is None:
                    # 担任クラスの空きコマ = 担任が自クラスで授業
                    cell = f"担任({_hr_cls})"
                else:
                    cell = "空き"
            col_data.append(cell)
        data[day] = col_data

    df = pd.DataFrame(data, index=[f"{p}限" for p in range(1, max_p + 1)])
    st.dataframe(df, use_container_width=True)
    st.caption("🚫 不在　担任(クラス)=担任として自クラスで授業　空き=空きコマ")


def display_all_teachers_timetable(timetable: dict, timetable_sg: dict):
    """全教員の時間割を一覧表示する。縦=教員、横=曜日×時限。"""
    teachers      = st.session_state["teachers"]
    periods_per_day = st.session_state["periods_per_day"]
    unavailable   = st.session_state["unavailable"]
    small_group_classes = st.session_state["small_group_classes"]

    # 列ラベルを生成（例: 月1, 月2, ..., 金6）
    columns = [
        f"{day}{p}"
        for day in DAYS
        for p in range(1, periods_per_day.get(day, 6) + 1)
    ]

    # 担任逆引き {教員名: 担任クラス}
    _homeroom_rev = {
        t: cls for cls, t in st.session_state.get("homeroom_teachers", {}).items() if t
    }

    rows = {}
    for teacher in teachers:
        _hr_cls = _homeroom_rev.get(teacher, "")
        row = []
        for day in DAYS:
            day_periods = periods_per_day.get(day, 6)
            for p in range(1, day_periods + 1):
                if p in unavailable.get(teacher, {}).get(day, []):
                    row.append("🚫不在")
                    continue
                cell = ""
                for cls, schedule in timetable.items():
                    subj = schedule.get(day, {}).get(p)
                    if subj and teacher in get_teachers_for_slot(cls, subj):
                        cell = f"{cls} {subj}"
                        break
                if not cell:
                    for sg_name, sg_schedule in timetable_sg.items():
                        subj = sg_schedule.get(day, {}).get(p)
                        if subj and teacher in get_sg_teachers(sg_name, small_group_classes):
                            cell = f"{sg_name} {subj}"
                            break
                if not cell and _hr_cls and timetable.get(_hr_cls, {}).get(day, {}).get(p) is None:
                    cell = f"担任({_hr_cls})"
                row.append(cell)
        rows[teacher] = row

    df = pd.DataFrame(rows, index=columns).T
    st.dataframe(df, use_container_width=True)
    st.caption("🚫 不在　空欄=空きコマ")


# ============================================================
# Part 4: Excel 出力
# ============================================================

def export_timetable_to_excel(
    timetable: dict,
    timetable_sg: dict,
    manual_timetable: dict
) -> bytes:
    """
    生成済み時間割をExcelに出力する。
    ①通常学級シート ②少人数学級シート ③教員別シート
    """
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_align  = Alignment(wrap_text=True, horizontal="center", vertical="center")
    max_p       = get_max_periods()

    def make_fill(color):
        return PatternFill("solid", fgColor=color)

    # 少人数カラー制約（Excel塗りつぶし用）
    _xl_color_constraints = get_color_constraints(
        st.session_state.get("small_group_classes", {})
    )
    # 色制約 → Excel fill color
    _SLOT_FILL = {
        "red":    "FFD0D0",   # 薄赤
        "blue":   "D0E8FF",   # 薄青
        "yellow": "FFF5CC",   # 薄黄
        "mixed":  "D0F0FF",   # 薄水色
    }

    def write_timetable_sheet(ws, cls, schedule, manual, is_sg=False):
        """共通の時間割シート書き込み処理。"""
        # ヘッダー行（曜日）
        ws.append(["時限"] + DAYS)
        _apply_header_style(ws)

        # 担任行（通常学級のみ）
        _homeroom = "" if is_sg else st.session_state.get("homeroom_teachers", {}).get(cls, "")
        if _homeroom:
            ws.append([f"担任: {_homeroom}"] + [""] * len(DAYS))
            _hr_row = 2
            ws.merge_cells(start_row=_hr_row, start_column=1,
                           end_row=_hr_row, end_column=len(DAYS) + 1)
            _hr_cell = ws.cell(row=_hr_row, column=1)
            _hr_cell.font      = Font(bold=True)
            _hr_cell.fill      = make_fill(COLOR_HEADER)
            _hr_cell.alignment = Alignment(horizontal="left", vertical="center")
            _hr_cell.border    = border
            ws.row_dimensions[_hr_row].height = 18
        _row_offset = 2 if _homeroom else 1  # 担任行がある場合はオフセット+1

        for p in range(1, max_p + 1):
            row_cells = []
            has_allowed_note = False
            for day in DAYS:
                day_periods = st.session_state["periods_per_day"].get(day, 6)
                if p > day_periods:
                    row_cells.append("─")
                    continue
                subj = schedule.get(day, {}).get(p, "")
                # 赤セルの許容教科（生徒ごとに括弧付き）
                slot_info_cell = None if is_sg else _xl_color_constraints.get(cls, {}).get(day, {}).get(p)
                allowed_note = ""
                if slot_info_cell and slot_info_cell["color"] == "red":
                    _ps = _format_red_allowed_per_student(slot_info_cell)
                    if _ps:
                        allowed_note = "\n" + _ps
                    has_allowed_note = True
                if not subj:
                    row_cells.append(allowed_note.lstrip("\n") if allowed_note else "")
                    continue
                if is_sg:
                    teachers = get_sg_teachers(cls)
                else:
                    teachers = get_teachers_for_slot(cls, subj)
                cell_val = f"{subj}\n{'・'.join(teachers)}{allowed_note}" if teachers else f"{subj}{allowed_note}"
                row_cells.append(cell_val)

            ws.append([f"{p}限"] + row_cells)
            row_idx = p + _row_offset
            ws.row_dimensions[row_idx].height = 50 if has_allowed_note else 35

            for col_idx, day in enumerate(DAYS, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border    = border
                cell.alignment = wrap_align
                day_periods = st.session_state["periods_per_day"].get(day, 6)
                if p > day_periods:
                    continue
                subj = schedule.get(day, {}).get(p, "")
                # 色制約に基づく塗りつぶし優先（赤スロットのみ色付け）
                # blue/yellow/mixed スロットは通常クラスにいる時間帯なので色不要
                slot_info = None if is_sg else _xl_color_constraints.get(
                    cls, {}).get(day, {}).get(p)
                if slot_info and slot_info["color"] == "red":
                    cell.fill = make_fill(_SLOT_FILL["red"])
                    if subj:
                        cell.font = Font(bold=True)
                elif not subj:
                    pass  # 空欄はそのまま
                else:
                    is_manual = manual.get(cls, {}).get(day, {}).get(p) is not None
                    if is_sg:
                        cell.fill = make_fill(COLOR_SG)
                    elif is_manual:
                        cell.fill = make_fill(COLOR_MANUAL)
                    else:
                        cell.fill = make_fill(COLOR_AUTO)

            # 時限列のスタイル
            lc = ws.cell(row=row_idx, column=1)
            lc.border    = border
            lc.alignment = wrap_align
            lc.fill      = make_fill(COLOR_HEADER)
            lc.font      = Font(bold=True)

        # 列幅設定
        ws.column_dimensions["A"].width = 8
        for col_idx in range(2, len(DAYS) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── ①全学年一覧シート（縦軸:学年クラス、横軸:曜日×時限） ──
    ws = wb.create_sheet("全学年一覧")
    ppd_grade = st.session_state.get("periods_per_day_grade", {})
    all_cls_list = get_all_classes()

    # 横軸: 曜日×時限の列を構築
    col_headers = []  # (day, period) のリスト
    for day in DAYS:
        day_max = max(
            (ppd_grade.get(g, {}).get(day, 6) for g in GRADES),
            default=max_p
        )
        for p in range(1, day_max + 1):
            col_headers.append((day, p))

    # ヘッダー行1: 学年クラス・担任（列1-2）＋曜日（列3+）
    for _hcol, _hval in [(1, "学年クラス"), (2, "担任")]:
        _hc = ws.cell(row=1, column=_hcol, value=_hval)
        _hc.fill      = make_fill(COLOR_HEADER)
        _hc.font      = Font(bold=True)
        _hc.alignment = Alignment(horizontal="center", vertical="center")
        _hc.border    = border
        ws.merge_cells(start_row=1, start_column=_hcol, end_row=2, end_column=_hcol)

    # 担任ヘッダーの行2セルにも罫線を設定（merge後の下セル）
    ws.cell(row=2, column=2).border = border

    _homeroom_teachers = st.session_state.get("homeroom_teachers", {})

    day_col_ranges = {}  # day -> (start_col, end_col)
    for col_i, (day, p) in enumerate(col_headers):
        col = col_i + 3  # 列1=学年クラス、列2=担任、列3+が時間割
        if day not in day_col_ranges:
            day_col_ranges[day] = [col, col]
        else:
            day_col_ranges[day][1] = col

    for day, (sc, ec) in day_col_ranges.items():
        if sc == ec:
            cell = ws.cell(row=1, column=sc, value=day)
        else:
            ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
            cell = ws.cell(row=1, column=sc, value=day)
        cell.fill      = make_fill(COLOR_HEADER)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    # ヘッダー行2: 時限（列3+）
    for col_i, (day, p) in enumerate(col_headers):
        col = col_i + 3
        cell = ws.cell(row=2, column=col, value=f"{p}限")
        cell.fill      = make_fill(COLOR_HEADER)
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    # データ行: 各クラス
    for row_i, cls in enumerate(all_cls_list):
        row = row_i + 3
        g   = get_grade(cls)

        # 列1: クラス名ラベル
        lc = ws.cell(row=row, column=1, value=cls)
        lc.fill      = make_fill(COLOR_HEADER)
        lc.font      = Font(bold=True)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border    = border

        # 列2: 担任名
        _hr = _homeroom_teachers.get(cls, "")
        hrc = ws.cell(row=row, column=2, value=_hr)
        hrc.border    = border
        hrc.alignment = Alignment(horizontal="center", vertical="center")
        if _hr:
            hrc.font = Font(bold=True)

        # 赤セル（許容教科付き）があれば行を高くする
        _cls_has_red = any(
            _xl_color_constraints.get(cls, {}).get(day, {}).get(p, {}).get("color") == "red"
            for day in DAYS
            for p in range(1, max_p + 1)
        )
        ws.row_dimensions[row].height = 32 if _cls_has_red else 20

        schedule = timetable.get(cls, {})
        for col_i, (day, p) in enumerate(col_headers):
            col  = col_i + 3
            cell = ws.cell(row=row, column=col)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # この学年のこの曜日のコマ数上限を超える列はグレー
            day_max = ppd_grade.get(g, {}).get(day, 6)
            if p > day_max:
                cell.value = ""
                cell.fill  = make_fill(COLOR_ABSENT)
                continue

            subj = schedule.get(day, {}).get(p, "")
            # 少人数色制約チェック
            _overview_slot = _xl_color_constraints.get(cls, {}).get(day, {}).get(p)
            if _overview_slot and _overview_slot["color"] == "red":
                # 赤スロットのみ色付け（blue/yellow/mixed は通常クラスなので色不要）
                cell.fill = make_fill(_SLOT_FILL["red"])
                allowed_str = _format_red_allowed_per_student(_overview_slot)
                if allowed_str:
                    cell.value = (subj + "\n" + allowed_str) if subj else allowed_str
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                if subj:
                    cell.font = Font(bold=True)
            else:
                cell.value = subj if subj else ""
                if subj:
                    is_manual = manual_timetable.get(cls, {}).get(day, {}).get(p) is not None
                    cell.fill = make_fill(COLOR_MANUAL if is_manual else COLOR_AUTO)

    # 列幅・行高
    ws.column_dimensions["A"].width = 10  # 学年クラス
    ws.column_dimensions["B"].width = 10  # 担任
    for col_i in range(len(col_headers)):
        ws.column_dimensions[get_column_letter(col_i + 3)].width = 6
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16

    # ── ②通常学級シート（個別） ──────────────────────────────
    for cls in get_all_classes():
        ws = wb.create_sheet(cls)
        write_timetable_sheet(
            ws, cls,
            timetable.get(cls, {}),
            manual_timetable, is_sg=False
        )

    # ── ②少人数学級シート ────────────────────────────────────
    for sg_name in st.session_state["small_group_classes"]:
        ws = wb.create_sheet(sg_name)
        write_timetable_sheet(
            ws, sg_name,
            timetable_sg.get(sg_name, {}),
            {}, is_sg=True
        )

    # ── ③教員別シート ────────────────────────────────────────
    # 担任逆引き {教員名: 担任クラス}
    _excel_homeroom_rev = {
        t: cls for cls, t in st.session_state.get("homeroom_teachers", {}).items() if t
    }
    for teacher in st.session_state["teachers"]:
        ws = wb.create_sheet(f"教員_{teacher}")
        ws.append(["時限"] + DAYS)
        _apply_header_style(ws)
        _excel_hr_cls = _excel_homeroom_rev.get(teacher, "")

        for p in range(1, max_p + 1):
            row_cells = []
            for day in DAYS:
                day_periods = st.session_state["periods_per_day"].get(day, 6)
                if p > day_periods:
                    row_cells.append("─")
                    continue
                if p in st.session_state["unavailable"].get(teacher, {}).get(day, []):
                    row_cells.append("不在")
                    continue
                cell_val = None
                for cls, schedule in timetable.items():
                    subj = schedule.get(day, {}).get(p)
                    if subj and teacher in get_teachers_for_slot(cls, subj):
                        cell_val = f"{cls}\n{subj}"
                        break
                if cell_val is None:
                    for sg_name, sg_schedule in timetable_sg.items():
                        subj = sg_schedule.get(day, {}).get(p)
                        if subj and teacher in get_sg_teachers(sg_name):
                            cell_val = f"{sg_name}\n{subj}"
                            break
                if cell_val is None:
                    if _excel_hr_cls and timetable.get(_excel_hr_cls, {}).get(day, {}).get(p) is None:
                        cell_val = f"担任({_excel_hr_cls})"
                    else:
                        cell_val = "空き"
                row_cells.append(cell_val)

            ws.append([f"{p}限"] + row_cells)
            row_idx = p + 1
            ws.row_dimensions[row_idx].height = 35

            for col_idx, day in enumerate(DAYS, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border    = border
                cell.alignment = wrap_align
                day_periods = st.session_state["periods_per_day"].get(day, 6)
                if p > day_periods:
                    continue
                val = row_cells[col_idx - 2]
                if val == "不在":
                    cell.fill = make_fill(COLOR_ABSENT)
                elif val == "空き":
                    pass
                elif val and str(val).startswith("担任("):
                    cell.fill = make_fill("FFF0E0")  # 薄橙 = 担任授業（暗黙）
                else:
                    cell.fill = make_fill(COLOR_AUTO)

            lc = ws.cell(row=row_idx, column=1)
            lc.border = border; lc.alignment = wrap_align
            lc.fill   = make_fill(COLOR_HEADER)
            lc.font   = Font(bold=True)

        ws.column_dimensions["A"].width = 8
        for col_idx in range(2, len(DAYS) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # ── ④全教員一覧シート ─────────────────────────────────────
    ws = wb.create_sheet("全教員一覧")
    periods_per_day = st.session_state["periods_per_day"]
    unavailable     = st.session_state["unavailable"]
    sg_classes      = st.session_state["small_group_classes"]
    _all_homeroom_rev = {
        t: cls for cls, t in st.session_state.get("homeroom_teachers", {}).items() if t
    }

    # ヘッダー行（教員名 | 月1 月2 ... 金6）
    col_labels = [
        f"{day}{p}"
        for day in DAYS
        for p in range(1, periods_per_day.get(day, 6) + 1)
    ]
    ws.append(["教員名"] + col_labels)
    _apply_header_style(ws)

    for teacher in st.session_state["teachers"]:
        _all_hr_cls = _all_homeroom_rev.get(teacher, "")
        row_vals = [teacher]
        for day in DAYS:
            day_periods = periods_per_day.get(day, 6)
            for p in range(1, day_periods + 1):
                if p in unavailable.get(teacher, {}).get(day, []):
                    row_vals.append("不在")
                    continue
                cell_val = ""
                for cls, schedule in timetable.items():
                    subj = schedule.get(day, {}).get(p)
                    if subj and teacher in get_teachers_for_slot(cls, subj):
                        cell_val = f"{cls} {subj}"
                        break
                if not cell_val:
                    for sg_name, sg_schedule in timetable_sg.items():
                        subj = sg_schedule.get(day, {}).get(p)
                        if subj and teacher in get_sg_teachers(sg_name, sg_classes):
                            cell_val = f"{sg_name} {subj}"
                            break
                if not cell_val and _all_hr_cls and timetable.get(_all_hr_cls, {}).get(day, {}).get(p) is None:
                    cell_val = f"担任({_all_hr_cls})"
                row_vals.append(cell_val)
        ws.append(row_vals)

    # スタイル適用
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border    = border
            cell.alignment = wrap_align
            if cell.column == 1:
                cell.fill = make_fill(COLOR_HEADER)
                cell.font = Font(bold=True)
            elif cell.value == "不在":
                cell.fill = make_fill(COLOR_ABSENT)
            elif cell.value and str(cell.value).startswith("担任("):
                cell.fill = make_fill("FFF0E0")  # 薄橙 = 担任授業（暗黙）
            elif cell.value:
                cell.fill = make_fill(COLOR_AUTO)
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(col_labels) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# Part 5: 手動編集 ── コマ交換機能
# ============================================================

def _get_room_for(cls: str, subj: str):
    """クラスと教科から対応する特別教室オブジェクトを返す（なければNone）"""
    grade = get_grade(cls)
    for room in st.session_state["special_rooms"]:
        for g, s in room.get("grade_subjects", []):
            if s == subj and g == grade:
                return room
    return None


def _count_room_usage(timetable: dict, room: dict, day: str, p: int, exclude_cls: str) -> int:
    """指定時限にその特別教室を使っているクラス数を数える（exclude_clsを除く）"""
    count = 0
    for cls, sched in timetable.items():
        if cls == exclude_cls:
            continue
        curr_s = sched.get(day, {}).get(p)
        if curr_s:
            for g, s in room.get("grade_subjects", []):
                if curr_s == s and get_grade(cls) == g:
                    count += 1
                    break
    return count


def is_valid_swap(
    timetable: dict,
    timetable_sg: dict,
    cls1: str, day1: str, p1: int,
    cls2: str, day2: str, p2: int
) -> bool:
    """
    (cls1, day1, p1) の教科と (cls2, day2, p2) の教科を交換したとき、
    全制約（R1/R2/R3/R4/R5/R7）を満たすかチェックする。
    """
    if cls1 == cls2 and day1 == day2 and p1 == p2:
        return False

    unavailable     = st.session_state["unavailable"]
    small_group_cls = st.session_state["small_group_classes"]
    r7_exempt       = st.session_state.get("r7_exempt_teachers", [])

    subj1 = timetable.get(cls1, {}).get(day1, {}).get(p1)
    subj2 = timetable.get(cls2, {}).get(day2, {}).get(p2)

    if not subj1:
        return False
    if cls1 == cls2 and subj1 == subj2:
        return False
    if p1 > get_periods_for_class(cls1, day1) or p2 > get_periods_for_class(cls2, day2):
        return False

    same_cls_same_day = (cls1 == cls2 and day1 == day2)

    # ── R4: 教員不在 ──────────────────────────────────────────
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            if p1 in unavailable.get(t, {}).get(day1, []):
                return False
    for t in get_teachers_for_slot(cls2, subj1):
        if p2 in unavailable.get(t, {}).get(day2, []):
            return False

    # ── R5: 同日同教科（クラス内）────────────────────────────
    if subj2:
        for p_x, s_x in timetable.get(cls1, {}).get(day1, {}).items():
            if p_x == p1:
                continue
            if same_cls_same_day and p_x == p2:
                continue
            if s_x == subj2:
                return False
    for p_x, s_x in timetable.get(cls2, {}).get(day2, {}).items():
        if p_x == p2:
            continue
        if same_cls_same_day and p_x == p1:
            continue
        if s_x == subj1:
            return False

    # ── R1: 教員重複 ─────────────────────────────────────────
    def busy_at(day, p, exclude_cls):
        busy = set()
        for cls, schedule in timetable.items():
            if cls == exclude_cls:
                continue
            s = schedule.get(day, {}).get(p)
            if s:
                for t in get_teachers_for_slot(cls, s):
                    busy.add(t)
        for sg_name, schedule in timetable_sg.items():
            s = schedule.get(day, {}).get(p)
            if s:
                for t in small_group_cls.get(sg_name, {}).get("teachers", []):
                    busy.add(t)
        return busy

    busy_src = busy_at(day1, p1, cls1)
    busy_dst = busy_at(day2, p2, cls2)

    t1_set = set(get_teachers_for_slot(cls2, subj1))
    t2_set = set(get_teachers_for_slot(cls1, subj2)) if subj2 else set()

    if t2_set & busy_src:
        return False
    if t1_set & busy_dst:
        return False
    if day1 == day2 and p1 == p2 and t2_set & t1_set:
        return False

    # ── R3: 特別教室容量 ──────────────────────────────────────
    if subj2:
        room2 = _get_room_for(cls1, subj2)
        if room2:
            cnt = _count_room_usage(timetable, room2, day1, p1, cls1)
            if cnt + 1 > room2["capacity"]:
                return False
    room1 = _get_room_for(cls2, subj1)
    if room1:
        cnt = _count_room_usage(timetable, room1, day2, p2, cls2)
        if cnt + 1 > room1["capacity"]:
            return False

    # ── R7: 教員日内コマ数上限 ───────────────────────────────
    teacher_day_cnt: dict = {}
    for cls, schedule in timetable.items():
        for day, periods in schedule.items():
            for p, subj in periods.items():
                if not subj:
                    continue
                if (cls == cls1 and day == day1 and p == p1) or \
                   (cls == cls2 and day == day2 and p == p2):
                    continue
                for t in get_teachers_for_slot(cls, subj):
                    teacher_day_cnt.setdefault(t, {})
                    teacher_day_cnt[t][day] = teacher_day_cnt[t].get(day, 0) + 1
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            teacher_day_cnt.setdefault(t, {})
            teacher_day_cnt[t][day1] = teacher_day_cnt[t].get(day1, 0) + 1
    for t in get_teachers_for_slot(cls2, subj1):
        teacher_day_cnt.setdefault(t, {})
        teacher_day_cnt[t][day2] = teacher_day_cnt[t].get(day2, 0) + 1

    affected = set()
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            affected.add((t, day1))
    for t in get_teachers_for_slot(cls2, subj1):
        affected.add((t, day2))
    for t, day in affected:
        if t in r7_exempt:
            continue
        total_today = get_periods_for_class(cls1 if day == day1 else cls2, day)
        unavail_cnt = len(unavailable.get(t, {}).get(day, []))
        available   = total_today - unavail_cnt
        max_ok      = available - 1
        if max_ok <= 0:
            continue
        if teacher_day_cnt.get(t, {}).get(day, 0) > max_ok:
            return False

    # ── R2: 週コマ数保全（異クラス交換時のみ） ────────────────
    if cls1 != cls2:
        wp = st.session_state["weekly_periods"]

        def _count_subj(cls, subj):
            return sum(
                1 for prs in timetable.get(cls, {}).values()
                for s in prs.values() if s == subj
            )

        if _count_subj(cls1, subj1) - 1 < wp.get(cls1, {}).get(subj1, 0):
            return False
        if _count_subj(cls2, subj1) + 1 > wp.get(cls2, {}).get(subj1, 0):
            return False

        if subj2:
            if _count_subj(cls2, subj2) - 1 < wp.get(cls2, {}).get(subj2, 0):
                return False
            if _count_subj(cls1, subj2) + 1 > wp.get(cls1, {}).get(subj2, 0):
                return False

    # ── R6: 少人数カラー制約（赤コマ） ───────────────────────
    color_constraints = get_color_constraints(st.session_state["small_group_classes"])
    # swap後: (cls1, day1, p1) に subj2 が入る
    slot_info_1 = color_constraints.get(cls1, {}).get(day1, {}).get(p1)
    if slot_info_1 and slot_info_1.get("color") == "red":
        if subj2 and subj2 not in slot_info_1.get("allowed", set()):
            return False
    # swap後: (cls2, day2, p2) に subj1 が入る
    slot_info_2 = color_constraints.get(cls2, {}).get(day2, {}).get(p2)
    if slot_info_2 and slot_info_2.get("color") == "red":
        if subj1 not in slot_info_2.get("allowed", set()):
            return False

    return True


def get_swap_violations(
    timetable: dict,
    timetable_sg: dict,
    cls1: str, day1: str, p1: int,
    cls2: str, day2: str, p2: int,
) -> list[str]:
    """
    is_valid_swap と同じ順序で全制約を検査し、
    違反しているものを日本語メッセージのリストで返す。
    空リストなら交換可能。
    """
    violations = []

    if cls1 == cls2 and day1 == day2 and p1 == p2:
        return ["同一スロットが選択されています。"]

    unavailable     = st.session_state["unavailable"]
    small_group_cls = st.session_state["small_group_classes"]
    r7_exempt       = st.session_state.get("r7_exempt_teachers", [])
    wp              = st.session_state["weekly_periods"]

    subj1 = timetable.get(cls1, {}).get(day1, {}).get(p1)
    subj2 = timetable.get(cls2, {}).get(day2, {}).get(p2)

    if not subj1:
        return ["コマAに授業が入っていません。"]
    if cls1 == cls2 and subj1 == subj2:
        return ["同じクラスの同じ教科同士のため入れ替えの意味がありません。"]
    if p1 > get_periods_for_class(cls1, day1) or p2 > get_periods_for_class(cls2, day2):
        return ["選択した時限がその曜日の授業コマ数を超えています。"]

    same_cls_same_day = (cls1 == cls2 and day1 == day2)

    # ── R4: 教員不在 ──────────────────────────────────────────
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            if p1 in unavailable.get(t, {}).get(day1, []):
                violations.append(
                    f"R4（教員不在）: {t} 先生は {day1}曜{p1}限が不在コマです"
                    f"（{cls1}「{subj2}」を移動できません）"
                )
    for t in get_teachers_for_slot(cls2, subj1):
        if p2 in unavailable.get(t, {}).get(day2, []):
            violations.append(
                f"R4（教員不在）: {t} 先生は {day2}曜{p2}限が不在コマです"
                f"（{cls2}「{subj1}」を移動できません）"
            )

    # ── R5: 同日同教科 ────────────────────────────────────────
    if subj2:
        for p_x, s_x in timetable.get(cls1, {}).get(day1, {}).items():
            if p_x == p1:
                continue
            if same_cls_same_day and p_x == p2:
                continue
            if s_x == subj2:
                violations.append(
                    f"R5（同日同教科）: {cls1} の {day1}曜日に「{subj2}」が既に {p_x}限に入っています"
                )
                break
    for p_x, s_x in timetable.get(cls2, {}).get(day2, {}).items():
        if p_x == p2:
            continue
        if same_cls_same_day and p_x == p1:
            continue
        if s_x == subj1:
            violations.append(
                f"R5（同日同教科）: {cls2} の {day2}曜日に「{subj1}」が既に {p_x}限に入っています"
            )
            break

    # ── R1: 教員重複 ─────────────────────────────────────────
    def busy_at_v(day, p, exclude_cls):
        busy = {}
        for cls, schedule in timetable.items():
            if cls == exclude_cls:
                continue
            s = schedule.get(day, {}).get(p)
            if s:
                for t in get_teachers_for_slot(cls, s):
                    busy[t] = cls
        for sg_name, schedule in timetable_sg.items():
            s = schedule.get(day, {}).get(p)
            if s:
                for t in small_group_cls.get(sg_name, {}).get("teachers", []):
                    busy[t] = sg_name
        return busy

    busy_src = busy_at_v(day1, p1, cls1)
    busy_dst = busy_at_v(day2, p2, cls2)

    t1_set = set(get_teachers_for_slot(cls2, subj1))
    t2_set = set(get_teachers_for_slot(cls1, subj2)) if subj2 else set()

    for t in t2_set:
        if t in busy_src:
            violations.append(
                f"R1（教員重複）: {t} 先生は {day1}曜{p1}限に "
                f"【{busy_src[t]}】で既に授業があります"
                f"（「{subj2}」をここに移動できません）"
            )
    for t in t1_set:
        if t in busy_dst:
            violations.append(
                f"R1（教員重複）: {t} 先生は {day2}曜{p2}限に "
                f"【{busy_dst[t]}】で既に授業があります"
                f"（「{subj1}」をここに移動できません）"
            )

    # ── R3: 特別教室容量 ──────────────────────────────────────
    if subj2:
        room2 = _get_room_for(cls1, subj2)
        if room2:
            cnt = _count_room_usage(timetable, room2, day1, p1, cls1)
            if cnt + 1 > room2["capacity"]:
                violations.append(
                    f"R3（特別教室）: {day1}曜{p1}限に「{subj2}」の教室を使う授業が"
                    f"既に {cnt} クラスあり、上限（{room2['capacity']} クラス）を超えます"
                )
    room1 = _get_room_for(cls2, subj1)
    if room1:
        cnt = _count_room_usage(timetable, room1, day2, p2, cls2)
        if cnt + 1 > room1["capacity"]:
            violations.append(
                f"R3（特別教室）: {day2}曜{p2}限に「{subj1}」の教室を使う授業が"
                f"既に {cnt} クラスあり、上限（{room1['capacity']} クラス）を超えます"
            )

    # ── R2: 週コマ数 ─────────────────────────────────────────
    if cls1 != cls2:
        def _cnt(cls, subj):
            return sum(1 for prs in timetable.get(cls, {}).values()
                       for s in prs.values() if s == subj)

        if _cnt(cls1, subj1) - 1 < wp.get(cls1, {}).get(subj1, 0):
            violations.append(
                f"R2（週コマ数）: {cls1} の「{subj1}」が"
                f" {_cnt(cls1, subj1) - 1} コマになり"
                f"必要数（{wp.get(cls1, {}).get(subj1, 0)} コマ）を下回ります"
            )
        if _cnt(cls2, subj1) + 1 > wp.get(cls2, {}).get(subj1, 0):
            violations.append(
                f"R2（週コマ数）: {cls2} の「{subj1}」が"
                f" {_cnt(cls2, subj1) + 1} コマになり"
                f"必要数（{wp.get(cls2, {}).get(subj1, 0)} コマ）を超えます"
            )
        if subj2:
            if _cnt(cls2, subj2) - 1 < wp.get(cls2, {}).get(subj2, 0):
                violations.append(
                    f"R2（週コマ数）: {cls2} の「{subj2}」が"
                    f" {_cnt(cls2, subj2) - 1} コマになり"
                    f"必要数（{wp.get(cls2, {}).get(subj2, 0)} コマ）を下回ります"
                )
            if _cnt(cls1, subj2) + 1 > wp.get(cls1, {}).get(subj2, 0):
                violations.append(
                    f"R2（週コマ数）: {cls1} の「{subj2}」が"
                    f" {_cnt(cls1, subj2) + 1} コマになり"
                    f"必要数（{wp.get(cls1, {}).get(subj2, 0)} コマ）を超えます"
                )

    # ── R7: 教員日内コマ数上限 ───────────────────────────────
    teacher_day_cnt: dict = {}
    for cls, schedule in timetable.items():
        for day, periods in schedule.items():
            for p, subj in periods.items():
                if not subj:
                    continue
                if (cls == cls1 and day == day1 and p == p1) or \
                   (cls == cls2 and day == day2 and p == p2):
                    continue
                for t in get_teachers_for_slot(cls, subj):
                    teacher_day_cnt.setdefault(t, {})
                    teacher_day_cnt[t][day] = teacher_day_cnt[t].get(day, 0) + 1
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            teacher_day_cnt.setdefault(t, {})
            teacher_day_cnt[t][day1] = teacher_day_cnt[t].get(day1, 0) + 1
    for t in get_teachers_for_slot(cls2, subj1):
        teacher_day_cnt.setdefault(t, {})
        teacher_day_cnt[t][day2] = teacher_day_cnt[t].get(day2, 0) + 1

    affected = set()
    if subj2:
        for t in get_teachers_for_slot(cls1, subj2):
            affected.add((t, day1))
    for t in get_teachers_for_slot(cls2, subj1):
        affected.add((t, day2))
    for t, day in affected:
        if t in r7_exempt:
            continue
        total_today  = get_periods_for_class(cls1 if day == day1 else cls2, day)
        unavail_cnt  = len(unavailable.get(t, {}).get(day, []))
        max_ok       = total_today - unavail_cnt - 1
        if max_ok <= 0:
            continue
        cnt = teacher_day_cnt.get(t, {}).get(day, 0)
        if cnt > max_ok:
            violations.append(
                f"R7（教員日内コマ数）: {t} 先生の {day}曜の授業が"
                f" {cnt} コマになり上限（{max_ok} コマ）を超えます"
            )

    # ── R6: 少人数カラー制約（赤コマ） ───────────────────────
    color_constraints = get_color_constraints(st.session_state["small_group_classes"])
    # swap後: (cls1, day1, p1) に subj2 が入る
    slot_info_1 = color_constraints.get(cls1, {}).get(day1, {}).get(p1)
    if slot_info_1 and slot_info_1.get("color") == "red":
        if subj2 and subj2 not in slot_info_1.get("allowed", set()):
            allowed_str = "、".join(sorted(slot_info_1.get("allowed", set()))) or "なし"
            violations.append(
                f"R6（🔴少人数コマ）: {cls1} の {day1}曜{p1}限は赤コマです。"
                f"「{subj2}」はこのコマに配置できません"
                f"（許容教科: {allowed_str}）"
            )
    # swap後: (cls2, day2, p2) に subj1 が入る
    slot_info_2 = color_constraints.get(cls2, {}).get(day2, {}).get(p2)
    if slot_info_2 and slot_info_2.get("color") == "red":
        if subj1 not in slot_info_2.get("allowed", set()):
            allowed_str = "、".join(sorted(slot_info_2.get("allowed", set()))) or "なし"
            violations.append(
                f"R6（🔴少人数コマ）: {cls2} の {day2}曜{p2}限は赤コマです。"
                f"「{subj1}」はこのコマに配置できません"
                f"（許容教科: {allowed_str}）"
            )

    return violations


def find_valid_swaps(
    timetable: dict,
    timetable_sg: dict,
    src_cls: str, src_day: str, src_period: int
) -> list:
    """src スロットと交換可能な全スロットを返す。"""
    src_subj = timetable.get(src_cls, {}).get(src_day, {}).get(src_period)
    if not src_subj:
        return []
    results = []
    for cls in timetable:
        for day in DAYS:
            day_periods = get_periods_for_class(cls, day)
            for p in range(1, day_periods + 1):
                if cls == src_cls and day == src_day and p == src_period:
                    continue
                if is_valid_swap(timetable, timetable_sg,
                                 src_cls, src_day, src_period,
                                 cls, day, p):
                    tgt_subj = timetable.get(cls, {}).get(day, {}).get(p) or ""
                    results.append((cls, day, p, tgt_subj))
    return results


def get_swap_impact(
    timetable: dict,
    teacher: str,
    cls_a: str, day_a: str, period_a: int,
    cls_b: str, day_b: str, period_b: int,
) -> list:
    """
    (cls_a,day_a,period_a) ↔ (cls_b,day_b,period_b) を入れ替えたとき、
    teacher 以外のどの教員のスケジュールが変化するかを返す。
    """
    subj_a = timetable.get(cls_a, {}).get(day_a, {}).get(period_a)
    subj_b = timetable.get(cls_b, {}).get(day_b, {}).get(period_b)

    impacts = []
    seen = set()

    if subj_a:
        for t in get_teachers_for_slot(cls_a, subj_a):
            if t != teacher and t not in seen:
                seen.add(t)
                impacts.append({
                    "teacher": t,
                    "from_cls": cls_a, "from_day": day_a,
                    "from_period": period_a, "from_subj": subj_a,
                    "to_cls": cls_b, "to_day": day_b, "to_period": period_b,
                })

    if subj_b:
        for t in get_teachers_for_slot(cls_b, subj_b):
            if t != teacher and t not in seen:
                seen.add(t)
                impacts.append({
                    "teacher": t,
                    "from_cls": cls_b, "from_day": day_b,
                    "from_period": period_b, "from_subj": subj_b,
                    "to_cls": cls_a, "to_day": day_a, "to_period": period_a,
                })

    return impacts


# ============================================================
# Part 4: 生成画面 render_generate()
# ============================================================

def render_generate():
    st.header("🎲 時間割生成・確認・出力")

    classes = get_all_classes()
    if not classes:
        st.warning("先にSTEP 1〜9を設定してください")
        return

    # ── バリデーション ────────────────────────────────────────
    warnings_list = validate_all_settings()
    if warnings_list:
        with st.expander(
            f"⚠️ 設定に {len(warnings_list)} 件の警告があります（クリックで展開）",
            expanded=False
        ):
            for w in warnings_list:
                st.warning(w)

    # ── ソフト制約設定 ────────────────────────────────────────
    render_soft_constraints()
    st.markdown("---")

    # ── 生成ボタン ────────────────────────────────────────────
    col1, col2, col3 = st.columns([3, 2, 1])
    with col1:
        generate_btn = st.button(
            "🚀 時間割を自動生成する",
            type="primary",
            use_container_width=True
        )
    with col2:
        has_prev = bool(st.session_state.get("generated_timetable"))
        alt_btn = st.button(
            "🔀 別パターンを生成する",
            use_container_width=True,
            disabled=not has_prev,
            help="現在と異なる配置パターンを探します（生成後に有効になります）"
        )
    with col3:
        time_limit = st.number_input(
            "制限（秒）",
            min_value=10, max_value=120, value=30, step=5,
            help="この秒数を超えるとOR-Toolsに切り替えます"
        )

    first_subjects = st.session_state.get("soft_first_subjects", [])
    subject_btn = st.button(
        "📚 教科別段階生成",
        use_container_width=True,
        disabled=not first_subjects,
        help="先に配置する教科を選択してから実行してください（ソフト制約設定で選択）"
    )

    # パターン番号表示
    pattern_no = st.session_state.get("timetable_pattern_no", 0)
    if has_prev:
        label = "標準パターン" if pattern_no == 1 else f"パターン #{pattern_no}"
        st.caption(f"🎲 現在表示中: {label}　　別パターンを生成するたびに番号が増えます")

    # ── 生成処理の共通ロジック ───────────────────────────────
    def _run_generation(forbidden_slots: frozenset = frozenset(), ort_seed: int = 0):
        """
        forbidden_slots: 別パターン生成時に禁止する (cls,day,period,subject) の集合。
                         通常生成時は空集合。
        ort_seed:        OR-Toolsフォールバック時のランダムシード。
        """
        with st.spinner("時間割を生成中..."):
            timetable = {
                cls: {day: {} for day in DAYS}
                for cls in classes
            }
            timetable_sg = {
                sg: {day: {} for day in DAYS}
                for sg in st.session_state["small_group_classes"]
            }
            weekly_periods_sg = {
                sg: {
                    subj: settings.get("weekly", 0)
                    for subj, settings in sg_data.get("subject_settings", {}).items()
                    if settings.get("weekly", 0) > 0
                }
                for sg, sg_data in st.session_state["small_group_classes"].items()
            }
            weekly_remaining, weekly_remaining_sg = preprocess(
                timetable, timetable_sg,
                st.session_state["manual_timetable"],
                st.session_state["manual_timetable_sg"],
                st.session_state["weekly_periods"],
                weekly_periods_sg
            )
            priority_subjects = st.session_state.get("soft_priority_subjects", [])
            # assignments のコピーを作り合成教科の担当教員を追加
            # （session_state 自体は変更しない）
            assignments = {
                cls: dict(v)
                for cls, v in st.session_state["assignments"].items()
            }
            for _cls in get_all_classes():
                for _ci in compute_combined_subjects(
                        st.session_state["weekly_periods"].get(_cls, {})):
                    _cn = _ci["name"]
                    _teachers_combined: list = []
                    for _s in _ci["subjects"]:
                        for _t in assignments.get(_cls, {}).get(_s, []):
                            if _t not in _teachers_combined:
                                _teachers_combined.append(_t)
                    assignments.setdefault(_cls, {})[_cn] = _teachers_combined
            special_rooms      = st.session_state["special_rooms"]
            unavailable        = st.session_state["unavailable"]
            periods_per_day    = st.session_state["periods_per_day"]
            sg_classes         = st.session_state["small_group_classes"]

            result = solve_ortools(
                timetable, timetable_sg,
                weekly_remaining, assignments,
                special_rooms, unavailable,
                periods_per_day, sg_classes,
                priority_subjects,
                random_seed=ort_seed,
                time_limit=float(time_limit)
            )

            used_grade_split = False

            # ── 学年別段階生成フォールバック ──────────────────
            # OR-Tools が失敗し、複数学年がある場合のみ試みる
            n_grades = len(set(get_grade(c) for c in classes))
            if result != "solved" and n_grades >= 2:
                st.info(
                    "🔀 学年別段階生成を試みます…"
                    "（各学年を順番に生成し、教員の衝突を調整します）"
                )
                # 新しいtimetableを用意してpreprocess
                timetable_gs = {cls: {day: {} for day in DAYS} for cls in classes}
                timetable_sg_gs = {
                    sg: {day: {} for day in DAYS} for sg in timetable_sg
                }
                wr_gs, _ = preprocess(
                    timetable_gs, timetable_sg_gs,
                    st.session_state["manual_timetable"],
                    st.session_state["manual_timetable_sg"],
                    st.session_state["weekly_periods"],
                    weekly_periods_sg
                )
                result = solve_grade_by_grade(
                    timetable_gs, timetable_sg_gs,
                    wr_gs, assignments,
                    special_rooms, unavailable,
                    periods_per_day, sg_classes,
                    priority_subjects,
                    ort_time_limit=float(time_limit) * 2
                )
                if result == "solved":
                    timetable    = timetable_gs
                    timetable_sg = timetable_sg_gs
                    weekly_remaining = wr_gs
                used_grade_split = True

            if result == "solved":
                st.session_state["generated_timetable"]    = timetable
                st.session_state["generated_timetable_sg"] = timetable_sg
                if used_grade_split:
                    method = "学年別段階生成"
                else:
                    method = "OR-Tools CP-SAT"
                pno = st.session_state.get("timetable_pattern_no", 0)
                lbl = "標準パターン" if pno == 1 else f"パターン #{pno}"
                from datetime import datetime as _dt
                history_entry = {
                    "timetable":    timetable,
                    "timetable_sg": timetable_sg,
                    "label":        f"{lbl}（{method}）",
                    "timestamp":    _dt.now().strftime("%H:%M:%S"),
                }
                st.session_state.setdefault("timetable_history", []).append(history_entry)
                st.success(f"✅ {lbl} の生成が完了しました！（{method}）")
                report_r6_violations(timetable)
            else:
                report_conflict(timetable, weekly_remaining)
        return result

    # ── 教科別段階生成 ──────────────────────────────────────────
    def _run_subject_generation():
        """選択した教科を先に配置し、残りを後から配置する段階的生成。"""
        fs = st.session_state.get("soft_first_subjects", [])
        with st.spinner(f"教科別段階生成中… Phase 1: {', '.join(fs)}"):
            timetable = {cls: {day: {} for day in DAYS} for cls in classes}
            timetable_sg = {
                sg: {day: {} for day in DAYS}
                for sg in st.session_state["small_group_classes"]
            }
            weekly_periods_sg = {
                sg: {
                    subj: settings.get("weekly", 0)
                    for subj, settings in sg_data.get("subject_settings", {}).items()
                    if settings.get("weekly", 0) > 0
                }
                for sg, sg_data in st.session_state["small_group_classes"].items()
            }
            weekly_remaining, _ = preprocess(
                timetable, timetable_sg,
                st.session_state["manual_timetable"],
                st.session_state["manual_timetable_sg"],
                st.session_state["weekly_periods"],
                weekly_periods_sg
            )
            priority_subjects = st.session_state.get("soft_priority_subjects", [])
            # assignments のコピーを作り合成教科の担当教員を追加
            assignments = {
                cls: dict(v)
                for cls, v in st.session_state["assignments"].items()
            }
            for _cls in get_all_classes():
                for _ci in compute_combined_subjects(
                        st.session_state["weekly_periods"].get(_cls, {})):
                    _cn = _ci["name"]
                    _teachers_combined: list = []
                    for _s in _ci["subjects"]:
                        for _t in assignments.get(_cls, {}).get(_s, []):
                            if _t not in _teachers_combined:
                                _teachers_combined.append(_t)
                    assignments.setdefault(_cls, {})[_cn] = _teachers_combined
            special_rooms   = st.session_state["special_rooms"]
            unavailable     = st.session_state["unavailable"]
            periods_per_day = st.session_state["periods_per_day"]
            sg_classes      = st.session_state["small_group_classes"]

            result = solve_subject_by_subject(
                timetable, timetable_sg,
                weekly_remaining, assignments,
                special_rooms, unavailable,
                periods_per_day, sg_classes,
                priority_subjects,
                first_subjects=fs,
                ort_time_limit=float(time_limit) * 2
            )

            if result == "solved":
                st.session_state["generated_timetable"]    = timetable
                st.session_state["generated_timetable_sg"] = timetable_sg
                pno = st.session_state.get("timetable_pattern_no", 0)
                lbl = "標準パターン" if pno == 1 else f"パターン #{pno}"
                from datetime import datetime as _dt
                history_entry = {
                    "timetable":    timetable,
                    "timetable_sg": timetable_sg,
                    "label":        f"{lbl}（教科別段階生成: {', '.join(fs)} → 残り）",
                    "timestamp":    _dt.now().strftime("%H:%M:%S"),
                }
                st.session_state.setdefault("timetable_history", []).append(history_entry)
                st.success(
                    f"✅ {lbl} の生成が完了しました！（教科別段階生成）\n"
                    f"Phase 1: {', '.join(fs)} → Phase 2: 残りの教科"
                )
                report_r6_violations(timetable)
            elif result == "failed_phase1":
                st.error(
                    f"❌ Phase 1（{', '.join(fs)}）の配置に失敗しました。\n"
                    "選択した教科の担当教員・週コマ数・教室設定を確認してください。"
                )
            else:
                st.error(
                    "❌ Phase 2（残りの教科）の配置に失敗しました。\n"
                    "Phase 1 の配置後に残るスロットが不足している可能性があります。"
                )
        return result

    # ── 通常生成 ─────────────────────────────────────────────
    if generate_btn:
        st.session_state["timetable_pattern_no"] = 1
        result = _run_generation()
        if result != "solved":
            return

    # ── 別パターン生成 ────────────────────────────────────────
    if alt_btn:
        import random as _rnd
        prev_tt = st.session_state.get("generated_timetable", {})
        # 前の解から全配置を収集し、約25%をランダムに禁止する
        all_placements = [
            (cls, day, p, subj)
            for cls, day_map in prev_tt.items()
            for day, period_map in day_map.items()
            for p, subj in period_map.items()
            if subj
        ]
        if all_placements:
            n_forbid = max(1, len(all_placements) // 4)
            forbidden = frozenset(_rnd.sample(all_placements, n_forbid))
        else:
            forbidden = frozenset()
        pno = st.session_state.get("timetable_pattern_no", 1) + 1
        st.session_state["timetable_pattern_no"] = pno
        result = _run_generation(
            forbidden_slots=forbidden,
            ort_seed=pno
        )
        if result != "solved":
            # 別パターンが見つからなかった場合はカウンタを戻す
            st.session_state["timetable_pattern_no"] = pno - 1
            return

    # ── 教科別段階生成ボタン ──────────────────────────────────
    if subject_btn:
        st.session_state["timetable_pattern_no"] = 1
        result = _run_subject_generation()
        if result != "solved":
            return

    # ── 生成済み時間割の表示 ──────────────────────────────────
    history = st.session_state.get("timetable_history", [])

    if not history:
        st.info("上の「時間割を自動生成する」ボタンを押してください")
        return

    st.markdown("---")

    # 履歴が複数ある場合は選択UIを表示
    if len(history) > 1:
        history_labels = [
            f"{h['timestamp']}  {h['label']}"
            for h in reversed(history)
        ]
        selected_label = st.selectbox(
            "📋 閲覧する時間割を選択",
            options=history_labels,
            index=0,
            help="過去に生成した時間割と現在の時間割を切り替えて閲覧できます",
        )
        selected_pos = history_labels.index(selected_label)
        selected_entry = list(reversed(history))[selected_pos]
        timetable    = selected_entry["timetable"]
        timetable_sg = selected_entry["timetable_sg"]
        if selected_pos == 0:
            st.caption("✅ 最新の時間割を表示しています")
        else:
            st.caption(f"🕐 過去の時間割を表示中: {selected_entry['label']}")
    else:
        timetable    = history[-1]["timetable"]
        timetable_sg = history[-1]["timetable_sg"]

    _TAB_LABELS = [
        "📘 クラス別", "📗 少人数学級", "👤 教員別",
        "👥 全教員一覧", "📥 ダウンロード", "✏️ 手動編集"
    ]
    active_tab = st.radio(
        "", _TAB_LABELS, horizontal=True,
        key="tt_active_tab", label_visibility="collapsed"
    )

    st.markdown("---")

    if active_tab == "📘 クラス別":
        selected_cls = st.selectbox(
            "クラスを選択",
            options=get_all_classes(),
            key="disp_cls_select"
        )
        if selected_cls:
            st.subheader(f"📘 {selected_cls} の時間割")
            display_class_timetable(
                selected_cls, timetable,
                st.session_state["manual_timetable"]
            )

    elif active_tab == "📗 少人数学級":
        sg_list = list(st.session_state["small_group_classes"].keys())
        if not sg_list:
            st.info("少人数学級が登録されていません")
        else:
            selected_sg = st.selectbox(
                "少人数学級を選択",
                options=sg_list,
                key="disp_sg_select"
            )
            if selected_sg:
                st.subheader(f"📗 {selected_sg} の時間割")
                display_class_timetable(
                    selected_sg, timetable_sg,
                    st.session_state["manual_timetable_sg"],
                    is_small_group=True
                )

    elif active_tab == "👤 教員別":
        teachers = st.session_state["teachers"]
        if not teachers:
            st.info("教員が登録されていません")
        else:
            selected_t = st.selectbox(
                "教員を選択",
                options=teachers,
                key="disp_teacher_select"
            )
            if selected_t:
                st.subheader(f"👤 {selected_t} の時間割")
                display_teacher_timetable(selected_t, timetable, timetable_sg)

                # ── コマの変更確認 ────────────────────────────────
                st.markdown("---")
                st.markdown("#### 🔍 コマの変更確認")
                chk_col1, chk_col2 = st.columns(2)
                with chk_col1:
                    chk_day = st.selectbox("曜日", DAYS, key="chk_day_sel")
                with chk_col2:
                    chk_max_p = get_periods_for_class(
                        next((c for c in get_all_classes()), "1年1組"), chk_day
                    )
                    chk_period = st.selectbox(
                        "時限",
                        list(range(1, chk_max_p + 1)),
                        format_func=lambda x: f"{x}限",
                        key="chk_period_sel"
                    )

                if st.button("🔍 変更可能か確認する", key="btn_chk_swap"):
                    slot_cls = slot_subj = None
                    for cls, schedule in timetable.items():
                        subj = schedule.get(chk_day, {}).get(chk_period)
                        if subj and selected_t in get_teachers_for_slot(cls, subj):
                            slot_cls, slot_subj = cls, subj
                            break

                    if slot_subj:
                        with st.spinner("全制約をチェック中..."):
                            cands = find_valid_swaps(
                                timetable, timetable_sg, slot_cls, chk_day, chk_period
                            )
                        st.session_state["chk_result"] = {
                            "teacher": selected_t, "day": chk_day, "period": chk_period,
                            "type": "occupied", "cls": slot_cls, "subj": slot_subj,
                            "candidates": cands,
                        }
                    else:
                        with st.spinner("全制約をチェック中..."):
                            cands = []
                            for cls, schedule in timetable.items():
                                for day, periods in schedule.items():
                                    for p, subj in periods.items():
                                        if not subj:
                                            continue
                                        if selected_t not in get_teachers_for_slot(cls, subj):
                                            continue
                                        if day == chk_day and p == chk_period:
                                            continue
                                        if is_valid_swap(
                                            timetable, timetable_sg,
                                            cls, day, p, cls, chk_day, chk_period
                                        ):
                                            cands.append((cls, day, p, subj))
                        st.session_state["chk_result"] = {
                            "teacher": selected_t, "day": chk_day, "period": chk_period,
                            "type": "empty", "candidates": cands,
                        }

                chk_result = st.session_state.get("chk_result")
                if (chk_result
                        and chk_result.get("teacher") == selected_t
                        and chk_result.get("day") == chk_day
                        and chk_result.get("period") == chk_period):

                    cands = chk_result["candidates"]

                    if chk_result["type"] == "occupied":
                        cls   = chk_result["cls"]
                        subj  = chk_result["subj"]
                        st.markdown(
                            f"**対象コマ:** {cls} {chk_day}曜 {chk_period}限 ─ "
                            f"**:blue[{subj}]**（担当: {selected_t}）"
                        )
                        show_cands = [
                            (c, d, p, s) for c, d, p, s in cands
                            if not s or selected_t not in get_teachers_for_slot(c, s)
                        ]
                        if not show_cands:
                            st.error("❌ このコマは交換できません。")
                        else:
                            st.success(f"✅ {len(show_cands)} 件の交換候補があります。")
                            labels = []
                            for c, d, p, s in show_cands:
                                t_list = get_teachers_for_slot(c, s) if s else []
                                t_str  = "・".join(t_list) if t_list else "─"
                                labels.append(
                                    f"【{c}】{d}曜{p}限「{s if s else '空き'}」"
                                    f" 担当:{t_str}"
                                )
                            sel_label = st.selectbox(
                                "交換候補を選択", labels, key="chk_cand_sel"
                            )
                            sel_idx = labels.index(sel_label)
                            tgt_c, tgt_d, tgt_p, tgt_s = show_cands[sel_idx]
                            if st.button(
                                "✏️ この候補で手動編集タブへ", key="btn_goto_edit_from_chk"
                            ):
                                st.session_state["tt_active_tab"]    = "✏️ 手動編集"
                                st.session_state["edit_select_mode"] = "クラスから選ぶ"
                                st.session_state["edit_cls_select"]  = cls
                                st.session_state["edit_cls_last"]    = cls
                                st.session_state["edit_cls_slot_a"]  = (chk_day, chk_period, cls, subj)
                                st.session_state["edit_cls_slot_b"]  = (tgt_d, tgt_p, tgt_c, tgt_s)
                                st.session_state["edit_cls_cands"]   = None
                                st.rerun()

                    else:  # empty
                        st.markdown(
                            f"**対象コマ:** {selected_t} {chk_day}曜 {chk_period}限 ─ "
                            f"**空きコマ**"
                        )
                        if not cands:
                            st.error("❌ このコマに移動できる授業がありません。")
                        else:
                            st.success(f"✅ {len(cands)} 件の授業をこのコマに移動できます。")
                            labels = []
                            for c, d, p, s in cands:
                                labels.append(
                                    f"【{c}】{d}曜{p}限「{s}」→ {chk_day}曜{chk_period}限に移動"
                                )
                            sel_label = st.selectbox(
                                "移動候補を選択", labels, key="chk_cand_sel"
                            )
                            sel_idx = labels.index(sel_label)
                            src_c, src_d, src_p, src_s = cands[sel_idx]
                            if st.button(
                                "✏️ この候補で手動編集タブへ", key="btn_goto_edit_from_chk"
                            ):
                                st.session_state["tt_active_tab"]    = "✏️ 手動編集"
                                st.session_state["edit_select_mode"] = "クラスから選ぶ"
                                st.session_state["edit_cls_select"]  = src_c
                                st.session_state["edit_cls_last"]    = src_c
                                st.session_state["edit_cls_slot_a"]  = (src_d, src_p, src_c, src_s)
                                st.session_state["edit_cls_slot_b"]  = (chk_day, chk_period, src_c, None)
                                st.session_state["edit_cls_cands"]   = None
                                st.rerun()

                # ── 2コマの入れ替え確認 ────────────────────────────────
                st.markdown("---")
                st.markdown("#### 🔄 同一教員の2コマ入れ替え確認")
                st.caption(
                    "コマをクリックして A・B を選択 → 「確認する」で入れ替え可否と他教員への影響を表示します。"
                )

                if st.session_state.get("swap2_last_teacher") != selected_t:
                    st.session_state["swap2_slot_a"] = None
                    st.session_state["swap2_slot_b"] = None
                    st.session_state["swap2_result"] = None
                    st.session_state["swap2_last_teacher"] = selected_t

                _slot_a = st.session_state.get("swap2_slot_a")
                _slot_b = st.session_state.get("swap2_slot_b")

                _disp_col1, _disp_col2 = st.columns(2)
                with _disp_col1:
                    if _slot_a:
                        _la = (f"【{_slot_a[2]}】{_slot_a[0]}曜{_slot_a[1]}限「{_slot_a[3]}」"
                               if _slot_a[3] else f"{_slot_a[0]}曜{_slot_a[1]}限（空き）")
                        st.info(f"🅐 コマA: {_la}")
                    else:
                        st.info("🅐 コマA: 未選択")
                with _disp_col2:
                    if _slot_b:
                        _lb = (f"【{_slot_b[2]}】{_slot_b[0]}曜{_slot_b[1]}限「{_slot_b[3]}」"
                               if _slot_b[3] else f"{_slot_b[0]}曜{_slot_b[1]}限（空き）")
                        st.info(f"🅑 コマB: {_lb}")
                    else:
                        st.info("🅑 コマB: 未選択")

                _ppd  = st.session_state["periods_per_day"]
                _maxp = max(_ppd.values()) if _ppd else 6

                _hdr_cols = st.columns([0.6] + [1] * len(DAYS))
                with _hdr_cols[0]:
                    st.markdown("**時限**")
                for _hi, _hd in enumerate(DAYS):
                    with _hdr_cols[_hi + 1]:
                        st.markdown(f"**{_hd}曜**")

                for _p in range(1, _maxp + 1):
                    _row = st.columns([0.6] + [1] * len(DAYS))
                    with _row[0]:
                        st.markdown(f"**{_p}限**")
                    for _di, _d in enumerate(DAYS):
                        with _row[_di + 1]:
                            if _p > _ppd.get(_d, 6):
                                st.write("─")
                                continue
                            _sc = _ss = None
                            for _c, _sched in timetable.items():
                                _s = _sched.get(_d, {}).get(_p)
                                if _s and selected_t in get_teachers_for_slot(_c, _s):
                                    _sc, _ss = _c, _s
                                    break
                            _is_a = bool(_slot_a and _slot_a[0] == _d and _slot_a[1] == _p)
                            _is_b = bool(_slot_b and _slot_b[0] == _d and _slot_b[1] == _p)
                            _mark  = "🅐 " if _is_a else ("🅑 " if _is_b else "")
                            _label = f"{_mark}{_ss}" if _ss else f"{_mark}空き"
                            if _sc:
                                _label += f"\n({_sc})"
                            if st.button(
                                _label,
                                key=f"s2_cell_{_d}_{_p}",
                                type="primary" if (_is_a or _is_b) else "secondary",
                                use_container_width=True,
                            ):
                                if _is_a:
                                    st.session_state["swap2_slot_a"] = None
                                elif _is_b:
                                    st.session_state["swap2_slot_b"] = None
                                elif not _slot_a:
                                    st.session_state["swap2_slot_a"] = (_d, _p, _sc, _ss)
                                elif not _slot_b:
                                    st.session_state["swap2_slot_b"] = (_d, _p, _sc, _ss)
                                else:
                                    st.session_state["swap2_slot_a"] = (_d, _p, _sc, _ss)
                                    st.session_state["swap2_slot_b"] = None
                                st.session_state["swap2_result"] = None
                                st.rerun()

                if _slot_a and _slot_b:
                    if st.button("🔄 入れ替え可能か確認する", key="btn_chk_swap2"):
                        st.session_state["swap2_result"] = {
                            "teacher":  selected_t,
                            "day_a":    _slot_a[0], "period_a": _slot_a[1],
                            "cls_a":    _slot_a[2], "subj_a":   _slot_a[3],
                            "day_b":    _slot_b[0], "period_b": _slot_b[1],
                            "cls_b":    _slot_b[2], "subj_b":   _slot_b[3],
                        }
                else:
                    st.caption("↑ コマを2つクリックして選択してください")

                swap2_result = st.session_state.get("swap2_result")
                _slot_a = st.session_state.get("swap2_slot_a")
                _slot_b = st.session_state.get("swap2_slot_b")
                if (swap2_result
                        and swap2_result.get("teacher") == selected_t
                        and _slot_a and _slot_b
                        and swap2_result.get("day_a") == _slot_a[0]
                        and swap2_result.get("period_a") == _slot_a[1]
                        and swap2_result.get("day_b") == _slot_b[0]
                        and swap2_result.get("period_b") == _slot_b[1]):

                    r2 = swap2_result
                    s2_day_a    = r2["day_a"];    s2_period_a = r2["period_a"]
                    s2_day_b    = r2["day_b"];    s2_period_b = r2["period_b"]
                    cls_a2  = r2["cls_a"];  subj_a2 = r2["subj_a"]
                    cls_b2  = r2["cls_b"];  subj_b2 = r2["subj_b"]

                    label_a2 = (
                        f"【{cls_a2}】{s2_day_a}曜{s2_period_a}限「{subj_a2}」"
                        if subj_a2 else f"{s2_day_a}曜{s2_period_a}限（空きコマ）"
                    )
                    label_b2 = (
                        f"【{cls_b2}】{s2_day_b}曜{s2_period_b}限「{subj_b2}」"
                        if subj_b2 else f"{s2_day_b}曜{s2_period_b}限（空きコマ）"
                    )
                    st.markdown(f"**コマ A:** {label_a2}")
                    st.markdown(f"**コマ B:** {label_b2}")

                    if s2_day_a == s2_day_b and s2_period_a == s2_period_b:
                        st.error("❌ 同じコマが選択されています。")
                    elif not subj_a2 and not subj_b2:
                        st.warning("⚠️ 両方とも空きコマです。入れ替えの意味がありません。")
                    else:
                        if subj_a2 and subj_b2:
                            eff_cls_a2, eff_cls_b2 = cls_a2, cls_b2
                            _vp2 = (cls_a2, s2_day_a, s2_period_a,
                                    cls_b2, s2_day_b, s2_period_b)
                            ok2 = is_valid_swap(timetable, timetable_sg, *_vp2)
                        elif subj_a2:
                            eff_cls_a2, eff_cls_b2 = cls_a2, cls_a2
                            _vp2 = (cls_a2, s2_day_a, s2_period_a,
                                    cls_a2, s2_day_b, s2_period_b)
                            ok2 = is_valid_swap(timetable, timetable_sg, *_vp2)
                        else:
                            eff_cls_a2, eff_cls_b2 = cls_b2, cls_b2
                            _vp2 = (cls_b2, s2_day_b, s2_period_b,
                                    cls_b2, s2_day_a, s2_period_a)
                            ok2 = is_valid_swap(timetable, timetable_sg, *_vp2)

                        if not ok2:
                            _viols2 = get_swap_violations(
                                timetable, timetable_sg, *_vp2
                            )
                            st.error("❌ このコマは入れ替えできません。")
                            for _v in _viols2:
                                st.write(f"　・{_v}")
                        else:
                            st.success("✅ このコマは入れ替えできます。")

                            impacts2 = get_swap_impact(
                                timetable, selected_t,
                                eff_cls_a2, s2_day_a, s2_period_a,
                                eff_cls_b2, s2_day_b, s2_period_b,
                            )
                            if not impacts2:
                                st.info("他の教員への影響はありません。")
                            else:
                                st.markdown("**他の教員への影響:**")
                                for imp in impacts2:
                                    st.write(
                                        f"・**{imp['teacher']}**: "
                                        f"【{imp['from_cls']}】"
                                        f"{imp['from_day']}曜{imp['from_period']}限"
                                        f"「{imp['from_subj']}」"
                                        f" → "
                                        f"【{imp['to_cls']}】"
                                        f"{imp['to_day']}曜{imp['to_period']}限"
                                    )

                            if st.button("✅ 入れ替えを実行する", key="btn_exec_swap2"):
                                snapshot2 = {
                                    c: {d: dict(prs) for d, prs in s.items()}
                                    for c, s in timetable.items()
                                }
                                st.session_state["swap_undo_stack"].append({
                                    "snapshot": snapshot2,
                                    "description": (
                                        f"{selected_t}: {label_a2} ↔ {label_b2}"
                                    ),
                                })
                                if subj_a2 and subj_b2:
                                    timetable[cls_a2][s2_day_a][s2_period_a] = subj_b2
                                    timetable[cls_b2][s2_day_b][s2_period_b] = subj_a2
                                elif subj_a2:
                                    _exist_b = timetable.get(cls_a2, {}).get(s2_day_b, {}).get(s2_period_b)
                                    timetable[cls_a2][s2_day_a][s2_period_a] = _exist_b
                                    timetable[cls_a2].setdefault(s2_day_b, {})[s2_period_b] = subj_a2
                                else:
                                    _exist_a = timetable.get(cls_b2, {}).get(s2_day_a, {}).get(s2_period_a)
                                    timetable[cls_b2][s2_day_b][s2_period_b] = _exist_a
                                    timetable[cls_b2].setdefault(s2_day_a, {})[s2_period_a] = subj_b2
                                st.session_state["swap2_result"] = None
                                st.rerun()

    elif active_tab == "👥 全教員一覧":
        teachers = st.session_state["teachers"]
        if not teachers:
            st.info("教員が登録されていません")
        else:
            st.subheader("👥 全教員の時間割一覧")
            display_all_teachers_timetable(timetable, timetable_sg)

    elif active_tab == "📥 ダウンロード":
        st.markdown("#### 📥 時間割をExcelでダウンロード")
        fname = f"時間割_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            label="📥 時間割Excelをダウンロード",
            data=export_timetable_to_excel(
                timetable, timetable_sg,
                st.session_state["manual_timetable"]
            ),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.caption("クラス別・少人数学級・教員別シートが含まれます")

    elif active_tab == "✏️ 手動編集":
        st.subheader("✏️ コマの手動交換")
        st.info(
            "交換したいコマを選択し「候補を検索」を押すと、"
            "全ての制約を満たす交換可能なコマが一覧表示されます。"
        )

        select_mode = st.radio(
            "コマの選択方法",
            ["クラスから選ぶ", "教員から選ぶ"],
            horizontal=True,
            key="edit_select_mode"
        )

        if select_mode == "クラスから選ぶ":
            edit_cls = st.selectbox(
                "クラス", get_all_classes(), key="edit_cls_select"
            )

            if st.session_state.get("edit_cls_last") != edit_cls:
                st.session_state["edit_cls_slot_a"]  = None
                st.session_state["edit_cls_slot_b"]  = None
                st.session_state["edit_cls_cands"]   = None
                st.session_state["edit_cls_last"]    = edit_cls

            _ec_sa = st.session_state.get("edit_cls_slot_a")
            _ec_sb = st.session_state.get("edit_cls_slot_b")

            _ec_cand_set = set()
            if _ec_sa and not _ec_sb:
                _ec_sa_day, _ec_sa_per, _ec_sa_cls, _ec_sa_subj = _ec_sa
                if st.session_state.get("edit_cls_cands") is None:
                    with st.spinner("候補を検索中..."):
                        if _ec_sa_subj:
                            _ec_all_cands = find_valid_swaps(
                                timetable, timetable_sg, _ec_sa_cls, _ec_sa_day, _ec_sa_per
                            )
                        else:
                            _ec_all_cands = []
                            for _d, _prs in timetable.get(edit_cls, {}).items():
                                for _p, _s in _prs.items():
                                    if not _s:
                                        continue
                                    if _d == _ec_sa_day and _p == _ec_sa_per:
                                        continue
                                    if is_valid_swap(
                                        timetable, timetable_sg,
                                        edit_cls, _d, _p,
                                        edit_cls, _ec_sa_day, _ec_sa_per
                                    ):
                                        _ec_all_cands.append((edit_cls, _d, _p, _s))
                    st.session_state["edit_cls_cands"] = _ec_all_cands
                _ec_all_cands = st.session_state.get("edit_cls_cands", [])
                _ec_cand_set  = {(d, p) for c, d, p, s in _ec_all_cands if c == edit_cls}

            _ec_col1, _ec_col2 = st.columns(2)
            with _ec_col1:
                if _ec_sa:
                    _ec_sa_lbl = (f"{_ec_sa[0]}曜{_ec_sa[1]}限「{_ec_sa[3]}」"
                                  if _ec_sa[3] else f"{_ec_sa[0]}曜{_ec_sa[1]}限（空き）")
                    _ec_hint = "（ここへ移動できる授業）" if not _ec_sa[3] else "（交換可能）"
                    if _ec_cand_set:
                        st.info(f"🅐 コマA: {_ec_sa_lbl}　🟢 {len(_ec_cand_set)} 件{_ec_hint}")
                    else:
                        st.info(f"🅐 コマA: {_ec_sa_lbl}")
                else:
                    st.info("🅐 コマA: 未選択")
            with _ec_col2:
                if _ec_sb:
                    _ec_sb_lbl = (f"{_ec_sb[0]}曜{_ec_sb[1]}限「{_ec_sb[3]}」"
                                  if _ec_sb[3] else f"{_ec_sb[0]}曜{_ec_sb[1]}限（空き）")
                    st.info(f"🅑 コマB: {_ec_sb_lbl}")
                else:
                    _b_hint2 = "🟢のコマを選択" if _ec_cand_set else "未選択"
                    st.info(f"🅑 コマB: {_b_hint2}")

            st.markdown("""
<style>
div[data-testid="stMarkdown"]:has(.ec-swap-candidate)
  + div[data-testid="stButton"] button {
    background-color: #c8f7c5 !important;
    border: 2px solid #28a745 !important;
    color: #155724 !important;
}
</style>
""", unsafe_allow_html=True)

            _ec_color_constraints = get_color_constraints(st.session_state["small_group_classes"])
            _ppd_c  = st.session_state["periods_per_day"]
            _maxp_c = max(_ppd_c.values()) if _ppd_c else 6

            _hdr_c = st.columns([0.6] + [1] * len(DAYS))
            with _hdr_c[0]:
                st.markdown("**時限**")
            for _hi_c, _hd_c in enumerate(DAYS):
                with _hdr_c[_hi_c + 1]:
                    st.markdown(f"**{_hd_c}曜**")

            for _p_c in range(1, _maxp_c + 1):
                _row_c = st.columns([0.6] + [1] * len(DAYS))
                with _row_c[0]:
                    st.markdown(f"**{_p_c}限**")
                for _di_c, _d_c in enumerate(DAYS):
                    with _row_c[_di_c + 1]:
                        if _p_c > get_periods_for_class(edit_cls, _d_c):
                            st.write("─")
                            continue
                        _ss_c    = timetable.get(edit_cls, {}).get(_d_c, {}).get(_p_c)
                        _is_a_c  = bool(_ec_sa and _ec_sa[0] == _d_c and _ec_sa[1] == _p_c)
                        _is_b_c  = bool(_ec_sb and _ec_sb[0] == _d_c and _ec_sb[1] == _p_c)
                        _is_cc   = ((_d_c, _p_c) in _ec_cand_set and not _is_a_c and not _is_b_c)
                        if _is_cc:
                            st.markdown(
                                '<div class="ec-swap-candidate"></div>',
                                unsafe_allow_html=True
                            )
                        _mark_c = "🅐 " if _is_a_c else ("🅑 " if _is_b_c else ("🔄 " if _is_cc else ""))
                        _slot_info_c = _ec_color_constraints.get(edit_cls, {}).get(_d_c, {}).get(_p_c)
                        _red_prefix_c = "🔴" if (_slot_info_c and _slot_info_c.get("color") == "red") else ""
                        _lbl_c  = f"{_mark_c}{_red_prefix_c}{_ss_c}" if _ss_c else f"{_mark_c}{_red_prefix_c}空き"
                        if st.button(
                            _lbl_c,
                            key=f"ec_cell_{_d_c}_{_p_c}",
                            type="primary" if (_is_a_c or _is_b_c) else "secondary",
                            use_container_width=True,
                        ):
                            if _is_a_c:
                                st.session_state["edit_cls_slot_a"] = None
                                st.session_state["edit_cls_cands"]  = None
                            elif _is_b_c:
                                st.session_state["edit_cls_slot_b"] = None
                            elif not _ec_sa:
                                st.session_state["edit_cls_slot_a"] = (_d_c, _p_c, edit_cls, _ss_c)
                                st.session_state["edit_cls_cands"]  = None
                            elif not _ec_sb:
                                st.session_state["edit_cls_slot_b"] = (_d_c, _p_c, edit_cls, _ss_c)
                            else:
                                st.session_state["edit_cls_slot_a"] = (_d_c, _p_c, edit_cls, _ss_c)
                                st.session_state["edit_cls_slot_b"] = None
                                st.session_state["edit_cls_cands"]  = None
                            st.rerun()

            _ec_sa = st.session_state.get("edit_cls_slot_a")
            _ec_sb = st.session_state.get("edit_cls_slot_b")
            if not (_ec_sa and _ec_sb):
                st.caption("↑ コマを2つクリックして選択してください")
            else:
                _ec_da, _ec_pa, _ec_ca, _ec_ss_a = _ec_sa
                _ec_db, _ec_pb, _ec_cb, _ec_ss_b = _ec_sb

                _ec_la = (f"【{_ec_ca}】{_ec_da}曜{_ec_pa}限「{_ec_ss_a}」"
                          if _ec_ss_a else f"{_ec_da}曜{_ec_pa}限（空きコマ）")
                _ec_lb = (f"【{_ec_cb}】{_ec_db}曜{_ec_pb}限「{_ec_ss_b}」"
                          if _ec_ss_b else f"{_ec_db}曜{_ec_pb}限（空きコマ）")

                st.markdown("---")
                if _ec_da == _ec_db and _ec_pa == _ec_pb and _ec_ca == _ec_cb:
                    st.error("❌ 同じコマが選択されています。")
                elif not _ec_ss_a and not _ec_ss_b:
                    st.warning("⚠️ 両方とも空きコマです。入れ替えの意味がありません。")
                else:
                    if _ec_ss_a and _ec_ss_b:
                        _ec_vp = (_ec_ca, _ec_da, _ec_pa, _ec_cb, _ec_db, _ec_pb)
                        _ec_eff_ca, _ec_eff_cb = _ec_ca, _ec_cb
                    elif _ec_ss_a:
                        _ec_vp = (_ec_ca, _ec_da, _ec_pa, _ec_ca, _ec_db, _ec_pb)
                        _ec_eff_ca = _ec_eff_cb = _ec_ca
                    else:
                        _ec_vp = (_ec_cb, _ec_db, _ec_pb, _ec_cb, _ec_da, _ec_pa)
                        _ec_eff_ca = _ec_eff_cb = _ec_cb

                    _ec_ok = is_valid_swap(timetable, timetable_sg, *_ec_vp)

                    if not _ec_ok:
                        _ec_viols = get_swap_violations(timetable, timetable_sg, *_ec_vp)
                        st.error("❌ このコマは入れ替えできません。")
                        for _v in _ec_viols:
                            st.write(f"　・{_v}")
                    else:
                        _ec_impacts = get_swap_impact(
                            timetable, None,
                            _ec_eff_ca, _ec_da, _ec_pa,
                            _ec_eff_cb, _ec_db, _ec_pb,
                        )
                        if not _ec_impacts:
                            st.info("他の教員への影響はありません。")
                        else:
                            st.markdown("**他の教員への影響:**")
                            for _imp in _ec_impacts:
                                st.write(
                                    f"・**{_imp['teacher']}**: "
                                    f"【{_imp['from_cls']}】"
                                    f"{_imp['from_day']}曜{_imp['from_period']}限"
                                    f"「{_imp['from_subj']}」"
                                    f" → "
                                    f"【{_imp['to_cls']}】"
                                    f"{_imp['to_day']}曜{_imp['to_period']}限"
                                )

                        if st.button("✅ 入れ替えを実行する", type="primary",
                                     key="btn_exec_edit_cls"):
                            _ec_snap = {
                                c: {d: dict(prs) for d, prs in s.items()}
                                for c, s in timetable.items()
                            }
                            st.session_state["swap_undo_stack"].append({
                                "snapshot": _ec_snap,
                                "description": f"{_ec_la} ↔ {_ec_lb}",
                            })
                            if _ec_ss_a and _ec_ss_b:
                                timetable[_ec_ca][_ec_da][_ec_pa] = _ec_ss_b
                                timetable[_ec_cb][_ec_db][_ec_pb] = _ec_ss_a
                            elif _ec_ss_a:
                                _ec_ex = timetable.get(_ec_ca, {}).get(_ec_db, {}).get(_ec_pb)
                                timetable[_ec_ca][_ec_da][_ec_pa] = _ec_ex
                                timetable[_ec_ca].setdefault(_ec_db, {})[_ec_pb] = _ec_ss_a
                            else:
                                _ec_ex = timetable.get(_ec_cb, {}).get(_ec_da, {}).get(_ec_pa)
                                timetable[_ec_cb][_ec_db][_ec_pb] = _ec_ex
                                timetable[_ec_cb].setdefault(_ec_da, {})[_ec_pa] = _ec_ss_b
                            st.session_state["generated_timetable"] = timetable
                            st.session_state["edit_cls_slot_a"] = None
                            st.session_state["edit_cls_slot_b"] = None
                            st.session_state["edit_cls_cands"]  = None
                            st.rerun()

        else:  # 教員から選ぶ
            teachers_list = st.session_state["teachers"]
            if not teachers_list:
                st.info("教員が登録されていません")
            else:
                edit_teacher = st.selectbox(
                    "教員", teachers_list, key="edit_teacher_select"
                )
                st.caption(
                    "コマをクリックして A・B を選択し、「確認する」で入れ替え可否と"
                    "他教員への影響を確認できます。"
                )

                if st.session_state.get("edit2_last_teacher") != edit_teacher:
                    st.session_state["edit2_slot_a"] = None
                    st.session_state["edit2_slot_b"] = None
                    st.session_state["edit2_result"] = None
                    st.session_state["edit2_cands"]  = None
                    st.session_state["edit2_last_teacher"] = edit_teacher

                _e2_slot_a = st.session_state.get("edit2_slot_a")
                _e2_slot_b = st.session_state.get("edit2_slot_b")

                _e2_cand_set = set()
                if _e2_slot_a and not _e2_slot_b:
                    if st.session_state.get("edit2_cands") is None:
                        with st.spinner("候補を検索中..."):
                            if _e2_slot_a[3]:
                                _a_cls = _e2_slot_a[2]
                                if _a_cls:
                                    _e2_all_cands = find_valid_swaps(
                                        timetable, timetable_sg,
                                        _a_cls, _e2_slot_a[0], _e2_slot_a[1]
                                    )
                                else:
                                    _e2_all_cands = []
                            else:
                                _e2_all_cands = []
                                for _cls, _sched in timetable.items():
                                    for _d, _prs in _sched.items():
                                        for _p, _s in _prs.items():
                                            if not _s:
                                                continue
                                            if edit_teacher not in get_teachers_for_slot(_cls, _s):
                                                continue
                                            if _d == _e2_slot_a[0] and _p == _e2_slot_a[1]:
                                                continue
                                            if is_valid_swap(
                                                timetable, timetable_sg,
                                                _cls, _d, _p,
                                                _cls, _e2_slot_a[0], _e2_slot_a[1]
                                            ):
                                                _e2_all_cands.append((_cls, _d, _p, _s))
                        st.session_state["edit2_cands"] = _e2_all_cands
                    _e2_cand_set = {(d, p) for _, d, p, _ in
                                    st.session_state.get("edit2_cands", [])}

                _e2_c1, _e2_c2 = st.columns(2)
                with _e2_c1:
                    if _e2_slot_a:
                        _la = (f"【{_e2_slot_a[2]}】{_e2_slot_a[0]}曜{_e2_slot_a[1]}限"
                               f"「{_e2_slot_a[3]}」" if _e2_slot_a[3]
                               else f"{_e2_slot_a[0]}曜{_e2_slot_a[1]}限（空き）")
                        _hint = ("（ここへ移動できる授業）" if not _e2_slot_a[3]
                                 else "（交換可能）")
                        if _e2_cand_set:
                            st.info(f"🅐 コマA: {_la}　🟢 {len(_e2_cand_set)} 件{_hint}")
                        else:
                            st.info(f"🅐 コマA: {_la}")
                    else:
                        st.info("🅐 コマA: 未選択")
                with _e2_c2:
                    if _e2_slot_b:
                        _lb = (f"【{_e2_slot_b[2]}】{_e2_slot_b[0]}曜{_e2_slot_b[1]}限"
                               f"「{_e2_slot_b[3]}」" if _e2_slot_b[3]
                               else f"{_e2_slot_b[0]}曜{_e2_slot_b[1]}限（空き）")
                        st.info(f"🅑 コマB: {_lb}")
                    else:
                        _b_hint = "🟢のコマを選択" if _e2_cand_set else "未選択"
                        st.info(f"🅑 コマB: {_b_hint}")

                st.markdown("""
<style>
div[data-testid="stMarkdown"]:has(.et-swap-candidate)
  + div[data-testid="stButton"] button {
    background-color: #c8f7c5 !important;
    border: 2px solid #28a745 !important;
    color: #155724 !important;
}
</style>
""", unsafe_allow_html=True)

                _et_color_constraints = get_color_constraints(st.session_state["small_group_classes"])
                _ppd_e  = st.session_state["periods_per_day"]
                _maxp_e = max(_ppd_e.values()) if _ppd_e else 6

                _hdr_e = st.columns([0.6] + [1] * len(DAYS))
                with _hdr_e[0]:
                    st.markdown("**時限**")
                for _hi_e, _hd_e in enumerate(DAYS):
                    with _hdr_e[_hi_e + 1]:
                        st.markdown(f"**{_hd_e}曜**")

                for _p_e in range(1, _maxp_e + 1):
                    _row_e = st.columns([0.6] + [1] * len(DAYS))
                    with _row_e[0]:
                        st.markdown(f"**{_p_e}限**")
                    for _di_e, _d_e in enumerate(DAYS):
                        with _row_e[_di_e + 1]:
                            if _p_e > _ppd_e.get(_d_e, 6):
                                st.write("─")
                                continue
                            _sc_e = _ss_e = None
                            for _c_e, _sched_e in timetable.items():
                                _s_e = _sched_e.get(_d_e, {}).get(_p_e)
                                if _s_e and edit_teacher in get_teachers_for_slot(_c_e, _s_e):
                                    _sc_e, _ss_e = _c_e, _s_e
                                    break
                            _is_a_e  = bool(_e2_slot_a and _e2_slot_a[0] == _d_e and _e2_slot_a[1] == _p_e)
                            _is_b_e  = bool(_e2_slot_b and _e2_slot_b[0] == _d_e and _e2_slot_b[1] == _p_e)
                            _is_ce   = ((_d_e, _p_e) in _e2_cand_set
                                        and not _is_a_e and not _is_b_e)
                            if _is_ce:
                                st.markdown(
                                    '<div class="et-swap-candidate"></div>',
                                    unsafe_allow_html=True
                                )
                            _mark_e = "🅐 " if _is_a_e else ("🅑 " if _is_b_e else ("🔄 " if _is_ce else ""))
                            _slot_info_e = _et_color_constraints.get(_sc_e, {}).get(_d_e, {}).get(_p_e) if _sc_e else None
                            _red_prefix_e = "🔴" if (_slot_info_e and _slot_info_e.get("color") == "red") else ""
                            _lbl_e  = f"{_mark_e}{_red_prefix_e}{_ss_e}" if _ss_e else f"{_mark_e}{_red_prefix_e}空き"
                            if _sc_e:
                                _lbl_e += f"\n({_sc_e})"
                            if st.button(
                                _lbl_e,
                                key=f"et_cell_{_d_e}_{_p_e}",
                                type="primary" if (_is_a_e or _is_b_e) else "secondary",
                                use_container_width=True,
                            ):
                                if _is_a_e:
                                    st.session_state["edit2_slot_a"] = None
                                    st.session_state["edit2_cands"]  = None
                                elif _is_b_e:
                                    st.session_state["edit2_slot_b"] = None
                                elif not _e2_slot_a:
                                    st.session_state["edit2_slot_a"] = (_d_e, _p_e, _sc_e, _ss_e)
                                    st.session_state["edit2_cands"]  = None
                                elif not _e2_slot_b:
                                    st.session_state["edit2_slot_b"] = (_d_e, _p_e, _sc_e, _ss_e)
                                else:
                                    st.session_state["edit2_slot_a"] = (_d_e, _p_e, _sc_e, _ss_e)
                                    st.session_state["edit2_slot_b"] = None
                                    st.session_state["edit2_cands"]  = None
                                st.session_state["edit2_result"] = None
                                st.rerun()

                _e2_slot_a = st.session_state.get("edit2_slot_a")
                _e2_slot_b = st.session_state.get("edit2_slot_b")
                if not (_e2_slot_a and _e2_slot_b):
                    st.caption("↑ コマを2つクリックして選択してください")
                else:
                    _e2_da, _e2_pa = _e2_slot_a[0], _e2_slot_a[1]
                    _e2_db, _e2_pb = _e2_slot_b[0], _e2_slot_b[1]
                    _e2_ca, _e2_sa = _e2_slot_a[2], _e2_slot_a[3]
                    _e2_cb, _e2_sb = _e2_slot_b[2], _e2_slot_b[3]

                    _e2_la = (f"【{_e2_ca}】{_e2_da}曜{_e2_pa}限「{_e2_sa}」"
                              if _e2_sa else f"{_e2_da}曜{_e2_pa}限（空きコマ）")
                    _e2_lb = (f"【{_e2_cb}】{_e2_db}曜{_e2_pb}限「{_e2_sb}」"
                              if _e2_sb else f"{_e2_db}曜{_e2_pb}限（空きコマ）")

                    st.markdown("---")
                    if _e2_da == _e2_db and _e2_pa == _e2_pb:
                        st.error("❌ 同じコマが選択されています。")
                    elif not _e2_sa and not _e2_sb:
                        st.warning("⚠️ 両方とも空きコマです。入れ替えの意味がありません。")
                    else:
                        if _e2_sa and _e2_sb:
                            _eff_ca, _eff_cb = _e2_ca, _e2_cb
                            _vp_e2 = (_e2_ca, _e2_da, _e2_pa, _e2_cb, _e2_db, _e2_pb)
                        elif _e2_sa:
                            _eff_ca, _eff_cb = _e2_ca, _e2_ca
                            _vp_e2 = (_e2_ca, _e2_da, _e2_pa, _e2_ca, _e2_db, _e2_pb)
                        else:
                            _eff_ca, _eff_cb = _e2_cb, _e2_cb
                            _vp_e2 = (_e2_cb, _e2_db, _e2_pb, _e2_cb, _e2_da, _e2_pa)

                        _ok_e2 = is_valid_swap(timetable, timetable_sg, *_vp_e2)

                        if not _ok_e2:
                            _viols_e2 = get_swap_violations(timetable, timetable_sg, *_vp_e2)
                            st.error("❌ このコマは入れ替えできません。")
                            for _v in _viols_e2:
                                st.write(f"　・{_v}")
                        else:
                            _impacts_e2 = get_swap_impact(
                                timetable, edit_teacher,
                                _eff_ca, _e2_da, _e2_pa,
                                _eff_cb, _e2_db, _e2_pb,
                            )
                            if not _impacts_e2:
                                st.info("他の教員への影響はありません。")
                            else:
                                st.markdown("**他の教員への影響:**")
                                for _imp in _impacts_e2:
                                    st.write(
                                        f"・**{_imp['teacher']}**: "
                                        f"【{_imp['from_cls']}】"
                                        f"{_imp['from_day']}曜{_imp['from_period']}限"
                                        f"「{_imp['from_subj']}」"
                                        f" → "
                                        f"【{_imp['to_cls']}】"
                                        f"{_imp['to_day']}曜{_imp['to_period']}限"
                                    )

                            if st.button("✅ 入れ替えを実行する", type="primary",
                                         key="btn_exec_edit2"):
                                _snap_e2 = {
                                    c: {d: dict(prs) for d, prs in s.items()}
                                    for c, s in timetable.items()
                                }
                                st.session_state["swap_undo_stack"].append({
                                    "snapshot": _snap_e2,
                                    "description": f"{edit_teacher}: {_e2_la} ↔ {_e2_lb}",
                                })
                                if _e2_sa and _e2_sb:
                                    timetable[_e2_ca][_e2_da][_e2_pa] = _e2_sb
                                    timetable[_e2_cb][_e2_db][_e2_pb] = _e2_sa
                                elif _e2_sa:
                                    _ex = timetable.get(_e2_ca, {}).get(_e2_db, {}).get(_e2_pb)
                                    timetable[_e2_ca][_e2_da][_e2_pa] = _ex
                                    timetable[_e2_ca].setdefault(_e2_db, {})[_e2_pb] = _e2_sa
                                else:
                                    _ex = timetable.get(_e2_cb, {}).get(_e2_da, {}).get(_e2_pa)
                                    timetable[_e2_cb][_e2_db][_e2_pb] = _ex
                                    timetable[_e2_cb].setdefault(_e2_da, {})[_e2_pa] = _e2_sb
                                st.session_state["generated_timetable"] = timetable
                                st.session_state["edit2_slot_a"] = None
                                st.session_state["edit2_slot_b"] = None
                                st.session_state["edit2_cands"]  = None
                                st.rerun()

        # ── 元に戻す ──────────────────────────────────────────
        undo_stack = st.session_state.get("swap_undo_stack", [])
        if undo_stack:
            st.markdown("---")
            last = undo_stack[-1]
            st.caption(f"直前の交換: {last['description']}")
            if st.button("↩️ 元に戻す", key="btn_undo_swap"):
                for cls, sched in last["snapshot"].items():
                    for day, periods in sched.items():
                        timetable[cls][day] = dict(periods)
                st.session_state["generated_timetable"] = timetable
                undo_stack.pop()
                st.session_state.pop("swap_candidates", None)
                st.session_state.pop("swap_src", None)
                st.rerun()


# ============================================================
# ⑤ main()（完成版）
# ============================================================

def main():
    st.set_page_config(
        page_title="小学校時間割自動生成",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    init_session_state()
    render_sidebar()

    step = st.session_state["current_step"]
    if   step == 1: render_step1()
    elif step == 2: render_step2()
    elif step == 3: render_step3()
    elif step == 4: render_step4()
    elif step == 5: render_step5()
    elif step == 6: render_step6()
    elif step == 7: render_step7()
    elif step == 8: render_step8()
    elif step == 9: render_step9()
    elif step == 10: render_step10()
    elif step == 0: render_generate()


if __name__ == "__main__":
    main()

